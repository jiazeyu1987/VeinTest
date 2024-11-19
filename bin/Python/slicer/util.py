from __future__ import print_function

import time

import logging

import os
import struct
import numpy as np
import qt
from threading import Lock
from datetime import datetime
import random
import string
import slicer

current_main_module_name = ""
hardware_state={}
username = None
user_id = 0
user_role = 0
db_conn = None
backup_db_conn = None
patient_name = "undefined"
patient_uid = "undefined"
volume_image_map = {}
global_data_map = {}
check_des_list = ['电源状态', '功放状态', '水处理状态', '主机状态', '机器人状态']
check_state_list = [0, 0, 0, 0, 0]
save_screeen_capture_list = []
error_des = ""
viewNodeIDs = ["vtkMRMLSliceNodeRed", "vtkMRMLSliceNodeGreen", "vtkMRMLSliceNodeYellow", "vtkMRMLViewNode1"]
mouse_listener = None
user_password = None
last_click_time = 0
last_event_time_lock = Lock()
ultra_width = 639
ultra_height = 479
ultra_y = 263
ultra_x = 0.5
l_width = 4.1
hifu_depth = 3.073
angle = 47.45
logger = None
extra_info = {'user': username}

vein_count = 0
segment_count = {}

scrcpy_command = [
    '--video-bit-rate', '32M',
    '--video-codec', 'h265',
    '--max-fps', '15',
    '--window-width', '550',
    '--window-height', '900',
    '--window-borderless',
    '--crop', '1800:2774:0:100',
    '--always-on-top',
    '--no-mouse-hover',
    '--turn-screen-off'
]
ultra_process = None
ultra_container = None

screenshot_command = [
    'adb', 'exec-out', 'screencap', '-p'
]

button1_style = """
    QPushButton {{
        font-size: 22px;
        color: rgba(255,255,255,217);
        text-align: center;
        background: #2075F5; 
        border-radius: {}px;
    }}
    QPushButton:pressed {{
        background: #004381;
        border: 1px solid #006DD1;
    }}
    QPushButton:checked {{
        background: #004381;
        border: 1px solid #006DD1;
    }}
    QPushButton:disabled {{
        background-color: rgba(0, 67, 129, 51);
        border: 1px solid #006DD1;
        color: rgba(255, 255, 255, 51);
    }}
"""

button2_style = """
    QPushButton {{
        font-size: 22px;
        color: rgba(255,255,255,217);
        text-align: center;
        background: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:1, 
                                stop:0 #00355D, stop:1 #000D1D); 
        border-radius: {}px;
        border: 1px solid #006DD1;
    }}
    QPushButton:pressed {{
        background: #004381;
        border: 1px solid #006DD1;
    }}
    QPushButton:checked {{
        background: #004381;
        border: 1px solid #006DD1;
    }}
    QPushButton:disabled {{
        background-color: rgba(0, 67, 129, 51);
        border: 1px solid #006DD1;
        color: rgba(255, 255, 255, 51);
    }}
"""

checkbox_style = '''               
    QCheckBox::indicator {
        width: 38px;
        height: 38px;
        background: #232428;
        border-radius: 4px;
        border: 1px solid #699FC3;
    }
    QCheckBox::indicator:unchecked {
        image: none;
    }
    QCheckBox::indicator:checked {
        background-color: #215BB3;
        border: 2px solid #215BB3;
    }
'''


class Communicate(qt.QObject):
    userActivity = qt.Signal()  # 定义用户激活信号


comm = Communicate()


def add_base_path():
    moduleManager = slicer.app.moduleManager()
    moduleFactoryManager = moduleManager.factoryManager()
    module_path = os.path.join(mainWindow().GetPythonBasePath(), "JExtension", "JBase")
    moduleFactoryManager.pyAddSearchPath(module_path)


def count_time(start_time, flag=1):
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"{flag} count_time elapse time：{elapsed_time}second\n")
    start_time = time.time()
    return start_time


def etimer(func):
    def wrapper(*args, **kwargs):
        import time
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f">> {func.__name__} elapse time：{elapsed_time}second <<")
        return result

    return wrapper


def unit_test_timer(func):
    def wrapper(*args, **kwargs):
        import time, math
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        elapsed_time = end_time - start_time
        elapsed_time = (int)(elapsed_time * 100000)
        elapsed_time = elapsed_time / 100
        write_test_log(f">> time cost：{elapsed_time} mill seccond <<")
        return result

    return wrapper


def memory_usage_timer(func):
    def wrapper(*args, **kwargs):
        import time, math, psutil
        start_memory = psutil.Process().memory_info().rss
        result = func(*args, **kwargs)
        end_memory = psutil.Process().memory_info().rss
        memory_change = end_memory - start_memory
        write_test_log(f">> memory change: {memory_change} bytes <<")
        return result

    return wrapper


node_id_to_pixelmap = {}
OPENAI_API_KEY = "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"
SERPAPI_API_KEY = ""
main_panel = None
QSS_QWidget_background_color = '#323232'
QSS_QWidget_MV_color = '#000000'
JsonData = None
InfoJson = None
parameter_config = None
parameter_config_path = None
global_int_index = 0
test_row_info = {}
loaded_modules = {}
unit_test = {}
vtkMRMLVolumeNode = "vtkMRMLVolumeNode"
vtkMRMLModelNode = "vtkMRMLModelNode"
vtkMRMLGridTransformNode = "vtkMRMLGridTransformNode"
vtkMRMLMarkupsNode = "vtkMRMLMarkupsNode"
vtkMRMLMarkupsCurveNode = "vtkMRMLMarkupsCurveNode"
vtkMRMLTransformNode = "vtkMRMLTransformNode"
vtkMRMLScalarVolumeNode = "vtkMRMLScalarVolumeNode"
vtkMRMLMultiVolumeNode = "vtkMRMLMultiVolumeNode"
vtkMRMLSegmentationNode = "vtkMRMLSegmentationNode"
vtkMRMLMarkupsFiducialNode = "vtkMRMLMarkupsFiducialNode"
vtkMRMLMarkupsPlaneNode = "vtkMRMLMarkupsPlaneNode"
vtkMRMLMarkupsLineNode = "vtkMRMLMarkupsLineNode"
vtkMRMLMarkupsROINode = "vtkMRMLMarkupsROINode"
vtkMRMLLabelMapVolumeNode = "vtkMRMLLabelMapVolumeNode"
vtkMRMLAnnotationROINode = "vtkMRMLAnnotationROINode"
vtkMRMLMarkupsAngleNode = "vtkMRMLMarkupsAngleNode"
vtkMRMLMarkupsClosedCurveNode = "vtkMRMLMarkupsClosedCurveNode"
vtkMRMLMarkupsCurveNode = "vtkMRMLMarkupsCurveNode"
vtkMRMLTableNode = "vtkMRMLTableNode"
vtkMRMLPlotSeriesNode = "vtkMRMLPlotSeriesNode"
vtkMRMLPlotChartNode = "vtkMRMLPlotChartNode"
FixedWidth = 400
UnitTestFinished = 2231220
NodeVolumeRegisterStatusChanged = 2231221
NodeInfochanged = 2231222
NodeLoadedEvent = 1231221
NodeRemovedEvent = 1231222
MainNodeLoadedEvent = 1231223
MainNodeRemovedEvent = 1231224
ArchiveFileLoadedEvent = 1231225
SceneDestroyEvent = 1231226
MultiVolumeLoadComplete = 1331211
MultiVolumeT1MappedLoadComplete = 1331212
JManagerModel_Modify = 1331213
JManagerModel_Delete = 1331214
JScalarVolumelUIColumn_Fresh = 1331215
JScalarVolumelUIColumn_2D_Pressed = 1331216
JScalarVolumelUIColumn_2D_Released = 1331217
Thanks_Friends = 1331218
Start_ModelList = 1331219
Return_ModelList = 1331220
JAddFiber_TargetPoint_Placed = 1331221
JAddFiber_EntryPoint_Placed = 1331222
BlendNodeListChanged = 1331223
TrackedDataUpdated = 1331224
GlobalPopupClose = 1331225
RecordScreenShot = 1331226
PlaybackCloseAllWindow = 1331227
RadiomicsComplete = 1331228
NewFileLoadedFromMain = 1331229
ThresholdApplyed = 1331230
FreshModelList = 1331231
JRemoveSkullBoardWidgetResult = 1331232
JSegmentHematomaConfirm = 1331233
JAddFiberReturn = 1331234
JAddFiberUnitRemvoed = 1331235
JMC_NormalRemoved = 1331236
SNormalScalarManagerVolumeNumber = 1331237
JPuncturGuideConfirm = 1331238
MarkupsModified = 1331239
BeforeSceneDestroyEvent = 1331240
JSegmentTumorConfirm = 1331241
JSurfaceCutWithUnitConfirm = 1331242
JSurfaceCutWithUnitCancel = 1331243
JAddFiberReturnSingle = 1331244
DoSNormalScalarManagerRemove = 1331245
DoSNormalScalarManagerAdd = 1431246
RadiomicsFinished = 1431247
ModelModifyComplete = 1431248
RDNCurveUpdate = 1431249
JAddFiberNormalReturnSingle = 1431250
JAddFiberNormalUnitRemvoed = 1431251
JRemoveSkullBoardWidgetMaskResult = 1431252
CombinePunctureGuidComplete = 1431253
JFootMaskResult = 1481254
DataDropedEvent = 1481255
ReconnectEvent = 1481256
ChangeFSMState = 1481257
ChangeToConnectState = 1481257
ChangeToDisConnectState = 1481258
PivotCalibProcessFinished = 1481259
HandEyeCalibProcessFinished = 1481260
SetCTMarkerProcessFinished = 1481261
MiniErrorProcessFinished = 1481262
DoctorAssitButttonResetState = 1481263
DoctorAssitLeftButttonResetState = 1481264
SpacePressEvent = 1481265
FrameworkTopLoginSucceed = 1481266
FrameworkTopRefreshCache = 1481267
FrameworkTopNeedUpdate = 1481268
ESCPressEvent = 1481269
DoctorAssitRightButttonResetState = 1481270
ResetSolution = 1481271
ResetVersion = 1481272
close_volume_rendering = 1481273
HideFrameButtomButton = 1481274
ShowFrameButtomButton = 1481275
AddSolutionScore = 1481276
ApplicationDisActivate = 9000001
ProgressStart = 1000001
ProgressValue = 1000002
TestScript_Return = 1000003
SetPage = 1000005
TestEvent = 1000006
RegisterActionComplete = 1000007
RegisterActionSuccess = 1000008
RegisterActionFail = 1000009
ChageGPTPage = 1000010
GotoPrePage = 1000011
GotoNextPage = 1000012
GotoCurrentPage = 1000013
InitPageList = 1000014
GotoLastPage = 1000015
TestInterval = 2
TickJump = 19870830
UltrasoundStartTreat = 2000001
SystemEmergeStart = 2000002
SystemEmergeStop = 2000003


def close_scene():
    import slicer
    send_event_str(BeforeSceneDestroyEvent, "1")
    slicer.mrmlScene.Clear(False)
    slicer.mrmlScene.SetURL("")
    slicer.app.coreIOManager().setDefaultSceneFileType("MRML Scene (.mrml)")
    send_event_str(SceneDestroyEvent, "1")
    # update_setting_from_config(mainWindow().GetProjectSettings())


def send_event_str(key, value=1):
    import slicer
    cn = AddNewNodeByClass("vtkMRMLColorNode")
    cn.SetAttribute("value", value.__str__())
    slicer.mrmlScene.InvokeEvent(key, cn)
    RemoveNode(cn)


def delete_all_children(widget):
    import qt
    child_widgets = widget.findChildren(qt.QWidget)  # 获取所有 QWidget 类型的子控件

    for child_widget in child_widgets:
        layout = widget.layout()  # 获取父控件的布局
        if layout is not None:
            layout.removeWidget(child_widget)  # 从布局中移除子控件
        child_widget.deleteLater()  # 删除子控件的实例

        if layout is not None:
            layout.update()  # 更新布局


def send_test_event(index, step, gap=300):
    import slicer, qt
    cn = AddNewNodeByClass("vtkMRMLColorNode")
    cn.SetAttribute("index", index.__str__())
    cn.SetAttribute("step", step.__str__())
    qt.QTimer.singleShot(gap, lambda: slicer.mrmlScene.InvokeEvent(TestEvent, cn))
    RemoveNode(cn)


#
# General
#

EXIT_SUCCESS = 0
EXIT_FAILURE = 1

TESTING_DATA_URL = "https://github.com/Slicer/SlicerTestingData/releases/download/"
"""Base URL for downloading testing data.

Datasets can be downloaded using URL of the form ``TESTING_DATA_URL + "SHA256/" + sha256ofDataSet``
"""

DATA_STORE_URL = "https://github.com/Slicer/SlicerDataStore/releases/download/"
"""Base URL for downloading data from Slicer Data Store. Data store contains atlases,
registration case library images, and various sample data sets.

Datasets can be downloaded using URL of the form ``DATA_STORE_URL + "SHA256/" + sha256ofDataSet``
"""


class ProgressTaskMultiProcessUnknown:

    def __init__(self, title="请稍候...") -> None:
        self.process_logic = None
        self.process = None
        self.interval = None
        self.target_finished_flag = False
        self.inner_index = 0
        self.m_ProgressBar = createProgressDialog2(autoClose=True, windowTitle=title)

        self.m_ProgressBar.connect('finished(int)', self.cancel_bar)

    def run(self, process_logic, process, output_func, interval=100.0):
        import slicer, qt
        self.output_func = output_func
        self.process_logic = process_logic
        self.process = process
        self.interval = interval
        self.run_flag = True
        self.process_logic.addProcess(self.process)
        self.process_logic.run()
        qt.QTimer.singleShot(1, self.onProgress)
        slicer.app.processEvents()
        return self.m_ProgressBar

    def on_process_output(self):
        print("on_process_output")
        self.output_func()
        self.target_finished_flag = True

    def add_step(self, value):
        import slicer
        import slicer.util as util
        self.inner_index += value
        util.setProgress((int)(self.inner_index))
        slicer.app.processEvents()

    def cancel_bar(self, int):
        import psutil
        PROCNAME = "python.exe"
        for proc in psutil.process_iter():
            # check whether the process name matches
            if proc.name() == PROCNAME:
                proc.kill()
        self.run_flag = False
        self.process.terminate()
        self.process.kill()

        pass

    def onProgress(self):
        import slicer, qt
        import slicer.util as util
        if self.target_finished_flag:
            step = 100 - self.inner_index
            if step > 10:
                step = 10
            else:
                self.run_flag = False
        else:
            step = (99 - self.inner_index) / self.interval
            if step > 1:
                step = 1
        self.inner_index += step
        util.setProgress((int)(self.inner_index))
        if self.run_flag:
            qt.QTimer.singleShot(100, lambda: self.onProgress())
        else:
            util.setProgress(100)
            util.HideProgressBar()
            if self.target_finished_flag:
                # self.success_func()
                pass
        slicer.app.processEvents()


class ProgressTaskCliUnknown():

    def __init__(self, run_func, success_func, fail_func, cancel_func, interval=100.0, parent=None) -> None:
        self.run_func = run_func
        self.success_func = success_func
        self.fail_func = fail_func
        self.progress_step = 0.0
        self.inner_index = 0
        self.run_flag = False
        self.target_finished_flag = False
        self.cli_node = None
        self.cancel_func = cancel_func
        self.interval = interval
        self.parent = parent

    def run(self, tag="请稍候"):
        import slicer, qt
        self.run_flag = True
        self.m_ProgressBar = createProgressDialog2(windowTitle=tag, parent=self.parent)
        # self.m_ProgressBar.show()
        # self.m_ProgressBar.connect('canceled()',self.cancel_bar)
        self.cli_node = self.run_func()
        self.cli_node.AddObserver('ModifiedEvent', self.onCliModified)
        qt.QTimer.singleShot(1, self.onProgress)
        slicer.app.processEvents()
        return self.m_ProgressBar

    def onProgress(self):
        import slicer, qt
        import slicer.util as util
        if self.target_finished_flag:
            step = 100 - self.inner_index
            if step > 10:
                step = 10
            else:
                self.run_flag = False
        else:
            step = (99 - self.inner_index) / self.interval
            if step > 1:
                step = 1
        self.inner_index += step
        util.setProgress((int)(self.inner_index))
        if self.run_flag:
            qt.QTimer.singleShot(100, lambda: self.onProgress())
        else:
            util.setProgress(100)
            util.HideProgressBar()
            if self.target_finished_flag:
                self.success_func()
        slicer.app.processEvents()

    def cancel_bar(self):
        self.run_flag = False
        self.cli_node.Cancel()
        self.cancel_func()

    def onCliModified(self, caller, event):
        import slicer, qt
        slicer.app.processEvents()
        if caller.GetStatus() == 32:
            self.target_finished_flag = True
        elif caller.GetStatus() == slicer.vtkMRMLCommandLineModuleNode.CompletedWithErrors:
            self.fail_func()
            self.target_finished_flag = True
        elif caller.GetStatus() == slicer.vtkMRMLCommandLineModuleNode.BusyMask:
            self.fail_func()
            self.target_finished_flag = True
            pass


view_tool_map = {}


def registe_view_tool(btn, key):
    if key in view_tool_map:
        values = view_tool_map[key]
        if btn in values:
            return
        else:
            view_tool_map[key].append(btn)
    else:
        view_tool_map[key] = [btn]


def trigger_view_tool(key):
    import slicer
    for inner in view_tool_map:
        if inner != key:
            btns = view_tool_map[inner]
            for btn in btns:
                btn.setChecked(False)


def get_meta_from_volume(node, key):
    import pydicom, slicer
    instanceUIDs = node.GetAttribute("DICOM.instanceUIDs")
    if instanceUIDs is None:
        return None
    instUids = instanceUIDs.split()
    filename = slicer.dicomDatabase.fileForInstance(instUids[0])
    ds = pydicom.dcmread(filename)
    # 检查Tag是否存在
    for elem in ds.iterall():
        if (key == str(elem.tag)):
            value = str(elem.value)
            return value
    return None


def quit():
    exit(EXIT_SUCCESS)


def exit(status=EXIT_SUCCESS):
    """Exits the application with the specified exit code.

  The method does not stops the process immediately but lets
  pending events to be processed.
  If exit() is called again while processing pending events,
  the error code will be overwritten.

  To make the application exit immediately, this code can be used.
  Note that forcing the application to exit may result in
  improperly released files and other resources.

  .. code-block:: python

    import sys
    sys.exit(status)

  """

    from slicer import app
    # Prevent automatic application exit (for example, triggered by starting Slicer
    # with "--testing" argument) from overwriting the exit code that we set now.
    app.commandOptions().runPythonAndExit = False
    app.exit(status)


def quit_without_warning():
    mainWindow().setNeedConfirm(False)
    quit()


def OnGlobalPopupMessage(message, alpha=0.5):
    mainWindow().OnGlobalPopupMessage(message, alpha)
    processEvents()


def HideGlobalPopup():
    mainWindow().HideGlobalPopup()
    processEvents()


def resourcePath():
    import qt
    base = qt.QApplication.applicationDirPath()
    resourcepath1 = base + "/Resources"
    return resourcepath1


def restart():
    """Restart the application.

  No confirmation popup is displayed.
  """

    from slicer import app
    app.restart()


def _readCMakeCache(var):
    import os
    from slicer import app

    prefix = var + ":"

    try:
        with open(os.path.join(app.slicerHome, "CMakeCache.txt")) as cache:
            for line in cache:
                if line.startswith(prefix):
                    return line.split("=", 1)[1].rstrip()

    except:
        pass

    return None


def sourceDir():
    """Location of the Slicer source directory.

  :type: :class:`str` or ``None``

  This provides the location of the Slicer source directory, if Slicer is being
  run from a CMake build directory. If the Slicer home directory does not
  contain a ``CMakeCache.txt`` (e.g. for an installed Slicer), the property
  will have the value ``None``.
  """
    return _readCMakeCache('Slicer_SOURCE_DIR')


def startupEnvironment():
    """Returns the environment without the Slicer specific values.

  Path environment variables like `PATH`, `LD_LIBRARY_PATH` or `PYTHONPATH`
  will not contain values found in the launcher settings.

  Similarly `key=value` environment variables also found in the launcher
  settings are excluded. Note that if a value was associated with a key prior
  starting Slicer, it will not be set in the environment returned by this
  function.

  The function excludes both the Slicer launcher settings and the revision
  specific launcher settings.
  """
    import slicer
    startupEnv = slicer.app.startupEnvironment()
    import os
    # "if varname" is added to reject empty key (it is invalid)
    if os.name == 'nt':
        # On Windows, subprocess functions expect environment to contain strings
        # and Qt provide us unicode strings, so we need to convert them.
        return {str(varname): str(startupEnv.value(varname)) for varname in list(startupEnv.keys()) if varname}
    else:
        return {varname: startupEnv.value(varname) for varname in list(startupEnv.keys()) if varname}


#
# Custom Import
#

def importVTKClassesFromDirectory(directory, dest_module_name, filematch='*'):
    from vtk import vtkObjectBase
    importClassesFromDirectory(directory, dest_module_name, vtkObjectBase, filematch)


def importQtClassesFromDirectory(directory, dest_module_name, filematch='*'):
    importClassesFromDirectory(directory, dest_module_name, 'PythonQtClassWrapper', filematch)


# To avoid globbing multiple times the same directory, successful
# call to ``importClassesFromDirectory()`` will be indicated by
# adding an entry to the ``__import_classes_cache`` set.
#
# Each entry is a tuple of form (directory, dest_module_name, type_info, filematch)
__import_classes_cache = set()


def importClassesFromDirectory(directory, dest_module_name, type_info, filematch='*'):
    # Create entry for __import_classes_cache
    cache_key = ",".join([str(arg) for arg in [directory, dest_module_name, type_info, filematch]])
    # Check if function has already been called with this set of parameters
    if cache_key in __import_classes_cache:
        return

    import glob, os, re, fnmatch
    re_filematch = re.compile(fnmatch.translate(filematch))
    for fname in glob.glob(os.path.join(directory, filematch)):
        if not re_filematch.match(os.path.basename(fname)):
            continue
        try:
            from_module_name = os.path.splitext(os.path.basename(fname))[0]
            importModuleObjects(from_module_name, dest_module_name, type_info)
        except ImportError as detail:
            import sys
            print(detail, file=sys.stderr)

    __import_classes_cache.add(cache_key)


def importModuleObjects(from_module_name, dest_module_name, type_info):
    """Import object of type 'type_info' (str or type) from module identified
  by 'from_module_name' into the module identified by 'dest_module_name'."""

    # Obtain a reference to the module identifed by 'dest_module_name'
    import sys
    dest_module = sys.modules[dest_module_name]

    # Skip if module has already been loaded
    if from_module_name in dir(dest_module):
        return

    # Obtain a reference to the module identified by 'from_module_name'
    import imp
    fp, pathname, description = imp.find_module(from_module_name)
    module = imp.load_module(from_module_name, fp, pathname, description)

    # Loop over content of the python module associated with the given python library
    for item_name in dir(module):

        # Obtain a reference associated with the current object
        item = getattr(module, item_name)

        # Check type match by type or type name
        match = False
        if isinstance(type_info, type):
            try:
                match = issubclass(item, type_info)
            except TypeError as e:
                pass
        else:
            match = type(item).__name__ == type_info

        if match:
            setattr(dest_module, item_name, item)


def show_left_right_style(left_width=395, right_width=395, docker_margin=10, left_module="Data", full_screen=True,
                          right_module="TestScript"):
    import slicer
    available_screenRect = slicer.app.desktop().availableGeometry()
    rect = mainWindow().GetScreenSize()
    title_height = slicer.app.style().pixelMetric(26)
    taskbar_height = rect.height() - available_screenRect.height()
    if full_screen:
        title_height = 0
        taskbar_height = 0

    layout_panel("left").setModule(left_module)
    layout_dock("left").setFixedSize(left_width, rect.height() - taskbar_height - title_height - docker_margin)
    layout_panel("left").show()
    layout_dock("left").show()

    layout_panel("right").setModule(right_module)
    layout_dock("right").setFixedSize(right_width, rect.height() - taskbar_height - title_height - docker_margin)
    layout_panel("right").show()
    layout_dock("right").show()

    layout_border_widget().show()
    layout_center_widget().show()
    layout_center_widget().setGeometry(0, 0, rect.width() - left_width - right_width,
                                       rect.height() - taskbar_height - title_height)


def set_stylesheet_green(pushbutton):
    btn_stylesheet = ""
    btn_stylesheet = btn_stylesheet + "QPushButton{background-color: green;border-radius: 4px;font-family: Source Han Sans CN-Regular, Source Han Sans CN; border: 0px}"
    btn_stylesheet = btn_stylesheet + "QPushButton:hover{background-color:#00c200 ;border: 1px solid #009900}"
    btn_stylesheet = btn_stylesheet + "QPushButton:pressed{background-color: green; border: 0px}"
    pushbutton.setStyleSheet(btn_stylesheet)


def set_stylesheet_orange(pushbutton):
    btn_stylesheet = ""
    btn_stylesheet = btn_stylesheet + "QPushButton{background-color: #ffaa00;border-radius: 4px;font-family: Source Han Sans CN-Regular, Source Han Sans CN; border: 0px}"
    btn_stylesheet = btn_stylesheet + "QPushButton:hover{background-color:#aa7100 ;border: 1px solid #aa7100}"
    btn_stylesheet = btn_stylesheet + "QPushButton:pressed{background-color:#ffaa00; border: 0px}"
    pushbutton.setStyleSheet(btn_stylesheet)


def set_stylesheet_normal(pushbutton):
    btn_stylesheet = ""
    btn_stylesheet = btn_stylesheet + "QPushButton{background-color: #525F7B;font-family: Source Han Sans CN-Regular, Source Han Sans CN;border-radius: 4px;}"
    btn_stylesheet = btn_stylesheet + "QPushButton:hover{background-color: #527FD9;}"
    btn_stylesheet = btn_stylesheet + "QPushButton:pressed{background-color: #0D1632;}"
    pushbutton.setStyleSheet(btn_stylesheet)


def show_left_style(left_width=495, docker_margin=10, measure_height=68, top_docker_height=56, middle_right_width=210,
                    middle_bottom_height=130, middle_top_module=None, middle_right_module=None,
                    middle_bottom_module=None, top_module=None, left_module=None, full_screen=False, show_main=True):
    import slicer
    available_screenRect = slicer.app.desktop().availableGeometry()
    rect = mainWindow().GetScreenSize()
    title_height = slicer.app.style().pixelMetric(26)
    taskbar_height = rect.height() - available_screenRect.height()

    if full_screen:
        taskbar_height = 0
        title_height = 0

    if top_module:
        layout_panel("top").setModule(top_module)
        layout_panel("top").setFixedSize(rect.width(), top_docker_height)
        layout_panel("top").show()
        layout_dock("top").show()
        layout_dock("top").setFixedHeight(top_docker_height)
    else:
        top_docker_height = -docker_margin

    if middle_right_module:
        layout_panel("middle_right").setModule(middle_right_module)
        layout_panel("middle_right").setGeometry(0, 0, middle_right_width,
                                                 rect.height() - top_docker_height - docker_margin)
        layout_panel("middle_right").show()
    else:
        middle_right_width = 0

    if middle_top_module:
        layout_panel("middle_top").show()
        layout_panel("middle_top").setModule(middle_top_module)
        layout_panel("middle_top").setGeometry(middle_right_width, 0, rect.width() - left_width - docker_margin,
                                               measure_height)
    else:
        measure_height = 0

    if middle_bottom_module:
        layout_panel("middle_bottom").setModule(middle_bottom_module)
        layout_panel("middle_bottom").setGeometry(middle_right_width,
                                                  rect.height() - top_docker_height - docker_margin - middle_bottom_height,
                                                  rect.width() - left_width - docker_margin - middle_right_width,
                                                  middle_bottom_height)
        layout_panel("middle_bottom").show()
    else:
        middle_bottom_height = 0

    if show_main:
        layout_border_widget().show()
        layout_center_widget().show()
        layout_border_widget().setGeometry(0, 0,
                                           rect.width() - left_width - docker_margin - middle_right_width,
                                           rect.height() - taskbar_height - title_height - top_docker_height - docker_margin)
        layout_center_widget().setGeometry(middle_right_width, measure_height,
                                           rect.width() - left_width - docker_margin - middle_right_width,
                                           rect.height() - measure_height - taskbar_height - title_height - top_docker_height - docker_margin - middle_bottom_height)

    if left_module:
        layout_panel("left").setModule(left_module)
        layout_dock("left").setFixedSize(left_width,
                                         rect.height() - taskbar_height - title_height - top_docker_height - docker_margin)
        layout_panel("left").show()
        layout_dock("left").show()
    else:
        left_width = 0
        docker_margin = 0
        layout_border_widget().setGeometry(0, 0, rect.width() - left_width - docker_margin,
                                           rect.height() - taskbar_height - title_height - top_docker_height - docker_margin)
        layout_center_widget().setGeometry(0, measure_height, rect.width() - left_width - docker_margin,
                                           rect.height() - measure_height - taskbar_height - title_height - top_docker_height - docker_margin - middle_bottom_height)


screen_width = 400


def get_width_extension():
    return screen_width


def get_height_extension():
    import slicer
    available_screenRect = slicer.app.desktop().availableGeometry()
    rect = mainWindow().GetScreenSize()
    title_height = slicer.app.style().pixelMetric(26)
    taskbar_height = rect.height() - available_screenRect.height()
    return available_screenRect.height()


def get_resource(resource_name, size=48, use_default_path=True):
    import os
    project_name_info = get_from_PAAA("current_project_selector_project_name")
    path = os.path.join(get_project_base_path(), "ProjectCache", project_name_info, "Icons", "Dark").replace("\\", "/")
    if not os.path.exists(path):
        os.makedirs(path)
    fullname = path + "/" + resource_name
    if not os.path.exists(fullname):
        if use_default_path:
            fullname = os.path.join(get_project_base_path(), "Resources", "Icons", f"empty{size}.png").replace("\\",
                                                                                                               "/")
    return fullname.replace('\\', '/')


def get_width_screen():
    import slicer
    available_screenRect = slicer.app.desktop().availableGeometry()
    return available_screenRect.width()


def get_height_screen():
    return mainWindow().GetScreenSize().height()


def show_right_research_style(right_module=None, middle_top_module=None, show_main=True, right_width=400):
    import qt

    if right_module:

        layout_dock("right").show()
        layout_panel("right").show()

        layout_panel("right").setModule(right_module)
        layout_panel("right").GetModuleWidget().setFixedWidth(right_width)
        layout_dock("right").setFixedWidth(right_width)
        if show_main:
            layout_dock("right").setSizePolicy(qt.QSizePolicy.Preferred, qt.QSizePolicy.Preferred)
            layout_dock("right").adjustSize()
        else:
            layout_dock("right").setSizePolicy(qt.QSizePolicy.Expanding, qt.QSizePolicy.Expanding)
            layout_dock("right").adjustSize()

        if middle_top_module:
            layout_panel("middle_top").show()
            layout_panel("middle_top").setModule(middle_top_module)

        if show_main:
            layout_border_widget().setSizePolicy(qt.QSizePolicy.Expanding, qt.QSizePolicy.Expanding)
            layout_border_widget().adjustSize()
            layout_border_widget().show()
        else:
            layout_border_widget().hide()
            pass


def hide_docker(docker_name):
    if layout_dock(docker_name).isVisible():
        layout_dock(docker_name).hide()


def show_docker(docker_name, module_name=None, width=0):
    layout_dock(docker_name).hide()
    if not layout_dock(docker_name).isVisible():
        layout_dock(docker_name).show()
    if module_name:
        if layout_panel(docker_name).currentModuleName() == module_name:
            return
        layout_panel(docker_name).setModule("")
        layout_panel(docker_name).setModule(module_name)
        layout_panel(docker_name).GetModuleWidget().setFixedWidth(width)
        layout_dock(docker_name).setFixedWidth(width)


def hide_center_widget():
    layout_border_widget().hide()


def show_center_widget():
    layout_border_widget().show()


def show_left_research_measure_style(module=None, show_main=True, width=400):
    setToolbarsVisible(False)
    setMenuBarsVisible(False)
    import qt, slicer
    if not show_main:
        hide_center_widget()
        show_docker("right", module, width)
    else:
        show_docker("right", module, width)
        show_center_widget()
        layout_panel("middle_right").show()
        layout_panel("middle_right").setModule("JMeasure")
        layout_panel("middle_right").setFixedWidth(47)
    return


def show_left_research_style(left_module=None, middle_top_module=None, middle_right_width=47, middle_right_module=None,
                             show_main=True, left_width=400, show_toolbar=True):
    import qt, slicer

    if not show_main:
        hide_center_widget()
        show_docker("left", left_module, left_width)
    else:
        show_docker("left", left_module, left_width)
        show_center_widget()
    if show_toolbar:
        setToolbarsVisible(True)
        setMenuBarsVisible(True)
    else:
        setToolbarsVisible(False)
        setMenuBarsVisible(False)
    return
    if left_module:

        layout_dock("left").show()
        layout_panel("left").show()

        layout_panel("left").setModule("")
        layout_panel("left").setModule(left_module)
        layout_panel("left").GetModuleWidget().setFixedWidth(left_width)
        layout_dock("left").setFixedWidth(left_width)

        if middle_top_module:
            layout_panel("middle_top").show()
            layout_panel("middle_top").setModule(middle_top_module)

        if middle_right_module:
            layout_panel("middle_right").show()
            layout_panel("middle_right").setModule(middle_right_module)
            layout_panel("middle_right").setFixedWidth(middle_right_width)

        if show_main:
            layout_border_widget().show()
        else:
            layout_border_widget().hide()
            pass
        # mainWindow().showNormal()
        # mainWindow().showMaximized()


def show_research_style(left_or_right="right", in_module=None, middle_top_module=None, show_main=True, width=400,
                        show_toolbar=True):
    import qt
    if show_toolbar:
        setToolbarsVisible(True)
        setMenuBarsVisible(True)
    else:
        setToolbarsVisible(False)
        setMenuBarsVisible(False)
        layout_panel("middle_right").setModule("JMeasure")
        layout_panel("middle_right").show()
        layout_panel("middle_right").setFixedWidth(47)

    if in_module:
        if left_or_right == "left":
            layout_dock("left").show()
            layout_panel("left").show()

            layout_panel("left").setModule(in_module)
            layout_panel("left").GetModuleWidget().setFixedWidth(width)
            layout_dock("left").setFixedWidth(width)
        else:
            layout_dock("right").show()
            layout_panel("right").show()

            layout_panel("right").setModule(in_module)
            layout_panel("right").GetModuleWidget().setFixedWidth(width)
            layout_dock("right").setFixedWidth(width)

        if middle_top_module:
            layout_panel("middle_top").show()
            layout_panel("middle_top").setModule(middle_top_module)

        if show_main:
            layout_border_widget().show()
        else:
            layout_border_widget().hide()
            pass

        mainWindow().showNormal()
        mainWindow().showMaximized()


def set_controller_vertical_layout():
    import slicer
    layoutManager = slicer.app.layoutManager()
    for sliceViewName in layoutManager.sliceViewNames():
        qMRMLSliceWidget = layoutManager.sliceWidget(sliceViewName)
        qMRMLSliceWidget.layout().setDirection(0)

        frame = qMRMLSliceWidget.frame
        frame.setSizePolicy(7, 7)

        sliceController = qMRMLSliceWidget.sliceController()

        BarWidget = sliceController.barWidget()
        BarWidget.layout().setDirection(2)

        SliceOffsetSlider = sliceController.sliceOffsetSlider()
        SliceOffsetSlider.SpinBox.hide()
        SliceOffsetSlider.Slider.setOrientation(0)
        SliceOffsetSlider.setSizePolicy(0, 1)
        qMRMLSliceControllerWidget = sliceController.qMRMLSliceControllerWidget
        qMRMLSliceControllerWidget.alignment = 34

        # for threeDViewIndex in range(layoutManager.threeDViewCount) :
        #   widget = layoutManager.threeDWidget(threeDViewIndex)
        #   widget.layout().setDirection(0)
        #   controller = layoutManager.threeDWidget(threeDViewIndex).threeDController()
        #   BarWidget = controller.barWidget()
        #   BarWidget.layout().setDirection(2)
        #   qMRMLThreeDViewControllerWidget = controller.qMRMLThreeDViewControllerWidget
        #   qMRMLThreeDViewControllerWidget.alignment = 34
        threeDWidget = slicer.app.layoutManager().threeDWidget(0)
        threeDWidget.layout().setDirection(0)

        threeDView = threeDWidget.threeDView()
        threeDView.setSizePolicy(7, 7)

        threeDController = threeDWidget.threeDController()

        BarWidget = threeDController.barWidget()
        BarWidget.layout().setDirection(2)

        qMRMLThreeDViewControllerWidget = threeDController.qMRMLThreeDViewControllerWidget
        qMRMLThreeDViewControllerWidget.alignment = 34


def show_right_style(right_width=495, right_only=False, docker_margin=10, measure_height=68, top_docker_height=55,
                     middle_right_width=47, middle_bottom_height=130, middle_top_module=None, middle_right_module=None,
                     middle_bottom_module=None, top_module=None, right_module=None, full_screen=False, show_main=True,
                     fixed=True, right_module_title=None, right_module_title_height=40, show_toolbar=False):
    import slicer, qt
    if show_toolbar:
        setToolbarsVisible(True)
        setMenuBarsVisible(True)
    else:
        setToolbarsVisible(False)
        setMenuBarsVisible(False)

    if right_only:
        layout_panel("right").setModule(right_module)
        layout_panel("right").show()
        layout_dock("right").show()
        return
    if top_module:
        layout_panel("top").setModule(top_module)
        layout_panel("top").show()
        layout_dock("top").show()

    if middle_right_module:
        layout_panel("middle_right").setModule("JMeasure")
        layout_panel("middle_right").show()
        layout_panel("middle_right").setFixedWidth(middle_right_width)

    if middle_top_module:
        layout_panel("middle_top").show()
        layout_panel("middle_top").setModule(middle_top_module)

    if middle_bottom_module:
        layout_panel("middle_bottom").setModule(middle_bottom_module)
        layout_panel("middle_bottom").show()

    if show_main:
        layout_border_widget().show()
        layout_center_widget().show()
    else:
        layout_border_widget().hide()
        layout_center_widget().hide()

    if right_module:
        layout_panel("right").setModule("")
        layout_panel("right").setModule(right_module)
        layout_panel("right").show()
        layout_dock("right").show()
        if fixed:
            layout_dock("right").setFixedWidth(right_width)
            layout_panel("right").setFixedWidth(right_width)
            layout_panel("right").GetModuleWidget().setFixedWidth(right_width)
        if full_screen:
            layout_border_widget().setFixedWidth(get_width_screen() - right_width)
    if right_module_title:
        layout_panel("right_title").setModule(right_module_title)
        layout_panel("right_title").show()
        layout_dock("right_title").show()
        layout_panel("right_title").setFixedHeight(right_module_title_height)
        layout_dock("right_title").setFixedHeight(right_module_title_height)


def show_right_style2(right_width=495, right_only=False, docker_margin=10, measure_height=68, top_docker_height=55,
                      middle_right_width=210, middle_bottom_height=130, middle_top_module=None,
                      middle_right_module=None, middle_bottom_module=None, top_module=None, right_module=None,
                      full_screen=False, show_main=True):
    import slicer

    available_screenRect = slicer.app.desktop().availableGeometry()
    rect = mainWindow().GetScreenSize()
    title_height = slicer.app.style().pixelMetric(26)
    taskbar_height = rect.height() - available_screenRect.height()

    if full_screen:
        taskbar_height = 0
        title_height = 0
    if right_only:
        layout_panel("right").setModule(right_module)
        # layout_dock("right").setFixedSize(right_width,rect.height()-taskbar_height-title_height-top_docker_height-docker_margin)
        layout_dock("right").setFixedWidth(right_width)
        layout_panel("right").show()
        layout_dock("right").show()
        return
    if top_module:
        layout_panel("top").setModule(top_module)
        layout_panel("top").setFixedHeight(top_docker_height)
        layout_panel("top").show()
        layout_dock("top").show()
        layout_dock("top").setFixedHeight(top_docker_height)
    else:
        top_docker_height = -docker_margin

    if middle_right_module:
        layout_panel("middle_right").setModule(middle_right_module)
        layout_panel("middle_right").setGeometry(rect.width() - right_width - docker_margin - middle_right_width, 0,
                                                 middle_right_width, rect.height() - top_docker_height - docker_margin)
        layout_panel("middle_right").show()
    else:
        middle_right_width = 0

    if middle_top_module:
        layout_panel("middle_top").show()
        layout_panel("middle_top").setModule(middle_top_module)
        layout_panel("middle_top").setGeometry(0, 0, rect.width() - right_width - docker_margin, measure_height)
    else:
        measure_height = 0

    if middle_bottom_module:
        layout_panel("middle_bottom").setModule(middle_bottom_module)
        layout_panel("middle_bottom").setGeometry(0,
                                                  rect.height() - top_docker_height - docker_margin - middle_bottom_height,
                                                  rect.width() - right_width - docker_margin - middle_right_width,
                                                  middle_bottom_height)
        layout_panel("middle_bottom").show()
    else:
        middle_bottom_height = 0

    if show_main:
        layout_border_widget().show()
        layout_center_widget().show()
        # layout_border_widget().setGeometry(0,0,
        # rect.width()-right_width-middle_right_width-docker_margin,
        # rect.height()-taskbar_height-title_height-top_docker_height-docker_margin)
        # layout_center_widget().setGeometry(0,measure_height,
        # rect.width()-right_width-middle_right_width-docker_margin,
        # rect.height()-taskbar_height-title_height-top_docker_height-docker_margin-measure_height-middle_bottom_height)

    if right_module:
        # layout_panel("right").setStyleSheet("background-color: red")
        # layout_dock("right").setStyleSheet("background-color: yellow")
        layout_panel("right").setModule(right_module)
        layout_dock("right").setFixedWidth(right_width)
        layout_panel("right").setFixedWidth(right_width)
        layout_panel("right").show()
        layout_dock("right").show()
    else:
        right_width = 0
        docker_margin = 0
        layout_border_widget().setGeometry(0, 0, rect.width() - right_width - docker_margin,
                                           rect.height() - taskbar_height - title_height - top_docker_height - docker_margin)
        layout_center_widget().setGeometry(0, measure_height, rect.width() - right_width - docker_margin,
                                           rect.height() - measure_height - taskbar_height - title_height - top_docker_height - docker_margin - middle_bottom_height)


def layout_panel(direction):
    return mainWindow().GetModulePanel(direction)


def layout_dock(direction):
    return mainWindow().GetDockWidget(direction)


def layout_center_widget():
    return mainWindow().GetCenterWidget()


def layout_border_widget():
    return mainWindow().GetCenterWidgetBorder()


def resetlayout():
    return mainWindow().ResetLayout()


def fresh_project_ui_frame():
    layout_panel("left").show()
    layout_dock("left").hide()
    layout_panel("right").show()
    layout_dock("right").hide()
    layout_dock("top").hide()
    layout_panel("top").show()
    layout_dock("bottom").hide()
    layout_panel("bottom").show()
    layout_dock("right_title").hide()
    layout_panel("right_title").show()

    layout_panel("middle_right").hide()
    layout_panel("middle_left").hide()
    layout_panel("middle_bottom").hide()
    layout_panel("middle_top").hide()

    layout_panel("left").setModule("")
    layout_panel("right").setModule("")
    layout_panel("top").setModule("")
    layout_panel("bottom").setModule("")
    layout_panel("middle_left").setModule("")
    layout_panel("middle_right").setModule("")
    layout_panel("middle_bottom").setModule("")
    layout_panel("middle_top").setModule("")


#
# UI
#

def lookupTopLevelWidget(objectName):
    """Loop over all top level widget associated with 'slicer.app' and
  return the one matching 'objectName'

  :raises RuntimeError: if no top-level widget is found by that name
  """
    from slicer import app
    for w in app.topLevelWidgets():
        if hasattr(w, 'objectName'):
            if w.objectName == objectName: return w
    # not found
    raise RuntimeError("Failed to obtain reference to '%s'" % objectName)


def getLogPath():
    import os
    slicerhome = mainWindow().slicerhome()
    path1 = os.path.join(slicerhome, "GLPyModule", "Project", 'tmp')
    path2 = os.path.join(slicerhome, "GLPyModule", "Project", "tmp")
    if isPackage():
        return path1
    else:
        return path2


def getScreenCapturePath():
    import os
    slicerhome = mainWindow().slicerhome()
    folder_path = os.path.join(slicerhome, "bin", "screencapture")
    if not isPackage():
        folder_path = os.path.join(slicerhome, "bin", "Release", "screencapture")
    if not os.path.exists(folder_path):
        # 如果文件夹不存在，则创建它
        os.makedirs(folder_path)
    folder_path = folder_path.replace('\\', '/')
    folder_path = folder_path.replace('bin/../', '')
    return folder_path


def isPackage():
    return mainWindow().IsPackages()


def show_gif_progress():
    if getjson("show_gif_progressbar") == "2":
        mainWindow().ShowGifProgressBar()
    processEvents()


def hide_gif_progress():
    if getjson("show_gif_progressbar") == "2":
        mainWindow().HideGifProgressBar()
    processEvents()


def mainWindow():
    """Get main window widget (qSlicerMainWindow object)

  :return: main window widget, or ``None`` if there is no main window
  """
    try:
        mw = lookupTopLevelWidget('qSlicerMainWindow')
    except RuntimeError:
        # main window not found, return None
        # Note: we do not raise an exception so that this function can be conveniently used
        # in expressions such as `parent if parent else mainWindow()`
        mw = None
    return mw


def pythonShell():
    """Get Python console widget (ctkPythonConsole object)

  :raises RuntimeError: if not found
  """
    from slicer import app
    console = app.pythonConsole()
    if not console:
        raise RuntimeError("Failed to obtain reference to python shell")
    return console


def showStatusMessage(message, duration=0):
    """Display ``message`` in the status bar.
  """
    mw = mainWindow()
    if not mw or not mw.statusBar():
        return False
    mw.statusBar().showMessage(message, duration)
    return True


def findChildren(widget=None, name="", text="", title="", className=""):
    """ Return a list of child widgets that meet all the given criteria.

  If no criteria are provided, the function will return all widgets descendants.
  If no widget is provided, slicer.util.mainWindow() is used.
  :param widget: parent widget where the widgets will be searched
  :param name: name attribute of the widget
  :param text: text attribute of the widget
  :param title: title attribute of the widget
  :param className: className() attribute of the widget
  :return: list with all the widgets that meet all the given criteria.
  """
    # TODO: figure out why the native QWidget.findChildren method does not seem to work from PythonQt
    import slicer, fnmatch
    if not widget:
        widget = mainWindow()
    if not widget:
        return []
    children = []
    parents = [widget]
    kwargs = {'name': name, 'text': text, 'title': title, 'className': className}
    expected_matches = []
    for kwarg in kwargs.keys():
        if kwargs[kwarg]:
            expected_matches.append(kwarg)
    while parents:
        p = parents.pop()
        # sometimes, p is null, f.e. when using --python-script or --python-code
        if not p:
            continue
        if not hasattr(p, 'children'):
            continue
        parents += p.children()
        matched_filter_criteria = 0
        for attribute in expected_matches:
            if hasattr(p, attribute):
                attr_name = getattr(p, attribute)
                if attribute == 'className':
                    # className is a method, not a direct attribute. Invoke the method
                    attr_name = attr_name()
                # Objects may have text attributes with non-string value (for example,
                # QUndoStack objects have text attribute of 'builtin_qt_slot' type.
                # We only consider string type attributes.
                if isinstance(attr_name, str):
                    if fnmatch.fnmatchcase(attr_name, kwargs[attribute]):
                        matched_filter_criteria = matched_filter_criteria + 1
        if matched_filter_criteria == len(expected_matches):
            children.append(p)
    return children


def findChild(widget, name):
    """Convenience method to access a widget by its ``name``.

  :raises RuntimeError: if the widget with the given ``name`` does not exist.
  """
    errorMessage = "Widget named " + str(name) + " does not exists."
    child = None
    try:
        child = findChildren(widget, name=name)[0]
        if not child:
            raise RuntimeError(errorMessage)
    except IndexError:
        raise RuntimeError(errorMessage)
    return child


def addWidget(parent, child, x, y):
    import qt
    widget = qt.QWidget(parent)
    layout = qt.QHBoxLayout(widget)
    layout.addWidget(child)
    widget.setFixedSize(child.width, child.height)
    widget.move(x, y)


def remove_layouts(widget):
    import qt
    # 遍历子控件列表
    for i in reversed(range(widget.layout().count())):
        # 获取子控件
        item = widget.layout().itemAt(i)
        # 检查是否为QLayout类型
        if isinstance(item, (qt.QVBoxLayout, qt.QHBoxLayout)):
            # 从QWidget中移除QLayout
            widget.layout().removeItem(item)
            # 删除QLayout对象
            item.widget().setParent(None)
        elif isinstance(item.widget(), qt.QWidget):
            # 递归删除子QWidget下的所有QLayout
            remove_layouts(item.widget())


def removeFromParent2(widget):
    if widget is None:
        return
    widget_a = widget.parentWidget()  # 获取部件 A
    layout_b = widget.layout()  # 获取布局 B
    if widget_a:
        widget_a.setLayout(None)
    if layout_b:
        layout_b.setParent(None)
    widget.setParent(None)


def addWidget2(parent, child):
    import qt
    child.setParent(parent)
    if parent.layout() is None:
        layout = qt.QHBoxLayout(parent)
        layout.addWidget(child)
        layout.setSpacing(0)
        layout.setContentsMargins(0, 0, 0, 0)
    else:
        parent.layout().addWidget(child)
    # print(parent, parent.layout())


def addLayoutOnly(parent, layout):
    import qt
    qwidget = qt.QWidget()
    qwidget.setLayout(layout)
    addWidgetOnly(parent, qwidget)


def addWidgetOnly(parent, child, width=None):
    import qt

    if child:
        child.setParent(parent)
    if parent.layout() is None:
        layout = qt.QHBoxLayout(parent)
        if child:
            layout.addWidget(child)
        layout.setSpacing(0)
        layout.setContentsMargins(0, 0, 0, 0)
    else:
        layout = parent.layout()

        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.setParent(None)

                # 可以选择删除部件
                # widget.deleteLater()

            del item
        if child:
            if width != None:
                child.setFixedWidth(width)
            parent.layout().addWidget(child)


def get_project_config(title, key, default_value):
    import configparser
    config = configparser.ConfigParser()
    try:
        with open(mainWindow().GetProjectIniPath().replace("\\", "/"), encoding='utf-8') as f:
            config.read_file(f)
            f.close()
        scene_extension = config.get(title, key, fallback=default_value)
    except AttributeError:
        return default_value
    return scene_extension


def loadUI(path):
    """ Load UI file ``path`` and return the corresponding widget.

  :raises RuntimeError: if the UI file is not found or if no
   widget was instantiated.
  """
    import qt
    qfile = qt.QFile(path)
    if not qfile.exists():
        errorMessage = "Could not load UI file: file not found " + str(path) + "\n\n"
        raise RuntimeError(errorMessage)
    qfile.open(qt.QFile.ReadOnly)
    loader = qt.QUiLoader()
    widget = loader.load(qfile)
    if not widget:
        errorMessage = "Could not load UI file: " + str(path) + "\n\n"
        raise RuntimeError(errorMessage)
    return widget


def startQtDesigner(args=None):
    """ Start Qt Designer application to allow editing UI files.
  """
    import slicer
    cmdLineArguments = []
    if args is not None:
        if isinstance(args, str):
            cmdLineArguments.append(args)
        else:
            cmdLineArguments.extend(args)
    return slicer.app.launchDesigner(cmdLineArguments)


def get_loading_movie(name=22):
    import os, qt
    path = os.path.join(get_project_base_path(), "Resources", "Icons", "loading%d.gif" % (name))
    movie = qt.QMovie(path)
    return movie


def multithread(func):
    import threading
    import time
    Newthread = threading.Thread(target=func)
    Newthread.start()


def save_scene_status():
    unit_test['scene_status'] = {}
    nodes = getNodes()
    for nodeid in nodes:
        unit_test['scene_status'][nodeid] = nodes[nodeid]


def compare_scene_status():
    if 'scene_status' not in unit_test:
        showWarningText("empty scene save information")
        return
    nodes = getNodes()
    added_nodes = []
    removed_nodes = []
    for nodeid in nodes:
        node = nodes[nodeid]
        if nodeid not in unit_test['scene_status']:
            added_nodes.append(node)
    for key in unit_test['scene_status']:
        value = unit_test['scene_status'][key]
        if key not in nodes:
            removed_nodes.append(value)
    return added_nodes, removed_nodes


def singleShot(time, func):
    import qt
    qt.QTimer.singleShot(time, func)


def copy_pixelmap(empty_label, target_label):
    pixmapA = target_label.pixmap
    if pixmapA.__str__() == "Pixmap 0, 0":
        print("empy pixel map")
        return
    empty_label.setPixmap(pixmapA)


def capture_snap_shot(volume_node, qlabel, orient=0):
    import qt, os

    import numpy as np
    from PIL import Image

    np_array = arrayFromVolume(volume_node)
    if len(np_array.shape) == 4:
        sd = int(np_array.shape[3] / 2)
        np_array = np_array.transpose(3, 0, 1, 2)
        np_array = np_array[sd]
    if orient == 0:
        pass
    elif orient == 1:
        np_array = np_array.transpose(1, 0, 2)
        rows = np_array.shape[0]
        for i in range(rows // 2):
            # 使用切片交换数据
            np_array[i], np_array[rows - i - 1] = np_array[rows - i - 1].copy(), np_array[i].copy()
    else:
        np_array = np_array.transpose(2, 0, 1)
    print("shape:", np_array.shape)
    np_shape = np_array.shape

    # DCE 中这里util.GetDisplayNode(node）是None，不用初始化
    # 如果其他地方也会出现这个问题，那么就考虑采取其他的方案
    if GetDisplayNode(volume_node) is None:
        return

    display_node = GetDisplayNode(volume_node)
    min_value = display_node.GetWindowLevelMin()
    max_value = display_node.GetWindowLevelMax()

    middle = int(np_shape[0] / 2)
    print("middle:", middle)
    pixels = np_array[middle]
    print("picture shape:", pixels.shape)
    pixels = np.clip(pixels, min_value, max_value)
    # 将灰度值转换为图像
    pixels = ((pixels - min_value) / (max_value - min_value) * 255).astype(np.uint8)

    image = Image.fromarray(pixels)
    current_file_directory = os.path.dirname(os.path.abspath(__file__))
    tempfile = os.path.join(current_file_directory, "tmp.png")
    image.save(tempfile)

    width = qlabel.width
    height = qlabel.height
    qimage = qt.QImage(tempfile)
    qpixmap = qt.QPixmap.fromImage(qimage)
    qpixmap = qpixmap.scaled(width, height)
    qlabel.setPixmap(qpixmap)
    os.remove(tempfile)


def ding():
    print("ding")
    import qt, os
    path = os.path.join(get_project_base_path(), "Resources", "Sound", "ding.wav").replace("\\", "/")
    qt.QSound.play(path)


def bind_button_with_pointset(btn, point_name, number=1, color=None, show_label=True):
    import slicer
    btn.setCheckable(True)
    point = getFirstNodeByName(point_name)
    if point is None:
        point = AddNewNodeByClass(vtkMRMLMarkupsFiducialNode)
        point.SetName(point_name)
    else:
        set_stylesheet_green(btn)
    GetDisplayNode(point).SetPointLabelsVisibility(show_label)
    if color is not None:
        # util.GetDisplayNode(point).SetColor(1,0,0)
        r, g, b = color
        GetDisplayNode(point).SetSelectedColor(r, g, b)

    btn.connect("toggled(bool)", lambda toggled: on_btn_toggled1513(toggled, btn, point, number))


def on_btn_toggled1513(toggled, btn, point, number):
    import slicer
    if toggled:
        interactionNode = slicer.app.applicationLogic().GetInteractionNode()
        selectionNode = slicer.app.applicationLogic().GetSelectionNode()
        selectionNode.SetReferenceActivePlaceNodeClassName("vtkMRMLMarkupsFiducialNode")
        selectionNode.SetActivePlaceNodeID(point.GetID())
        interactionNode.SetCurrentInteractionMode(interactionNode.Place)
        point.AddObserver(slicer.vtkMRMLMarkupsNode.PointPositionDefinedEvent,
                          lambda c, e: on_point_placed1513(c, e, btn, point, number))
    else:
        interactionNode = slicer.app.applicationLogic().GetInteractionNode()
        interactionNode.SetCurrentInteractionMode(interactionNode.ViewTransform)


def on_point_placed1513(caller, event, btn, point, number):
    n = point.GetNumberOfControlPoints()
    while n > number:
        point.RemoveNthControlPoint(0)
        n = point.GetNumberOfControlPoints()
    for i in range(n):
        point.SetNthControlPointLabel(i, (i + 1).__str__())
    btn.setChecked(False)
    set_stylesheet_green(btn)


def childWidgetVariables(widget):
    """ Get child widgets as attributes of an object.

  Each named child widget is accessible as an attribute of the returned object,
  with the attribute name matching the child widget name.
  This function provides convenient access to widgets in a loaded UI file.

  Example::

    uiWidget = slicer.util.loadUI(myUiFilePath)
    self.ui = slicer.util.childWidgetVariables(uiWidget)
    self.ui.inputSelector.setMRMLScene(slicer.mrmlScene)
    self.ui.outputSelector.setMRMLScene(slicer.mrmlScene)

  """
    ui = type('', (), {})()  # empty object
    childWidgets = findChildren(widget)
    for childWidget in childWidgets:
        if hasattr(childWidget, "name"):
            setattr(ui, childWidget.name, childWidget)
    return ui


def get_class_name(obj):
    return type(obj).__name__


def addParameterEditWidgetConnections(parameterEditWidgets, updateParameterNodeFromGUI):
    """ Add connections to get notification of a widget change.

  The function is useful for calling updateParameterNodeFromGUI method in scripted module widgets.

  .. note:: Not all widget classes are supported yet. Report any missing classes at https://discourse.slicer.org.

  Example::

    class SurfaceToolboxWidget(ScriptedLoadableModuleWidget, VTKObservationMixin):
      ...
      def setup(self):
        ...
        self.parameterEditWidgets = [
          (self.ui.inputModelSelector, "inputModel"),
          (self.ui.outputModelSelector, "outputModel"),
          (self.ui.decimationButton, "decimation"),
          ...]
        slicer.util.addParameterEditWidgetConnections(self.parameterEditWidgets, self.updateParameterNodeFromGUI)

      def updateGUIFromParameterNode(self, caller=None, event=None):
        if self._parameterNode is None or self._updatingGUIFromParameterNode:
          return
        self._updatingGUIFromParameterNode = True
        slicer.util.updateParameterEditWidgetsFromNode(self.parameterEditWidgets, self._parameterNode)
        self._updatingGUIFromParameterNode = False

      def updateParameterNodeFromGUI(self, caller=None, event=None):
        if self._parameterNode is None or self._updatingGUIFromParameterNode:
          return
        wasModified = self._parameterNode.StartModify()  # Modify all properties in a single batch
        slicer.util.updateNodeFromParameterEditWidgets(self.parameterEditWidgets, self._parameterNode)
        self._parameterNode.EndModify(wasModified)
  """

    for (widget, parameterName) in parameterEditWidgets:
        widgetClassName = widget.className()
        if widgetClassName == "QSpinBox":
            widget.connect("valueChanged(int)", updateParameterNodeFromGUI)
        elif widgetClassName == "QCheckBox":
            widget.connect("clicked()", updateParameterNodeFromGUI)
        elif widgetClassName == "QPushButton":
            widget.connect("toggled(bool)", updateParameterNodeFromGUI)
        elif widgetClassName == "qMRMLNodeComboBox":
            widget.connect("currentNodeIDChanged(QString)", updateParameterNodeFromGUI)
        elif widgetClassName == "QComboBox":
            widget.connect("currentIndexChanged(int)", updateParameterNodeFromGUI)
        elif widgetClassName == "ctkSliderWidget":
            widget.connect("valueChanged(double)", updateParameterNodeFromGUI)


def removeParameterEditWidgetConnections(parameterEditWidgets, updateParameterNodeFromGUI):
    """ Remove connections created by :py:meth:`addParameterEditWidgetConnections`.
  """

    for (widget, parameterName) in parameterEditWidgets:
        widgetClassName = widget.className()
        if widgetClassName == "QSpinBox":
            widget.disconnect("valueChanged(int)", updateParameterNodeFromGUI)
        elif widgetClassName == "QPushButton":
            widget.disconnect("toggled(bool)", updateParameterNodeFromGUI)
        elif widgetClassName == "qMRMLNodeComboBox":
            widget.disconnect("currentNodeIDChanged(QString)", updateParameterNodeFromGUI)
        elif widgetClassName == "QComboBox":
            widget.disconnect("currentIndexChanged(int)", updateParameterNodeFromGUI)
        elif widgetClassName == "ctkSliderWidget":
            widget.disconnect("valueChanged(double)", updateParameterNodeFromGUI)


def updateParameterEditWidgetsFromNode(parameterEditWidgets, parameterNode):
    """ Update widgets from values stored in a vtkMRMLScriptedModuleNode.

  The function is useful for implementing updateGUIFromParameterNode.

  Note: Only a few widget classes are supported now. More will be added later. Report any missing classes at discourse.slicer.org.

  See example in :py:meth:`addParameterEditWidgetConnections` documentation.
  """

    for (widget, parameterName) in parameterEditWidgets:
        widgetClassName = widget.className()
        parameterValue = parameterNode.GetParameter(parameterName)
        if widgetClassName == "QSpinBox":
            if parameterValue:
                widget.value = int(float(parameterValue))
            else:
                widget.value = 0
        if widgetClassName == "ctkSliderWidget":
            if parameterValue:
                widget.value = float(parameterValue)
            else:
                widget.value = 0.0
        elif widgetClassName == "QCheckBox" or widgetClassName == "QPushButton":
            widget.checked = (parameterValue == "true")
        elif widgetClassName == "QComboBox":
            widget.setCurrentText(parameterValue)
        elif widgetClassName == "qMRMLNodeComboBox":
            widget.currentNodeID = parameterNode.GetNodeReferenceID(parameterName)


def updateNodeFromParameterEditWidgets(parameterEditWidgets, parameterNode):
    """ Update vtkMRMLScriptedModuleNode from widgets.

  The function is useful for implementing updateParameterNodeFromGUI.

  Note: Only a few widget classes are supported now. More will be added later. Report any missing classes at discourse.slicer.org.

  See example in :py:meth:`addParameterEditWidgetConnections` documentation.
  """

    for (widget, parameterName) in parameterEditWidgets:
        widgetClassName = widget.className()
        if widgetClassName == "QSpinBox" or widgetClassName == "ctkSliderWidget":
            parameterNode.SetParameter(parameterName, str(widget.value))
        elif widgetClassName == "QCheckBox" or widgetClassName == "QPushButton":
            parameterNode.SetParameter(parameterName, "true" if widget.checked else "false")
        elif widgetClassName == "QComboBox":
            parameterNode.SetParameter(parameterName, widget.currentText)
        elif widgetClassName == "qMRMLNodeComboBox":
            parameterNode.SetNodeReferenceID(parameterName, widget.currentNodeID)


def setSliceViewerLayers(background='keep-current', foreground='keep-current', label='keep-current',
                         foregroundOpacity=None, labelOpacity=None, fit=False, rotateToVolumePlane=False):
    """ Set the slice views with the given nodes.

  If node ID is not specified (or value is 'keep-current') then the layer will not be modified.

  :param background: node or node ID to be used for the background layer
  :param foreground: node or node ID to be used for the foreground layer
  :param label: node or node ID to be used for the label layer
  :param foregroundOpacity: opacity of the foreground layer
  :param labelOpacity: opacity of the label layer
  :param rotateToVolumePlane: rotate views to closest axis of the selected background, foreground, or label volume
  :param fit: fit slice views to their content (position&zoom to show all visible layers)
  """
    import slicer
    def _nodeID(nodeOrID):
        nodeID = nodeOrID
        if isinstance(nodeOrID, slicer.vtkMRMLNode):
            nodeID = nodeOrID.GetID()
        return nodeID

    num = slicer.mrmlScene.GetNumberOfNodesByClass('vtkMRMLSliceCompositeNode')
    for i in range(num):
        sliceViewer = slicer.mrmlScene.GetNthNodeByClass(i, 'vtkMRMLSliceCompositeNode')
        if background != 'keep-current':
            sliceViewer.SetBackgroundVolumeID(_nodeID(background))
        if foreground != 'keep-current':
            sliceViewer.SetForegroundVolumeID(_nodeID(foreground))
        if foregroundOpacity is not None:
            sliceViewer.SetForegroundOpacity(foregroundOpacity)
        if label != 'keep-current':
            sliceViewer.SetLabelVolumeID(_nodeID(label))
        if labelOpacity is not None:
            sliceViewer.SetLabelOpacity(labelOpacity)

    if rotateToVolumePlane:
        if background != 'keep-current':
            volumeNode = slicer.mrmlScene.GetNodeByID(_nodeID(background))
        elif foreground != 'keep-current':
            volumeNode = slicer.mrmlScene.GetNodeByID(_nodeID(foreground))
        elif label != 'keep-current':
            volumeNode = slicer.mrmlScene.GetNodeByID(_nodeID(label))
        else:
            volumeNode = None
        if volumeNode:
            layoutManager = slicer.app.layoutManager()
            for sliceViewName in layoutManager.sliceViewNames():
                layoutManager.sliceWidget(sliceViewName).mrmlSliceNode().RotateToVolumePlane(volumeNode)

    if fit:
        layoutManager = slicer.app.layoutManager()
        if layoutManager is not None:
            sliceLogics = layoutManager.mrmlSliceLogics()
            for i in range(sliceLogics.GetNumberOfItems()):
                sliceLogic = sliceLogics.GetItemAsObject(i)
                if sliceLogic:
                    sliceLogic.FitSliceToAll()


def getPlaneIntersectionPoint():
    import slicer, vtk

    axialNode = slicer.mrmlScene.GetNodeByID('vtkMRMLSliceNodeRed')
    ortho1Node = slicer.mrmlScene.GetNodeByID('vtkMRMLSliceNodeYellow')
    ortho2Node = slicer.mrmlScene.GetNodeByID('vtkMRMLSliceNodeGreen')

    axialSliceToRas = axialNode.GetSliceToRAS()
    n1 = [axialSliceToRas.GetElement(0, 2), axialSliceToRas.GetElement(1, 2), axialSliceToRas.GetElement(2, 2)]
    x1 = [axialSliceToRas.GetElement(0, 3), axialSliceToRas.GetElement(1, 3), axialSliceToRas.GetElement(2, 3)]

    ortho1SliceToRas = ortho1Node.GetSliceToRAS()
    n2 = [ortho1SliceToRas.GetElement(0, 2), ortho1SliceToRas.GetElement(1, 2), ortho1SliceToRas.GetElement(2, 2)]
    x2 = [ortho1SliceToRas.GetElement(0, 3), ortho1SliceToRas.GetElement(1, 3), ortho1SliceToRas.GetElement(2, 3)]

    ortho2SliceToRas = ortho2Node.GetSliceToRAS()
    n3 = [ortho2SliceToRas.GetElement(0, 2), ortho2SliceToRas.GetElement(1, 2), ortho2SliceToRas.GetElement(2, 2)]
    x3 = [ortho2SliceToRas.GetElement(0, 3), ortho2SliceToRas.GetElement(1, 3), ortho2SliceToRas.GetElement(2, 3)]

    # Computed intersection point of all planes
    x = [0, 0, 0]

    n2_xp_n3 = [0, 0, 0]
    x1_dp_n1 = vtk.vtkMath.Dot(x1, n1)
    vtk.vtkMath.Cross(n2, n3, n2_xp_n3)
    vtk.vtkMath.MultiplyScalar(n2_xp_n3, x1_dp_n1)
    vtk.vtkMath.Add(x, n2_xp_n3, x)

    n3_xp_n1 = [0, 0, 0]
    x2_dp_n2 = vtk.vtkMath.Dot(x2, n2)
    vtk.vtkMath.Cross(n3, n1, n3_xp_n1)
    vtk.vtkMath.MultiplyScalar(n3_xp_n1, x2_dp_n2)
    vtk.vtkMath.Add(x, n3_xp_n1, x)

    n1_xp_n2 = [0, 0, 0]
    x3_dp_n3 = vtk.vtkMath.Dot(x3, n3)
    vtk.vtkMath.Cross(n1, n2, n1_xp_n2)
    vtk.vtkMath.MultiplyScalar(n1_xp_n2, x3_dp_n3)
    vtk.vtkMath.Add(x, n1_xp_n2, x)

    normalMatrix = vtk.vtkMatrix3x3()
    normalMatrix.SetElement(0, 0, n1[0])
    normalMatrix.SetElement(1, 0, n1[1])
    normalMatrix.SetElement(2, 0, n1[2])
    normalMatrix.SetElement(0, 1, n2[0])
    normalMatrix.SetElement(1, 1, n2[1])
    normalMatrix.SetElement(2, 1, n2[2])
    normalMatrix.SetElement(0, 2, n3[0])
    normalMatrix.SetElement(1, 2, n3[1])
    normalMatrix.SetElement(2, 2, n3[2])
    normalMatrixDeterminant = normalMatrix.Determinant()

    if abs(normalMatrixDeterminant) > 0.01:
        # there is an intersection point
        vtk.vtkMath.MultiplyScalar(x, 1 / normalMatrixDeterminant)
    else:
        # no intersection point can be determined, use just the position of the axial slice
        x = x1

    # 为了解决负数问题
    node = None
    layoutManager = slicer.app.layoutManager()
    for sliceViewName in layoutManager.sliceViewNames():
        compositeNode = layoutManager.sliceWidget(sliceViewName).sliceLogic().GetSliceCompositeNode()
        backgroundid = compositeNode.GetBackgroundVolumeID()
        node = GetNodeByID(backgroundid)
    if node:
        bounds = [0, 0, 0, 0, 0, 0]
        node.GetRASBounds(bounds)
        x[0] = x[0] - bounds[0]
        x[1] = x[1] - bounds[2]
        x[2] = x[2] - bounds[4]
    return x


def reinit_reverse():
    import slicer
    layoutManager = slicer.app.layoutManager()
    backgroundid = None
    forgroundid = None
    for sliceViewName in layoutManager.sliceViewNames():
        compositeNode = layoutManager.sliceWidget(sliceViewName).sliceLogic().GetSliceCompositeNode()
        backgroundid = compositeNode.GetBackgroundVolumeID()
        forgroundid = compositeNode.GetForegroundVolumeID()
    if backgroundid is not None and forgroundid is not None:
        background_node1 = GetNodeByID(backgroundid)
        forground_node1 = GetNodeByID(forgroundid)
        if background_node1 and forground_node1:
            reinit(background_node=forground_node1, foreground_node=background_node1)


def fit_to_background():
    import slicer
    layoutManager = slicer.app.layoutManager()
    for sliceViewName in layoutManager.sliceViewNames():
        controller = layoutManager.sliceWidget(sliceViewName).sliceController()
        node = layoutManager.sliceWidget(sliceViewName).mrmlSliceNode()
        # node.SetOrientationToDefault()
        controller.fitSliceToBackground()


def reset_focal_point():
    import slicer
    layoutManager = slicer.app.layoutManager()
    if layoutManager is not None:
        sliceLogics = layoutManager.mrmlSliceLogics()
        for i in range(sliceLogics.GetNumberOfItems()):
            sliceLogic = sliceLogics.GetItemAsObject(i)
            if sliceLogic:
                sliceLogic.FitSliceToAll()


def reinit(background_node=None, foreground_node=None, alpha=0.3, ori=False):
    import slicer
    layoutManager = slicer.app.layoutManager()

    for threeDViewIndex in range(layoutManager.threeDViewCount):
        controller = layoutManager.threeDWidget(threeDViewIndex).threeDController()
        controller.resetFocalPoint()
        threeDView = layoutManager.threeDWidget(threeDViewIndex).threeDView()
        threeDView.resetCamera()

    if background_node is None:
        for sliceViewName in layoutManager.sliceViewNames():
            controller = layoutManager.sliceWidget(sliceViewName).sliceController()
            node = layoutManager.sliceWidget(sliceViewName).mrmlSliceNode()
            # if ori:
            #  node.SetOrientationToDefault()
            controller.fitSliceToBackground()
        for threeDViewIndex in range(layoutManager.threeDViewCount):
            controller = layoutManager.threeDWidget(threeDViewIndex).threeDController()
            controller.resetFocalPoint()
        return

    for sliceViewName in layoutManager.sliceViewNames():
        compositeNode = layoutManager.sliceWidget(sliceViewName).sliceLogic().GetSliceCompositeNode()
        compositeNode.SetBackgroundVolumeID(background_node.GetID())
        if foreground_node:
            compositeNode.SetForegroundVolumeID(foreground_node.GetID())
        else:
            compositeNode.SetForegroundVolumeID(None)
        compositeNode.SetForegroundOpacity(alpha)

    if True:
        layoutManager = slicer.app.layoutManager()
        if layoutManager is not None:
            sliceLogics = layoutManager.mrmlSliceLogics()
            for i in range(sliceLogics.GetNumberOfItems()):
                sliceLogic = sliceLogics.GetItemAsObject(i)
                if sliceLogic:
                    sliceLogic.FitSliceToAll()


def setToolbarsVisible(visible, ignore=None):
    """Show/hide all existing toolbars, except those listed in ignore list.

  If there is no main window then the function has no effect.
  """
    import qt
    mw = mainWindow()
    if not mw:
        return
    for toolbar in mainWindow().findChildren('QToolBar'):
        if ignore is not None and toolbar in ignore:
            continue
        toolbar.setVisible(visible)
        toolbar.setContextMenuPolicy(qt.Qt.NoContextMenu)
        toolbar.setContextMenuPolicy(qt.Qt.PreventContextMenu)

    # Prevent sequence browser toolbar showing up automatically
    # when a sequence is loaded.
    # (put in try block because Sequence Browser module is not always installed)
    try:
        import slicer
        slicer.modules.sequences.autoShowToolBar = visible
    except:
        # Sequences module is not installed
        pass


def setMenuBarsVisible(visible, ignore=None):
    """Show/hide all menu bars, except those listed in ignore list.

  If there is no main window then the function has no effect.
  """
    mw = mainWindow()
    if not mw:
        return
    for menubar in mw.findChildren('QMenuBar'):
        if ignore is not None and menubar in ignore:
            continue
        menubar.setVisible(visible)


def setPythonConsoleVisible(visible):
    """Show/hide Python console.

  If there is no main window then the function has no effect.
  """
    mw = mainWindow()
    if not mw:
        return
    mw.pythonConsole().parent().setVisible(visible)


def setStatusBarVisible(visible):
    """Show/hide status bar

  If there is no main window or status bar then the function has no effect.
  """
    mw = mainWindow()
    if not mw or not mw.statusBar():
        return
    mw.statusBar().setVisible(visible)


def setViewControllersVisible(visible):
    """Show/hide view controller toolbar at the top of slice and 3D views"""
    import slicer
    slicer.app.layoutManager().setLayout(slicer.vtkMRMLLayoutNode.SlicerLayoutThreeOverThreeView)
    lm = slicer.app.layoutManager()
    for viewIndex in range(lm.threeDViewCount):
        lm.threeDWidget(viewIndex).threeDController().setVisible(visible)
    for sliceViewName in lm.sliceViewNames():
        lm.sliceWidget(sliceViewName).sliceController().setVisible(visible)
    for viewIndex in range(lm.tableViewCount):
        lm.tableWidget(viewIndex).tableController().setVisible(visible)
    for viewIndex in range(lm.plotViewCount):
        lm.plotWidget(viewIndex).plotController().setVisible(False)
    slicer.app.layoutManager().setLayout(slicer.vtkMRMLLayoutNode.SlicerLayoutInitialView)


def forceRenderAllViews():
    """Force rendering of all views"""
    import slicer
    lm = slicer.app.layoutManager()
    for viewIndex in range(lm.threeDViewCount):
        lm.threeDWidget(viewIndex).threeDView().forceRender()
    for sliceViewName in lm.sliceViewNames():
        lm.sliceWidget(sliceViewName).sliceView().forceRender()
    for viewIndex in range(lm.tableViewCount):
        lm.tableWidget(viewIndex).tableView().repaint()
    for viewIndex in range(lm.plotViewCount):
        lm.plotWidget(viewIndex).plotView().repaint()


def AddSignalNode(key, value):
    import slicer
    node1 = slicer.mrmlScene.CreateNodeByClass("vtkMRMLColorNode")
    node1.SetAttribute("Help_Node", "true")
    node1.SetAttribute("Help_Name", key.__str__())
    node1.SetAttribute("Help_Info_0", value.__str__())
    slicer.mrmlScene.AddNode(node1)


#
# IO
#

def loadNodeFromFile(filename, filetype, properties={}, returnNode=False):
    """Load node into the scene from a file.

  :param filename: full path of the file to load.
  :param filetype: specifies the file type, which determines which IO class will load the file.
  :param properties: map containing additional parameters for the loading.
  :param returnNode: Deprecated. If set to true then the method returns status flag and node
    instead of signalling error by throwing an exception.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  :raises RuntimeError: in case of failure
  """
    from slicer import app
    from vtk import vtkCollection

    properties['fileName'] = filename

    loadedNodesCollection = vtkCollection()
    success = app.coreIOManager().loadNodes(filetype, properties, loadedNodesCollection)
    loadedNode = loadedNodesCollection.GetItemAsObject(0) if loadedNodesCollection.GetNumberOfItems() > 0 else None

    # Deprecated way of returning status and node
    if returnNode:
        import logging
        logging.warning(
            "loadNodeFromFile `returnNode` argument is deprecated. Loaded node is now returned directly if `returnNode` is not specified.")
        return success, loadedNode

    if not success:
        errorMessage = "Failed to load node from file: " + str(filename)
        raise RuntimeError(errorMessage)

    return loadedNode


def loadNodesFromFile(filename, filetype, properties={}, returnNode=False):
    """Load nodes into the scene from a file.

  It differs from `loadNodeFromFile` in that it returns loaded node(s) in an iterator.

  :param filename: full path of the file to load.
  :param filetype: specifies the file type, which determines which IO class will load the file.
  :param properties: map containing additional parameters for the loading.
  :return: loaded node(s) in an iterator object.
  :raises RuntimeError: in case of failure
  """
    from slicer import app
    from vtk import vtkCollection
    properties['fileName'] = filename

    loadedNodesCollection = vtkCollection()
    success = app.coreIOManager().loadNodes(filetype, properties, loadedNodesCollection)
    if not success:
        errorMessage = "Failed to load nodes from file: " + str(filename)
        raise RuntimeError(errorMessage)

    return iter(loadedNodesCollection)


def loadColorTable(filename, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    return loadNodeFromFile(filename, 'ColorTableFile', {}, returnNode)


def loadFiberBundle(filename, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    return loadNodeFromFile(filename, 'FiberBundleFile', {}, returnNode)


def loadFiducialList(filename, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    return loadNodeFromFile(filename, 'FiducialListFile', {}, returnNode)


def loadAnnotationFiducial(filename, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    return loadNodeFromFile(filename, 'AnnotationFile', {'fiducial': 1}, returnNode)


def loadAnnotationRuler(filename, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    return loadNodeFromFile(filename, 'AnnotationFile', {'ruler': 1}, returnNode)


def loadAnnotationROI(filename, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    return loadNodeFromFile(filename, 'AnnotationFile', {'roi': 1}, returnNode)


def loadMarkupsFiducialList(filename, returnNode=False):
    """Load markups fiducials from file.

  .. deprecated:: 4.13.0
    Use the universal :func:`loadMarkups` function instead.
  """
    if returnNode:
        return loadMarkups(filename)
    else:
        node = loadMarkups(filename)
        return [node is not None, node]


def loadMarkupsCurve(filename):
    """Load markups curve from file.

  .. deprecated:: 4.13.0
    Use the universal :func:`loadMarkups` function instead.
  """
    return loadMarkups(filename)


def loadMarkupsClosedCurve(filename):
    """Load markups closed curve from file.

  .. deprecated:: 4.13.0
    Use the universal :func:`loadMarkups` function instead.
  """
    return loadMarkups(filename)


def loadMarkups(filename):
    """Load node from file.

  :param filename: full path of the file to load.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
  """
    return loadNodeFromFile(filename, 'MarkupsFile')


def loadModel(filename, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    return loadNodeFromFile(filename, 'ModelFile', {}, returnNode)


def loadScalarOverlay(filename, modelNodeID, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    return loadNodeFromFile(filename, 'ScalarOverlayFile', {'modelNodeId': modelNodeID}, returnNode)


def loadSegmentation(filename, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    return loadNodeFromFile(filename, 'SegmentationFile', {}, returnNode)


def loadTransform(filename, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    return loadNodeFromFile(filename, 'TransformFile', {}, returnNode)


def loadTable(filename):
    """Load table node from file.

  :param filename: full path of the file to load.
  :return: loaded table node
  """
    return loadNodeFromFile(filename, 'TableFile')


def loadLabelVolume(filename, properties={}, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    properties['labelmap'] = True
    return loadNodeFromFile(filename, 'VolumeFile', properties, returnNode)


def loadShaderProperty(filename, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    return loadNodeFromFile(filename, 'ShaderPropertyFile', {}, returnNode)


def loadText(filename):
    """Load node from file.

  :param filename: full path of the text file to load.
  :return: loaded text node.
  """
    return loadNodeFromFile(filename, 'TextFile')


def loadVolume(filename, properties={}, returnNode=False):
    """Load node from file.

  :param filename: full path of the file to load.
  :param properties:
    - name: this name will be used as node name for the loaded volume
    - labelmap: interpret volume as labelmap
    - singleFile: ignore all other files in the directory
    - center: ignore image position
    - discardOrientation: ignore image axis directions
    - autoWindowLevel: compute window/level automatically
    - show: display volume in slice viewers after loading is completed
    - fileNames: list of filenames to load the volume from
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    filetype = 'VolumeFile'
    return loadNodeFromFile(filename, filetype, properties, returnNode)


def loadSequence(filename, properties={}):
    """Load sequence (4D data set) from file.

  :param filename: full path of the file to load.
  :param properties:
    - name: this name will be used as node name for the loaded volume
    - show: display volume in slice viewers after loading is completed
    - colorNodeID: color node to set in the proxy nodes's display node
  :return: loaded sequence node.
  """
    filetype = 'SequenceFile'
    return loadNodeFromFile(filename, filetype, properties)


def loadScene(filename, properties={}):
    """Load node from file.

  :param filename: full path of the file to load.
  :param returnNode: Deprecated.
  :return: loaded node (if multiple nodes are loaded then a list of nodes).
    If returnNode is True then a status flag and loaded node are returned.
  """
    filetype = 'SceneFile'
    return loadNodeFromFile(filename, filetype, properties, returnNode=False)


def openAddDataDialog():
    from slicer import app
    return app.coreIOManager().openAddDataDialog()


def openAddVolumeDialog():
    from slicer import app
    return app.coreIOManager().openAddVolumeDialog()


def openAddModelDialog():
    from slicer import app
    return app.coreIOManager().openAddModelDialog()


def openAddScalarOverlayDialog():
    from slicer import app
    return app.coreIOManager().openAddScalarOverlayDialog()


def openAddSegmentationDialog():
    from slicer import app, qSlicerFileDialog
    return app.coreIOManager().openDialog('SegmentationFile', qSlicerFileDialog.Read)


def openAddTransformDialog():
    from slicer import app
    return app.coreIOManager().openAddTransformDialog()


def openAddColorTableDialog():
    from slicer import app
    return app.coreIOManager().openAddColorTableDialog()


def openAddFiducialDialog():
    from slicer import app
    return app.coreIOManager().openAddFiducialDialog()


def openAddMarkupsDialog():
    from slicer import app
    return app.coreIOManager().openAddMarkupsDialog()


def openAddFiberBundleDialog():
    from slicer import app
    return app.coreIOManager().openAddFiberBundleDialog()


def openAddShaderPropertyDialog():
    from slicer import app, qSlicerFileDialog
    return app.coreIOManager().openDialog('ShaderPropertyFile', qSlicerFileDialog.Read)


def openSaveDataDialog():
    from slicer import app
    return app.coreIOManager().openSaveDataDialog()


def AddControlPointGlobal(pos_array):
    import slicer, vtk
    markupsNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLMarkupsFiducialNode")
    markupsNode.AddControlPointWorld(vtk.vtkVector3d(pos_array[0], pos_array[1], pos_array[2]))
    return markupsNode


def CreateControlPointGlobal(pos_array):
    import slicer, vtk
    markupsNode = slicer.mrmlScene.CreateNodeByClass("vtkMRMLMarkupsFiducialNode")
    markupsNode.AddControlPointWorld(vtk.vtkVector3d(pos_array[0], pos_array[1], pos_array[2]))
    return markupsNode


def saveNode(node, filename, properties={}):
    """Save 'node' data into 'filename'.

  It is the user responsibility to provide the appropriate file extension.

  User has also the possibility to overwrite the fileType internally retrieved using
  method 'qSlicerCoreIOManager::fileWriterFileType(vtkObject*)'. This can be done
  by specifying a 'fileType'attribute to the optional 'properties' dictionary.
  """
    from slicer import app
    properties["nodeID"] = node.GetID()
    properties["fileName"] = filename
    if hasattr(properties, "fileType"):
        filetype = properties["fileType"]
    else:
        filetype = app.coreIOManager().fileWriterFileType(node)
    return app.coreIOManager().saveNodes(filetype, properties)


def saveScene(filename, properties={}):
    """Save the current scene.

  Based on the value of 'filename', the current scene is saved either
  as a MRML file, MRB file or directory.

  If filename ends with '.mrml', the scene is saved as a single file
  without associated data.

  If filename ends with '.mrb', the scene is saved as a MRML bundle (Zip
  archive with scene and data files).

  In every other case, the scene is saved in the directory
  specified by 'filename'. Both MRML scene file and data
  will be written to disk. If needed, directories and sub-directories
  will be created.
  """
    from slicer import app
    filetype = 'SceneFile'
    properties['fileName'] = filename
    return app.coreIOManager().saveNodes(filetype, properties)


def GetVolumeFilePath(node):
    import slicer
    instanceUIDs = node.GetAttribute("DICOM.instanceUIDs")
    if instanceUIDs is None:
        showWarningText("没有找到对应的Dicom文件")
        return None
    instUids = instanceUIDs.split()
    filename = slicer.dicomDatabase.fileForInstance(instUids[0])
    return filename


def GetVolumeFolderPath(node):
    import os
    filepath = GetVolumeFilePath(node)
    if filepath is None:
        return None
    return os.path.dirname(filepath)


#
# Module
#

def GetVolume2D(alias_name, is_registed=False):
    for node in getNodes("*").values():
        if node.GetAttribute("alias_name") == alias_name:
            if is_registed:
                if node.GetAttribute("is_registed"):
                    return node
            else:
                if not node.GetAttribute("is_registed"):
                    return node
    return None


def moduleSelector():
    """Return module selector widget.

  :return: module widget object
  :raises RuntimeError: if there is no module selector (for example, the application runs without a main window).
  """
    mw = mainWindow()
    if not mw:
        raise RuntimeError("Could not find main window")
    return mw.moduleSelector()


def selectModule(module):
    """Set currently active module.

  Throws a RuntimeError exception in case of failure (no such module or the application runs without a main window).
  :param module: module name or object
  :raises RuntimeError: in case of failure
  """
    moduleName = module
    if not isinstance(module, str):
        moduleName = module.name
    selector = moduleSelector()
    if not selector:
        raise RuntimeError("Could not find moduleSelector in the main window")
    moduleSelector().selectModule(moduleName)


def selectedModule():
    """Return currently active module.

  :return: module object
  :raises RuntimeError: in case of failure (no such module or the application runs without a main window).
  """
    selector = moduleSelector()
    if not selector:
        raise RuntimeError("Could not find moduleSelector in the main window")
    return selector.selectedModule


def moduleNames():
    """Get list containing name of all successfully loaded modules.

  :return: list of module names
  """
    from slicer import app
    return app.moduleManager().factoryManager().loadedModuleNames()


def getModule(moduleName):
    """Get module object from module name.

  :return: module object
  :raises RuntimeError: in case of failure (no such module).
  """
    from slicer import app
    module = app.moduleManager().module(moduleName)
    if not module:
        raise RuntimeError("Could not find module with name '%s'" % moduleName)
    return module


def getModuleGui(module):
    """Get module widget.

  .. deprecated:: 4.13.0
    Use the universal :func:`getModuleWidget` function instead.
  """
    return getModuleWidget(module)


def getNewModuleGui(module):
    """Create new module widget.

  .. deprecated:: 4.13.0
    Use the universal :func:`getNewModuleWidget` function instead.
  """
    return getNewModuleWidget(module)


def add_pixelmap_to_label(path, img):
    import qt
    pixelmap = qt.QPixmap(path)
    pixelmap_scaled = pixelmap.scaled(img.width * 0.95, img.height * 0.95, 0, 1)
    img.setPixmap(pixelmap_scaled)


def add_pixelmap_to_label(path, img, scale=1):
    import qt
    pixelmap = qt.QPixmap(path)
    pixelmap_scaled = pixelmap.scaled(img.width * scale, img.height * scale, 0, 1)
    img.setPixmap(pixelmap_scaled)


# 存储临时文件到filename里
# 将临时文件读取到image里
def CaptureScreenShot(filename, image):
    import slicer, qt

    sliceViewName = "Red"
    # Set view background to white
    view = slicer.app.layoutManager().sliceWidget(sliceViewName).sliceView()
    view.forceRender()

    # Capture a screenshot
    import ScreenCapture
    cap = ScreenCapture.ScreenCaptureLogic()
    cap.captureImageFromView(view, filename)

    img = qt.QImage()
    img.load(filename)

    pixelmap = qt.QPixmap.fromImage(img)
    if img.width() > img.height():
        gap = (img.width() - img.height()) / 2
        rect = qt.QRect(gap, 0, img.height(), img.height())
        pixelmap_cropped = pixelmap.copy(rect)
        pixelmap_scaled = pixelmap_cropped.scaled(80, 80, 0, 1)
        image.setPixmap(pixelmap_scaled)
    else:
        gap = (img.height() - img.width()) / 2
        rect = qt.QRect(0, gap, img.width(), img.width())
        pixelmap_cropped = pixelmap.copy(rect)
        pixelmap_scaled = pixelmap_cropped.scaled(80, 80, 0, 1)
        image.setPixmap(pixelmap_scaled)


def addPictureFromFile(filename, image, size_width=80, size_height=80):
    import qt
    img = qt.QImage()
    img.load(filename)

    pixelmap = qt.QPixmap.fromImage(img)

    if img.width() > img.height():
        gap = (img.width() - img.height()) / 2
        rect = qt.QRect(gap, 0, img.height(), img.height())
        pixelmap_cropped = pixelmap.copy(rect)
        pixelmap_scaled = pixelmap_cropped.scaled(size_width, size_height, 0, 1)
        image.setPixmap(pixelmap_scaled)
    else:
        gap = (img.height() - img.width()) / 2
        rect = qt.QRect(0, gap, img.width(), img.width())
        pixelmap_cropped = pixelmap.copy(rect)
        pixelmap_scaled = pixelmap_cropped.scaled(size_width, size_height, 0, 1)
        image.setPixmap(pixelmap_scaled)
    return pixelmap


def shorten_string(text, max_length):
    if len(text) > max_length:
        return text[:max_length] + "..."
    else:
        return text


def set_pixelmap_to_qlabel(pixelmap, image):
    import qt
    if image.width > image.height:
        gap = (image.width - image.height) / 2
        rect = qt.QRect(gap, 0, image.height, image.height)
        pixelmap_scaled = pixelmap.scaled(image.height, image.height, 0, 1)
        image.setPixmap(pixelmap_scaled)
    else:
        gap = (image.height - image.width) / 2
        rect = qt.QRect(0, gap, image.width, image.width)
        pixelmap_scaled = pixelmap.scaled(image.height, image.height, 0, 1)
        image.setPixmap(pixelmap_scaled)


def make_lock_file(key_file):
    key = ""
    with open(key_file, 'r') as f:
        key = f.read()
    print("key_file:", key_file)
    print("key_value:", key)
    lock_file = "D:/lock.txt"
    print("lock_file:", lock_file)
    with open(lock_file, 'w') as f:
        import hashlib
        lock = hashlib.sha256(key.encode()).hexdigest()
        lock = lock[-6:]
        print("lock_value:", lock)
        f.write(lock)


def get_token_number(text):
    import regex
    chinese_characters = regex.findall(r'\p{Script=Han}', text)
    english_words = regex.findall(r'\b[a-zA-Z]+\b', text)
    return len(chinese_characters) + len(english_words)


def exportNode(node, filename, properties={}, world=False):
    """Export 'node' data into 'filename'.

    If `world` is set to True then the node will be exported in the world coordinate system
    (equivalent to hardening the transform before exporting).

    This method is different from saveNode in that it does not modify any existing storage node
    and therefore does not change the filename or filetype that is used when saving the scene.
    """
    from slicer import app, vtkDataFileFormatHelper, vtkMRMLMessageCollection
    nodeIDs = [node.GetID()]
    fileNames = [filename]
    hardenTransform = world

    if "fileFormat" not in properties:
        foundFileFormat = None
        currentExtension = app.coreIOManager().extractKnownExtension(filename, node)
        fileWriterExtensions = app.coreIOManager().fileWriterExtensions(node)
        for fileFormat in fileWriterExtensions:
            extension = vtkDataFileFormatHelper.GetFileExtensionFromFormatString(fileFormat)
            if extension == currentExtension:
                foundFileFormat = fileFormat
                break
        if not foundFileFormat:
            raise ValueError(
                f"Failed to export {node.GetID()} - no known file format was found for filename {filename}")
        properties["fileFormat"] = foundFileFormat

    userMessages = vtkMRMLMessageCollection()
    success = app.coreIOManager().exportNodes(nodeIDs, fileNames, properties, hardenTransform, userMessages)

    if not success:
        import logging
        errorMessage = f"Failed to export node to file: {filename}"
        if userMessages.GetNumberOfMessages() > 0:
            errorMessage += "\n" + userMessages.GetAllMessagesAsString()
        logging.error(errorMessage)

    return success


def getModuleWidget(module):
    """Return module widget (user interface) object for a module.

  :param module: module name or module object
  :return: module widget object
  :raises RuntimeError: if the module does not have widget.
  """
    import slicer
    if isinstance(module, str):
        module = getModule(module)
    widgetRepr = module.widgetRepresentation()
    if not widgetRepr:
        raise RuntimeError("Could not find module widget representation with name '%s'" % module.name)
    if isinstance(widgetRepr, slicer.qSlicerScriptedLoadableModuleWidget):
        # Scripted module, return the Python class
        return widgetRepr.self()
    else:
        # C++ module
        return widgetRepr


def getNewModuleWidget(module):
    """Create new module widget instance.

  In general, not recommended, as module widget may be developed expecting that there is only a single
  instance of this widget. Instead, of instantiating a complete module GUI, it is recommended to create
  only selected widgets that are used in the module GUI.

  :param module: module name or module object
  :return: module widget object
  :raises RuntimeError: if the module does not have widget.
  """
    import slicer
    if isinstance(module, str):
        module = getModule(module)
    widgetRepr = module.createNewWidgetRepresentation()
    if not widgetRepr:
        raise RuntimeError("Could not find module widget representation with name '%s'" % module.name)
    if isinstance(widgetRepr, slicer.qSlicerScriptedLoadableModuleWidget):
        # Scripted module, return the Python class
        return widgetRepr.self()
    else:
        # C++ module
        return widgetRepr


def show_file_dialog(file_mode, file_format=""):
    import qt
    res_path = "empty data"
    fileDialog = qt.QFileDialog()
    fileDialog.setFileMode(file_mode)
    if file_format != "":
        fileDialog.setNameFilter(file_format)
    result = fileDialog.exec()
    selectDir = fileDialog.selectedFiles()
    if result == qt.QFileDialog.Rejected:
        return "", False
    if len(selectDir) == 0:
        return res_path, result
    saveDirName = selectDir[0]
    if not saveDirName:
        return res_path, result
    res_path = saveDirName
    return res_path, result


def generate_random_string(length=16):
    import random
    import string
    letters = string.ascii_letters  # 包含所有字母的字符串
    random_string = ''.join(random.choice(letters) for _ in range(length))
    return random_string


def getModuleLogic(module):
    """Get module logic object.

  Module logic allows a module to use features offered by another module.

  :param module: module name or module object
  :return: module logic object
  :raises RuntimeError: if the module does not have widget.
  """
    import slicer
    if isinstance(module, str):
        module = getModule(module)
    if isinstance(module, slicer.qSlicerScriptedLoadableModule):
        try:
            logic = getModuleGui(module).logic
        except AttributeError:
            # This widget does not have a logic instance in its widget class
            logic = None
    else:
        logic = module.logic()
    if not logic:
        raise RuntimeError("Could not find module widget representation with name '%s'" % module.name)
    return logic


def send_fsm_event(event):
    if 'fsm' in global_data_map:
        global_data_map['fsm'].send_event(event)


FSM_INITIAL_STATE = "FSM_INITIAL_STATE"
FSM_CHECK_DEVICE_STATE = "FSM_CHECK_DEVICE_STATE"
FSM_CHECK_DEVICE_INFO = "FSM_CHECK_DEVICE_INFO"
FSM_READY_STATE = "FSM_READY_STATE"
FSM_HAND_EYE_CALIBRATION_STATE = "FSM_HAND_EYE_CALIBRATION_STATE"
FSM_TREATMENT_STATE = "FSM_TREATMENT_STATE"
FSM_END_STATE = "FSM_END_STATE"


def is_fsm_state_equal(state):
    return global_data_map['fsm'].state == state


def rdn_communicate_log_init():
    project_name = get_from_PAAA("current_project_selector_project_name")
    base_project_path = mainWindow().GetProjectBasePath()
    communicate = os.path.join(base_project_path, "ProjectCache", project_name, "communicate.log")
    global_data_map['rdn_file_path'] = communicate
    if communicate:
        if os.path.exists(communicate):
            os.remove(communicate)


def write_rdn_communicate_log(data):
    return
    from datetime import datetime
    with open(global_data_map['rdn_file_path'], "a") as f:
        f.write(data)
        f.write("\n")
        f.flush()  # 强制刷新缓冲区，将数据写入磁盘


robot_arm_type = "Elibot"
ROBOT_ARM_TYPE_ELIBOT = "Elibot"
ROBOT_ARM_TYPE_UR = "UR"


def get_robot_arm_module_widget():
    if robot_arm_type == ROBOT_ARM_TYPE_ELIBOT:
        widget = getModuleWidget("JRobotArm")
    elif robot_arm_type == ROBOT_ARM_TYPE_UR:
        widget = getModuleWidget("JURRobotArm")
    else:
        raise Exception("unsupport _robot_arm type:" + double_camera_type.__str__())
    return widget


def get_robot_arm_widget():
    if robot_arm_type == ROBOT_ARM_TYPE_ELIBOT:
        widget = getModuleWidget("JRobotArm").ui.widget
    elif robot_arm_type == ROBOT_ARM_TYPE_UR:
        widget = getModuleWidget("JURRobotArm").ui.widget
    else:
        raise Exception("unsupport robot_arm type:" + robot_arm_type.__str__())
    return widget


double_camera_type = "IME"
DOUBLE_CAMERA_TYPE_IME = "IME"
DOUBLE_CAMERA_TYPE_NDI = "NDI"


def get_double_camera_module_widget():
    if double_camera_type == DOUBLE_CAMERA_TYPE_IME:
        widget = getModuleWidget("JIMEConnector")
    elif double_camera_type == DOUBLE_CAMERA_TYPE_NDI:
        widget = getModuleWidget("JNDIConnector")
    else:
        raise Exception("unsupport double camera type:" + double_camera_type.__str__())
    return widget


def get_double_camera_widget():
    if double_camera_type == DOUBLE_CAMERA_TYPE_IME:
        widget = getModuleWidget("JIMEConnector").ui.widget
    elif double_camera_type == DOUBLE_CAMERA_TYPE_NDI:
        widget = getModuleWidget("JNDIConnector").ui.widget
    else:
        raise Exception("unsupport double camera type:" + double_camera_type.__str__())
    return widget


def modulePath(moduleName):
    """Get module logic object.

  Module logic allows a module to use features offered by another module.
  Throws a RuntimeError exception if the module does not have widget.
  :param moduleName: module name
  :return: file path of the module
  """
    import slicer
    return eval('slicer.modules.%s.path' % moduleName.lower())


def reloadScriptedModule(moduleName):
    """Generic reload method for any scripted module.

  The function performs the following:

  * Ensure ``sys.path`` includes the module path and use ``imp.load_module``
    to load the associated script.
  * For the current module widget representation:

    * Hide all children widgets
    * Call ``cleanup()`` function and disconnect ``ScriptedLoadableModuleWidget_onModuleAboutToBeUnloaded``
    * Remove layout items

  * Instantiate new widget representation
  * Call ``setup()`` function
  * Update ``slicer.modules.<moduleName>Widget`` attribute
  """
    import imp, sys, os
    import slicer

    widgetName = moduleName + "Widget"

    # reload the source code
    filePath = modulePath(moduleName)
    p = os.path.dirname(filePath)
    if not p in sys.path:
        sys.path.insert(0, p)
    with open(filePath, "r") as fp:
        reloaded_module = imp.load_module(
            moduleName, fp, filePath, ('.py', 'r', imp.PY_SOURCE))

    # find and hide the existing widget
    parent = eval('slicer.modules.%s.widgetRepresentation()' % moduleName.lower())
    for child in parent.children():
        try:
            child.hide()
        except AttributeError:
            pass

    # if the module widget has been instantiated, call cleanup function and
    # disconnect "_onModuleAboutToBeUnloaded" (this avoids double-cleanup on
    # application exit)
    if hasattr(slicer.modules, widgetName):
        widget = getattr(slicer.modules, widgetName)
        widget.cleanup()

        if hasattr(widget, '_onModuleAboutToBeUnloaded'):
            slicer.app.moduleManager().disconnect('moduleAboutToBeUnloaded(QString)', widget._onModuleAboutToBeUnloaded)

    # remove layout items (remaining spacer items would add space above the widget)
    items = []
    for itemIndex in range(parent.layout().count()):
        items.append(parent.layout().itemAt(itemIndex))
    for item in items:
        parent.layout().removeItem(item)

    # create new widget inside existing parent
    widget = eval('reloaded_module.%s(parent)' % widgetName)
    widget.setup()
    widget.enter()
    setattr(slicer.modules, widgetName, widget)

    return reloaded_module


def setModulePanelTitleVisible(visible):
    """Show/hide module panel title bar at the top of module panel.

  If the title bar is not visible then it is not possible to drag and dock the
  module panel to a different location.

  If there is no main window then the function has no effect.
  """
    mw = mainWindow()
    if mw is None:
        return
    modulePanelDockWidget = mw.findChildren('QDockWidget', 'PanelDockWidget')[0]
    if visible:
        modulePanelDockWidget.setTitleBarWidget(None)
    else:
        import qt
        modulePanelDockWidget.setTitleBarWidget(qt.QWidget(modulePanelDockWidget))


def setApplicationLogoVisible(visible):
    """Show/hide application logo at the top of module panel.

  If there is no main window then the function has no effect.
  """
    mw = mainWindow()
    if mw is None:
        return
    widget = findChild(mw, "LogoLabel")
    widget.setVisible(visible)


def setModuleHelpSectionVisible(visible):
    """Show/hide Help section at the top of module panel.

  If there is no main window then the function has no effect.
  """
    mw = mainWindow()
    if mw is None:
        return
    modulePanel = findChild(mw, "ModulePanel")
    modulePanel.helpAndAcknowledgmentVisible = visible


def setDataProbeVisible(visible):
    """Show/hide Data probe at the bottom of module panel.

  If there is no main window then the function has no effect.
  """
    mw = mainWindow()
    if mw is None:
        return
    widget = findChild(mw, "DataProbeCollapsibleWidget")
    widget.setVisible(visible)


#
# Layout
#

def resetThreeDViews():
    """Reset focal view around volumes"""
    import slicer
    slicer.app.layoutManager().resetThreeDViews()


def resetSliceViews():
    """Reset focal view around volumes"""
    import slicer
    manager = slicer.app.layoutManager().resetSliceViews()


#
# MRML
#

class MRMLNodeNotFoundException(Exception):
    """Exception raised when a requested MRML node was not found."""
    pass


def getNodesByTypeByList(type, list1):
    valuelist = []
    for node in list1:
        if node.IsA(type):
            valuelist.append(node)
    return valuelist


def getNodes(pattern="*", scene=None, useLists=False):
    """Return a dictionary of nodes where the name or id matches the ``pattern``.

  By default, ``pattern`` is a wildcard and it returns all nodes associated
  with ``slicer.mrmlScene``.

  If multiple node share the same name, using ``useLists=False`` (default behavior)
  returns only the last node with that name. If ``useLists=True``, it returns
  a dictionary of lists of nodes.
  """
    import slicer, collections, fnmatch
    nodes = collections.OrderedDict()
    if scene is None:
        scene = slicer.mrmlScene
    count = scene.GetNumberOfNodes()
    for idx in range(count):
        node = scene.GetNthNode(idx)
        name = node.GetName()
        id = node.GetID()
        if (fnmatch.fnmatchcase(name, pattern) or
                fnmatch.fnmatchcase(id, pattern)):
            if useLists:
                nodes.setdefault(node.GetName(), []).append(node)
            else:
                nodes[node.GetName()] = node
    return nodes


def loadModelSecret(stl_path):
    import os
    KEY = b'\x1b\x5e\x30\x27\x77\x34\x6a\x50\x3f\x63\x3f\x2b\x44\x3a\x6c\x52'
    with open(stl_path, 'rb') as input_file:
        ciphertext = input_file.read()

    # 对每个字节进行异或操作
    plaintext = bytes([byte ^ KEY[i % len(KEY)] for i, byte in enumerate(ciphertext)])
    output_file_path = os.path.join(stl_path[:-4] + '.stl')
    with open(output_file_path, 'wb') as output_file:
        output_file.write(plaintext)

    # new_name = stl_path+".stl"
    # os.rename(stl_path,new_name)
    model = loadModel(output_file_path)
    # os.rename(new_name,stl_path)
    os.remove(output_file_path)
    return model


def getNode(pattern="*", index=0, scene=None):
    """Return the indexth node where name or id matches ``pattern``.

  By default, ``pattern`` is a wildcard and it returns the first node
  associated with ``slicer.mrmlScene``.

  :raises MRMLNodeNotFoundException: if no node is found
   that matches the specified pattern.
  """
    nodes = getNodes(pattern, scene)
    if not nodes:
        raise MRMLNodeNotFoundException(
            "could not find nodes in the scene by name or id '%s'" % (pattern if (isinstance(pattern, str)) else ""))
    return list(nodes.values())[index]


'''
  获取到SegmentationNode的第0个Segment对应到的volumeNode区域的平均灰度值

  1.获取到需要的segment在SegmentationNode中对应的 np_segment
  2.获取到需要的segment在SegmentationNode中对应的 extent  (index单位)
  3.获取到SegmentationNode中的IndexToWorld变换
  4.通过 extent 和 np_segment 的遍历,将segment上点的坐标转换成全局坐标
  5.将全局坐标点转换成 volume 的index坐标点,通过volume对应的的np_volume,获取到segment上的点对应到volume点的灰度值
  6.将所有的volume对应的灰度值相加,获取到平均灰度值
'''


def get_mean_pixel_value_from_segment_to_volume(segmentation_node, volumeNode):
    import slicer, vtk, numpy as np
    segmentId = GetNthSegmentID(segmentation_node, 0)
    seg = arrayFromSegment(segmentation_node, segmentId).T
    segImage = slicer.vtkOrientedImageData()
    segmentation_node.GetBinaryLabelmapRepresentation(segmentId, segImage)
    segImageExtent = segImage.GetExtent()
    ijkToWorld = vtk.vtkMatrix4x4()
    segImage.GetImageToWorldMatrix(ijkToWorld)
    np_volume = arrayFromVolume(volumeNode).T
    it = np.nditer(seg, flags=['multi_index'])

    pixel_values = []
    while not it.finished:
        # 将Segment上的点从Segmentation的局部坐标转换成全局坐标
        if seg[it.multi_index[0], it.multi_index[1], it.multi_index[2]] == 0:
            it.iternext()
            continue
        world = [0, 0, 0, 1]
        ijk = [segImageExtent[0] + it.multi_index[0], segImageExtent[2] + it.multi_index[1],
               segImageExtent[4] + it.multi_index[2]]
        ijkToWorld.MultiplyPoint(np.append(ijk, 1.0), world)

        # 将全局坐标转换成Volume的局部坐标
        volumeRasToIjk = vtk.vtkMatrix4x4()
        volumeNode.GetRASToIJKMatrix(volumeRasToIjk)
        point_Ijk = [0, 0, 0, 1]
        volumeRasToIjk.MultiplyPoint(world, point_Ijk)

        # 获取到Volume局部坐标对应的灰度值
        pixel_value = np_volume[round(point_Ijk[0]), round(point_Ijk[1]), round(point_Ijk[2])]

        pixel_values.append(pixel_value)

        it.iternext()
    return np.mean(np.array(pixel_values))


def get_multi_mean_pixel_value_from_segment_to_volume(segmentation_node, multi_volume_node):
    import slicer, vtk, numpy as np
    segmentId = GetNthSegmentID(segmentation_node, 0)
    seg = arrayFromSegment(segmentation_node, segmentId).T
    segImage = slicer.vtkOrientedImageData()
    segmentation_node.GetBinaryLabelmapRepresentation(segmentId, segImage)
    segImageExtent = segImage.GetExtent()

    ijkToWorld = vtk.vtkMatrix4x4()
    segImage.GetImageToWorldMatrix(ijkToWorld)

    # 将全局坐标转换成Volume的局部坐标
    volumeRasToIjk = vtk.vtkMatrix4x4()
    multi_volume_node.GetRASToIJKMatrix(volumeRasToIjk)
    it = np.nditer(seg, flags=['multi_index'])
    indexlist = []
    while not it.finished:
        # 将Segment上的点从Segmentation的局部坐标转换成全局坐标
        if seg[it.multi_index[0], it.multi_index[1], it.multi_index[2]] == 0:
            it.iternext()
            continue
        world = [0, 0, 0, 1]
        ijk = [segImageExtent[0] + it.multi_index[0], segImageExtent[2] + it.multi_index[1],
               segImageExtent[4] + it.multi_index[2]]
        ijkToWorld.MultiplyPoint(np.append(ijk, 1.0), world)
        point_Ijk = [0, 0, 0, 1]
        volumeRasToIjk.MultiplyPoint(world, point_Ijk)
        indexlist.append(point_Ijk)
        it.iternext()
    res = []
    multi_arr = arrayFromVolume(multi_volume_node).T
    for index in range(multi_arr[0]):
        pixel_values = []
        np_volume = multi_arr[index]
        for index in indexlist:
            pixel_value = np_volume[round(index[0]), round(index[1]), round(index[2])]
            pixel_values.append(pixel_value)
        res.append(np.mean(np.array(pixel_values)))
    return res


def getFirstNodeByClass(classname):
    import slicer
    nodes = slicer.mrmlScene.GetNodesByClass(classname)
    for node in nodes:
        return node
    return None


def getFirstNodeByClassByAttribute(classname, attname, attvalue):
    import slicer
    nodes = slicer.mrmlScene.GetNodesByClass(classname)
    for node in nodes:
        if node.GetAttribute(attname) == attvalue:
            return node
    return None


def getFirstNodeByNameByAttribute(name, attname, attvalue):
    import slicer
    nodes = getNodesByName(name)
    for node in nodes:
        if node.GetAttribute(attname) == attvalue:
            return node
    return None


def getNodeByAttribute(att, value=None):
    vals = []
    for node in getNodes("*").values():
        if value:
            if node.GetAttribute(att) == value:
                vals.append(node)
        else:
            if node.GetAttribute(att) != None:
                vals.append(node)
    return vals


def getNodesByAttribute(key, value):
    list1 = []
    for node in getNodes("*").values():
        if node.GetAttribute(key) == value:
            list1.append(node)
    return list1


def getNodesByClass(className, scene=None):
    """Return all nodes in the scene of the specified class."""
    import slicer
    if scene is None:
        scene = slicer.mrmlScene
    nodes = slicer.mrmlScene.GetNodesByClass(className)
    nodes.UnRegister(slicer.mrmlScene)
    nodeList = []
    nodes.InitTraversal()
    node = nodes.GetNextItemAsObject()
    while node:
        nodeList.append(node)
        node = nodes.GetNextItemAsObject()
    return nodeList


def getFirstNodeByClassByName(className, name, scene=None):
    """Return the first node in the scene that matches the specified node name and node class."""
    import slicer
    if scene is None:
        scene = slicer.mrmlScene
    return scene.GetFirstNode(name, className)


def getFirstNodeByExactName(name):
    for node in getNodes("*").values():
        if node.GetName() == name:
            return node
    return None


def RemoveAllControlPoints(node):
    node.RemoveAllControlPoints()


def getNodesByName(name):
    import slicer
    scene = slicer.mrmlScene
    return scene.GetNodesByName(name)


def getNumberOfNodesByName(name):
    nodes = getNodesByName(name)
    index = 0
    for node in nodes:
        index = index + 1
    return index


def getVolumeAndFaceOfModel(model):
    import vtk
    if model is None:
        return 0, 0
    polyData = model.GetPolyData()

    # 创建vtkMassProperties对象
    massProperties = vtk.vtkMassProperties()
    massProperties.SetInputData(polyData)
    massProperties.Update()

    # 获取模型的表面积和体积
    surfaceArea = massProperties.GetSurfaceArea()
    volume = massProperties.GetVolume()

    return volume, surfaceArea


def getFirstNodeByName(name, className=None):
    """Get the first MRML node that name starts with the specified name.

  Optionally specify a classname that must also match.
  """
    import slicer
    scene = slicer.mrmlScene
    return scene.GetFirstNode(name, className, False, False)


def GetNodeByID(id):
    if not id:
        return None
    import slicer
    scene = slicer.mrmlScene
    return scene.GetNodeByID(id)


def create_bounding_box(segmentationlist):
    import slicer
    new_segmentation = CreateDefaultSegmentationNode("tmp")
    segmentEditorWidget = slicer.modules.segmenteditor.widgetRepresentation().self().editor
    segmentEditorWidget.setSegmentationNode(new_segmentation)
    segmentEditorWidget.setActiveEffectByName("Logical operators")
    effect = segmentEditorWidget.activeEffect()
    effect.setParameter("Operation", "UNION")

    for subnode in segmentationlist:
        segment = GetNthSegment(subnode, 0)
        segmentid = GetNthSegmentID(subnode, 0)
        new_segmentation.GetSegmentation().AddSegment(segment, segmentid)
        effect.setParameter("ModifierSegmentID", segmentid)
        effect.self().onApply()
        new_segmentation.GetSegmentation().RemoveSegment(segmentid)

    return new_segmentation


def get_roi_by_segmentation(segmentation1):
    print("get_roi_by_segmentation segmentation node id is :", segmentation1.GetID())
    import SegmentStatistics, slicer
    segStatLogic = SegmentStatistics.SegmentStatisticsLogic()
    segStatLogic.getParameterNode().SetParameter("Segmentation", segmentation1.GetID())
    segStatLogic.getParameterNode().SetParameter("LabelmapSegmentStatisticsPlugin.obb_origin_ras.enabled", str(True))
    segStatLogic.getParameterNode().SetParameter("LabelmapSegmentStatisticsPlugin.obb_diameter_mm.enabled", str(True))
    segStatLogic.getParameterNode().SetParameter("LabelmapSegmentStatisticsPlugin.obb_direction_ras_x.enabled",
                                                 str(True))
    segStatLogic.getParameterNode().SetParameter("LabelmapSegmentStatisticsPlugin.obb_direction_ras_y.enabled",
                                                 str(True))
    segStatLogic.getParameterNode().SetParameter("LabelmapSegmentStatisticsPlugin.obb_direction_ras_z.enabled",
                                                 str(True))
    segStatLogic.computeStatistics()
    stats = segStatLogic.getStatistics()
    print("get_roi_by_segmentation stats is ", stats)
    import numpy as np
    for segmentId in stats["SegmentIDs"]:
        # Get bounding box
        obb_origin_ras = np.array(stats[segmentId, "LabelmapSegmentStatisticsPlugin.obb_origin_ras"])
        obb_diameter_mm = np.array(stats[segmentId, "LabelmapSegmentStatisticsPlugin.obb_diameter_mm"])
        obb_direction_ras_x = np.array(stats[segmentId, "LabelmapSegmentStatisticsPlugin.obb_direction_ras_x"])
        obb_direction_ras_y = np.array(stats[segmentId, "LabelmapSegmentStatisticsPlugin.obb_direction_ras_y"])
        obb_direction_ras_z = np.array(stats[segmentId, "LabelmapSegmentStatisticsPlugin.obb_direction_ras_z"])
        # Create ROI
        segment = segmentation1.GetSegmentation().GetSegment(segmentId)
        roi = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLMarkupsROINode")
        roi.SetName(segment.GetName() + " OBB")
        roi.GetDisplayNode().SetHandlesInteractive(False)  # do not let the user resize the box
        roi.SetSize(obb_diameter_mm)
        # Position and orient ROI using a transform
        obb_center_ras = obb_origin_ras + 0.5 * (
                obb_diameter_mm[0] * obb_direction_ras_x + obb_diameter_mm[1] * obb_direction_ras_y +
                obb_diameter_mm[2] * obb_direction_ras_z)
        boundingBoxToRasTransform = np.row_stack((np.column_stack(
            (obb_direction_ras_x, obb_direction_ras_y, obb_direction_ras_z, obb_center_ras)), (0, 0, 0, 1)))
        boundingBoxToRasTransformMatrix = slicer.util.vtkMatrixFromArray(boundingBoxToRasTransform)
        roi.SetAndObserveObjectToNodeMatrix(boundingBoxToRasTransformMatrix)
        return roi


def CreateDefaultSegmentationNode(name1):
    import slicer
    segment_node = getFirstNodeByName(name1)
    if segment_node is None:
        segment_node = AddNewNodeByClass("vtkMRMLSegmentationNode")
        segment_node.SetName(name1)
        segment = slicer.vtkSegment()
        segment.SetName("layer_1")
        segment.SetColor(0.9, 0.9, 0)
        segment_node.GetSegmentation().AddSegment(segment)
        segment_node.CreateDefaultDisplayNodes()
    return segment_node


def EnsureFirstNodeByNameByClass(nodename, classname="vtkMRMLScalarVolumeNode", delete=False):
    import slicer
    referenceVolumeNode = slicer.mrmlScene.GetFirstNodeByName(nodename)
    if referenceVolumeNode:
        if delete == True:
            RemoveNode(referenceVolumeNode)
            return EnsureFirstNodeByNameByClass(nodename, classname=classname)
        else:
            return referenceVolumeNode
    else:
        if classname == vtkMRMLSegmentationNode:
            referenceVolumeNode = CreateDefaultSegmentationNode(nodename)
        else:
            referenceVolumeNode = slicer.mrmlScene.AddNewNodeByClass(classname)
        referenceVolumeNode.SetName(nodename)
        return referenceVolumeNode


from contextlib import contextmanager


@contextmanager
def tryWithErrorDisplay(message=None, show=True, waitCursor=False):
    """Show an error display with the error details if an exception is raised.

    :param message: Text shown in the message box.
    :param show: If show is False, the context manager has no effect.
    :param waitCursor: If waitCrusor is set to True then mouse cursor is changed to
       wait cursor while the context manager is being run.

    .. code-block:: python

      import random

      def risky():
        if random.choice((True, False)):
          raise Exception('Error while trying to do some internal operations.')

      with slicer.util.tryWithErrorDisplay("Risky operation failed."):
        risky()
    """
    try:
        if waitCursor:
            import slicer, qt
            slicer.app.setOverrideCursor(qt.Qt.WaitCursor)
        yield
        if waitCursor:
            slicer.app.restoreOverrideCursor()
    except Exception as e:
        import slicer
        if waitCursor:
            slicer.app.restoreOverrideCursor()
        if show and not slicer.app.testingEnabled():
            if message is not None:
                errorMessage = f'{message}\n\n{e}'
            else:
                errorMessage = str(e)
            import traceback
            errorDisplay(errorMessage, detailedText=traceback.format_exc())
        raise


def EnsureFirstNodeByAttributeByNameByClass(attkey, attvalue, nodename="", classname="vtkMRMLScalarVolumeNode"):
    import slicer
    for node in getNodes("*").values():
        if node.GetAttribute(attkey) == attvalue:
            return node
    if classname is None:
        return None
    referenceVolumeNode = slicer.mrmlScene.AddNewNodeByClass(classname)
    if nodename != "":
        referenceVolumeNode.SetName(nodename)
    referenceVolumeNode.SetAttribute(attkey, attvalue)
    return referenceVolumeNode


def HideNode2D(node, is_show):
    if node is None:
        return
    displayNode = GetDisplayNode(node)
    if not displayNode:
        return
    displayNode.SetVisibility(True)
    displayNode.SetVisibility2D(is_show)
    displayNode.SetSliceIntersectionVisibility(is_show)


def Visibility3D(node):
    displaynode = GetDisplayNode(node)
    if not displaynode.GetVisibility():
        return False
    if not displaynode.GetVisibility3D():
        return False
    return True


def Visibility(node):
    displaynode = GetDisplayNode(node)
    if not displaynode.GetVisibility():
        return False
    return True


def RemoveTransfromNode(node):
    if node is None:
        return
    tranid = node.GetTransformNodeID()
    if tranid is None:
        return
    node.SetAndObserveTransformNodeID(None)


def ShowNode3D(node, is_show=True):
    if node is None:
        return
    displayNode = GetDisplayNode(node)
    if is_show:
        displayNode.SetVisibility(True)
    if displayNode:
        displayNode.SetVisibility3D(is_show)
    if is_show:
        if node.IsA("vtkMRMLSegmentationNode"):
            node.CreateClosedSurfaceRepresentation()


def is_folder_exist(path, subname):
    import os
    if os.path.exists(path):
        files = os.listdir(path)
        for f in files:
            if f == subname and os.path.isdir(os.path.join(path, f)):
                return True
    return False


def get_path_base():
    import os
    # D:\S521\S4R\Slicer-build\Applications\SlicerApp
    #
    parent_path = os.path.abspath("./")
    is_package = True
    if is_folder_exist(parent_path, 'CMakeFiles'):
        is_package = False
    if is_package:
        # D:\S521\S4RO\Slicer-build\_CPack_Packages\win-amd64\NSIS\Slicer-5.2.1--win-amd64
        parent_path = parent_path
    else:
        # D:\S521\S4R\Slicer-build\Applications\
        parent_path = os.path.dirname(parent_path)
        # D:\S521\S4R\Slicer-build
        parent_path = os.path.dirname(parent_path)
        # D:\S521\S4R
        parent_path = os.path.dirname(parent_path)
        # D:\S521\
        parent_path = os.path.dirname(parent_path)
    return parent_path


def HideNode(node):
    if node is None:
        return
    displayNode = GetDisplayNode(node)
    if displayNode:
        displayNode.SetVisibility(False)


def copy_color_and_opacity(origin_node, copy_node):
    scolor = GetDisplayNode(origin_node).GetColor()
    GetDisplayNode(copy_node).SetColor(scolor[0], scolor[1], scolor[2])
    percent = GetDisplayNode(origin_node).GetOpacity()
    GetDisplayNode(copy_node).SetOpacity(percent)
    percent = GetDisplayNode(origin_node).GetSliceIntersectionOpacity()
    GetDisplayNode(copy_node).SetSliceIntersectionOpacity(percent)


def copy_preset(origin_node, copy_node):
    import slicer
    volRenLogic = slicer.modules.volumerendering.logic()
    displayNode = volRenLogic.GetFirstVolumeRenderingDisplayNode(origin_node)
    if not displayNode:
        displayNode = volRenLogic.CreateDefaultVolumeRenderingNodes(origin_node)
    # Set up gradient vs opacity transfer function
    json_data = {}
    gradientOpacityTransferFunction = displayNode.GetVolumePropertyNode().GetVolumeProperty().GetGradientOpacity()
    size = gradientOpacityTransferFunction.GetSize()
    json_data["GradientOpacity"] = []
    for i in range(size):
        sub_json = {}
        val = [0, 0, 0, 0]
        gradientOpacityTransferFunction.GetNodeValue(i, val)
        sub_json[i] = val
        json_data["GradientOpacity"].append(sub_json)

    scalarOpacityTransferFunction = displayNode.GetVolumePropertyNode().GetVolumeProperty().GetScalarOpacity()
    size = scalarOpacityTransferFunction.GetSize()
    json_data["ScalarOpacity"] = []
    for i in range(size):
        sub_json = {}
        val = [0, 0, 0, 0]
        scalarOpacityTransferFunction.GetNodeValue(i, val)
        sub_json[i] = val
        json_data["ScalarOpacity"].append(sub_json)

    RGBTransferFunction = displayNode.GetVolumePropertyNode().GetVolumeProperty().GetRGBTransferFunction()
    size = RGBTransferFunction.GetSize()
    json_data["RGBTransferFunction"] = []
    for i in range(size):
        sub_json = {}
        val = [0, 0, 0, 0, 0, 0]
        RGBTransferFunction.GetNodeValue(i, val)
        sub_json[i] = val
        json_data["RGBTransferFunction"].append(sub_json)

    displayNode = volRenLogic.GetFirstVolumeRenderingDisplayNode(copy_node)
    if not displayNode:
        displayNode = volRenLogic.CreateDefaultVolumeRenderingNodes(copy_node)

    gradientOpacityTransferFunction = displayNode.GetVolumePropertyNode().GetVolumeProperty().GetGradientOpacity()
    gradientOpacityTransferFunction.RemoveAllPoints()
    GradientOpacityJsonData = json_data["GradientOpacity"]
    for data in GradientOpacityJsonData:
        for key in data:
            val = data[key]
            if len(val) != 4:
                showWarningText("json文件损坏")
                return
            gradientOpacityTransferFunction.AddPoint(val[0], val[1], val[2], val[3])

    scalarOpacityTransferFunction = displayNode.GetVolumePropertyNode().GetVolumeProperty().GetScalarOpacity()
    scalarOpacityTransferFunction.RemoveAllPoints()
    ScalarOpacityData = json_data["ScalarOpacity"]
    for data in ScalarOpacityData:
        for key in data:
            val = data[key]
            if len(val) != 4:
                showWarningText("json文件损坏")
                return
            scalarOpacityTransferFunction.AddPoint(val[0], val[1], val[2], val[3])

    RGBTransferFunction = displayNode.GetVolumePropertyNode().GetVolumeProperty().GetRGBTransferFunction()
    RGBTransferFunction.RemoveAllPoints()
    RGBTransferFunctionData = json_data["RGBTransferFunction"]
    for data in RGBTransferFunctionData:
        for key in data:
            val = data[key]
            if len(val) != 6:
                showWarningText("json文件损坏")
                return
            RGBTransferFunction.AddRGBPoint(val[0], val[1], val[2], val[3], val[4], val[5])


def get_widget_from_node(node):
    import slicer.util as util
    import slicer
    view = slicer.app.layoutManager().threeDWidget(0).threeDView()
    markupsDisplayableManager = view.displayableManagerByClassName('vtkMRMLMarkupsDisplayableManager')
    widget = markupsDisplayableManager.GetWidget(util.GetDisplayNode(node))
    return widget


def GetDisplayNode(node):
    if node is None:
        return None
    displayNode = node.GetDisplayNode()
    if displayNode is None:
        node.CreateDefaultDisplayNodes()
        displayNode = node.GetDisplayNode()
    return displayNode


def harden_transform(model, matrix_arr):
    import vtk, slicer

    matrix = vtk.vtkMatrix4x4()
    for i in range(0, 4):
        for j in range(0, 4):
            matrix.SetElement(i, j, matrix_arr[i * 4 + j])
    TNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLLinearTransformNode")
    TNode.SetAndObserveMatrixTransformToParent(
        matrix
    )
    model.SetAndObserveTransformNodeID(TNode.GetID())
    model.HardenTransform()
    RemoveNode(TNode)


def ShowModel2D(node, lineThickness):
    node.GetDisplayNode().SetVisibility2D(True)
    node.GetDisplayNode().SetSliceIntersectionThickness(lineThickness)


def ShowNode(node, is_show=True):
    if node is None:
        return
    displayNode = GetDisplayNode(node)
    if displayNode:
        displayNode.SetVisibility(is_show)
    if node.IsA("vtkMRMLSegmentationNode"):
        segmentid = GetNthSegmentID(node, 0)
        displayNode.SetSegmentVisibility(segmentid, is_show)


def AddNode(node):
    if node is None:
        return
    import slicer
    exist_id = node.GetID()
    if type(exist_id) is str and GetNodeByID(exist_id) is not None:
        return
    slicer.mrmlScene.AddNode(node)


def RemoveNodeByName(nodename):
    import slicer
    node = getFirstNodeByName(nodename)
    if node is None:
        return
    RemoveNode(node)


def distance_between_list(list1, list2):
    import math
    val = 0
    for i in range(len(list1)):
        val = val + (list1[i] - list2[i]) * (list1[i] - list2[i])
    val = math.sqrt(val)
    return val


def RemoveNode(node):
    import slicer
    if node is None:
        return
    slicer.mrmlScene.RemoveNode(node)


def RemoveNodes(nodes):
    for node in nodes:
        RemoveNode(node)


def convert_markups_curve_to_model(markups_curve_node, model_node_name):
    import slicer, vtk
    # Check if the input node is a vtkMRMLMarkupsCurveNode
    if not markups_curve_node.IsA('vtkMRMLMarkupsCurveNode'):
        print("Error: The input node is not a vtkMRMLMarkupsCurveNode.")
        return None

    # Get the markups curve points
    points = vtk.vtkPoints()
    num_points = markups_curve_node.GetNumberOfControlPoints()
    for i in range(num_points):
        point = [0, 0, 0]
        markups_curve_node.GetNthControlPointPositionWorld(i, point)
        points.InsertNextPoint(point)

    # Create a vtkPolyData with the points
    poly_data = vtk.vtkPolyData()
    poly_data.SetPoints(points)

    # Create a vtkCellArray to define the curve
    lines = vtk.vtkCellArray()
    lines.InsertNextCell(num_points)
    for i in range(num_points):
        lines.InsertCellPoint(i)

    poly_data.SetLines(lines)

    model_node = getFirstNodeByClassByName(vtkMRMLModelNode, model_node_name)
    if model_node:
        RemoveNode(model_node)
    # Create a model node and add the poly data to it
    model_node = slicer.mrmlScene.AddNewNodeByClass('vtkMRMLModelNode', model_node_name)
    model_node.SetAndObservePolyData(poly_data)
    ShowNode(model_node)
    GetDisplayNode(model_node).SetColor(1, 1, 0)
    return model_node


def RemoveConfig(key):
    import slicer
    saveConfig = slicer.mrmlScene.GetFirstNodeByName("saveconfig")
    saveConfig.RemoveAttribute(key)


def combineModelList(modellist, output_node):
    import vtk, slicer, os, shutil, qt
    appendFilter = vtk.vtkAppendPolyData()
    for model in modellist:
        vtkpolydata = vtk.vtkPolyData()
        vtkpolydata.DeepCopy(model.GetPolyData())
        appendFilter.AddInputData(vtkpolydata)
    appendFilter.Update()
    clean3 = vtk.vtkCleanPolyData()
    clean3.SetTolerance(0.001)
    clean3.SetInputData(appendFilter.GetOutput())
    clean3.Update()
    # 不能加normalfilter,会删除模型数据
    output_node.SetAndObservePolyData(clean3.GetOutput())
    RemoveNode(output_node)
    AddNode(output_node)
    tmp_file = os.path.join(slicer.app.slicerHome, "mtmpjoin/tmp.stl").replace('\\', '/')
    saveNode(output_node, tmp_file)
    node_tmp = loadModel(tmp_file)
    import shutil
    shutil.rmtree(os.path.join(slicer.app.slicerHome, "mtmpjoin").replace('\\', '/'))
    output_node.SetAndObservePolyData(node_tmp.GetPolyData())
    RemoveNode(node_tmp)
    return output_node


def ShowProgressBar(title="请稍候"):
    send_event_str(ProgressStart, title)


def SetProgress(val):
    send_event_str(ProgressValue, val.__str__())


def SetProgressCounts(val):
    mainWindow().SetProgressCounts(val)


def SetProgressStep(val):
    mainWindow().SetProgressStep(val)


def UpdateProgress():
    mainWindow().UpdateProgress()


def combineModel(model1, model2):
    import slicer.util as util
    import vtk, slicer

    clean1 = vtk.vtkCleanPolyData()
    clean1.SetInputData(model1.GetPolyData())
    clean1.Update()

    clean2 = vtk.vtkCleanPolyData()
    clean2.SetInputData(model2.GetPolyData())
    clean2.Update()

    model3 = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLModelNode")
    model3.SetName("output")

    appendFilter = vtk.vtkAppendPolyData()
    appendFilter.AddInputData(clean1.GetOutput())
    appendFilter.AddInputData(clean2.GetOutput())
    appendFilter.Update()

    clean3 = vtk.vtkCleanPolyData()
    clean3.SetInputData(appendFilter.GetOutput())
    clean3.Update()

    connectivityFilter = vtk.vtkConnectivityFilter()
    connectivityFilter.SetInputData(clean3.GetOutput())
    connectivityFilter.SetExtractionModeToLargestRegion()
    connectivityFilter.Update()

    model3.SetAndObservePolyData(connectivityFilter.GetOutput())
    return model3


def largestRegion(model3):
    import vtk
    connectivityFilter = vtk.vtkPolyDataConnectivityFilter()
    connectivityFilter.ScalarConnectivityOff()
    connectivityFilter.SetInputData(model3.GetPolyData())
    connectivityFilter.SetExtractionModeToLargestRegion()
    connectivityFilter.Update()
    model3.SetAndObservePolyData(connectivityFilter.GetOutput())


def AddNewNodeByNameByClass(nodename, classname="vtkMRMLScalarVolumeNode"):
    import slicer
    referenceVolumeNode = slicer.mrmlScene.AddNewNodeByClass(classname)
    referenceVolumeNode.SetName(nodename)
    return referenceVolumeNode


def AddNewNodeByClass(classname="vtkMRMLScalarVolumeNode"):
    import slicer
    referenceVolumeNode = slicer.mrmlScene.AddNewNodeByClass(classname)
    return referenceVolumeNode


def CreateNodeByClass(classname="vtkMRMLScalarVolumeNode"):
    import slicer
    referenceVolumeNode = slicer.mrmlScene.CreateNodeByClass(classname)
    return referenceVolumeNode


def CreateNodeByNameByClass(nodename, classname="vtkMRMLScalarVolumeNode"):
    import slicer
    referenceVolumeNode = slicer.mrmlScene.CreateNodeByClass(classname)
    referenceVolumeNode.SetName(nodename)
    return referenceVolumeNode


class NodeModify(object):
    """Context manager to conveniently compress mrml node modified event."""

    def __init__(self, node):
        self.node = node

    def __enter__(self):
        self.wasModifying = self.node.StartModify()
        return self.node

    def __exit__(self, type, value, traceback):
        self.node.EndModify(self.wasModifying)


def GetGlobalSaveValue(key):
    import slicer
    node = slicer.mrmlScene.GetFirstNodeByName("saveconfig")
    if node is None:
        return None
    val = node.GetAttribute(key)
    return val


def SetGlobalSaveValue(key, value):
    import slicer
    node = slicer.mrmlScene.GetFirstNodeByName("saveconfig")
    if node is None:
        node = CreateNodeByClass("vtkMRMLModelNode")
        node.SetName("saveconfig")
        node.SetAttribute("hide_in_manager", "1")
        AddNode(node)
    node.SetAttribute(key, value.__str__())


def showVolumeRendering(volumeNode):
    import slicer
    print("Show volume rendering of node " + volumeNode.GetName())
    volRenLogic = slicer.modules.volumerendering.logic()
    displayNode = volRenLogic.CreateDefaultVolumeRenderingNodes(volumeNode)
    displayNode.SetVisibility(True)
    scalarRange = volumeNode.GetImageData().GetScalarRange()
    if scalarRange[1] - scalarRange[0] < 1500:
        # Small dynamic range, probably MRI
        displayNode.GetVolumePropertyNode().Copy(volRenLogic.GetPresetByName("MR-Default"))
    else:
        # Larger dynamic range, probably CT
        displayNode.GetVolumePropertyNode().Copy(volRenLogic.GetPresetByName("CT-AAA"))


#
# Subject hierarchy
#
def getSubjectHierarchyItemChildren(parentItem=None, recursive=False):
    """Convenience method to get children of a subject hierarchy item.

  :param vtkIdType parentItem: Item for which to get children for. If omitted
         or None then use scene item (i.e. get all items)
  :param bool recursive: Whether the query is recursive. False by default
  :return: List of child item IDs
  """
    children = []
    shNode = slicer.mrmlScene.GetSubjectHierarchyNode()
    # Use scene as parent item if not given
    if not parentItem:
        parentItem = shNode.GetSceneItemID()
    childrenIdList = vtk.vtkIdList()
    shNode.GetItemChildren(parentItem, childrenIdList, recursive)
    for childIndex in range(childrenIdList.GetNumberOfIds()):
        children.append(childrenIdList.GetId(childIndex))
    return children


def isRelease():
    return mainWindow().isRelease()


#
# MRML-numpy
#

def array(pattern="", index=0):
    """Return the array you are "most likely to want" from the indexth

  MRML node that matches the pattern.

  :raises RuntimeError: if the node cannot be accessed as an array.

  .. warning::

    Meant to be used in the python console for quick debugging/testing.

  More specific API should be used in scripts to be sure you get exactly
  what you want, such as :py:meth:`arrayFromVolume`, :py:meth:`arrayFromModelPoints`,
  and :py:meth:`arrayFromGridTransform`.
  """
    node = getNode(pattern=pattern, index=index)
    import slicer
    if isinstance(node, slicer.vtkMRMLVolumeNode):
        return arrayFromVolume(node)
    elif isinstance(node, slicer.vtkMRMLModelNode):
        return arrayFromModelPoints(node)
    elif isinstance(node, slicer.vtkMRMLGridTransformNode):
        return arrayFromGridTransform(node)
    elif isinstance(node, slicer.vtkMRMLMarkupsNode):
        return arrayFromMarkupsControlPoints(node)
    elif isinstance(node, slicer.vtkMRMLTransformNode):
        return arrayFromTransformMatrix(node)

    # TODO: accessors for other node types: polydata (verts, polys...), colors
    raise RuntimeError("Cannot get node " + node.GetID() + " as array")


def multivolume_to_scalarvolume_array(mvNode):
    import vtk
    if not mvNode:
        return
    nodes = getNodesByClass("vtkMRMLScalarVolumeNode")
    for node in nodes:
        if node.GetAttribute("MultiVolumeTag") == mvNode.GetName():
            RemoveNode(node)
    imagedata = mvNode.GetImageData()
    if not imagedata:
        return
    extent = imagedata.GetExtent()
    nparray = vtk.util.numpy_support.vtk_to_numpy(imagedata.GetPointData().GetScalars()).T
    nFrames = mvNode.GetNumberOfFrames()
    list1 = []
    for index in range(nFrames):
        numpy1 = nparray[index]
        numpy1 = numpy1.reshape(extent[5] + 1, extent[3] + 1, extent[1] + 1)
        volumenode = addVolumeFromArray(numpy1)
        volumenode.SetAttribute("MultiVolumeTag", mvNode.GetName())
        volumenode.SetAttribute("MultiVolumeOriginID", mvNode.GetID())
        volumenode.CopyOrientation(mvNode)
        volumenode.SetName(mvNode.GetName() + "-" + index.__str__())
        list1.append(volumenode)
    return list1


def arrayFromVolume(volumeNode):
    """Return voxel array from volume node as numpy array.

  Voxels values are not copied. Voxel values in the volume node can be modified
  by changing values in the numpy array.
  After all modifications has been completed, call :py:meth:`arrayFromVolumeModified`.

  :raises RuntimeError: in case of failure

  .. warning:: Memory area of the returned array is managed by VTK, therefore
    values in the array may be changed, but the array must not be reallocated
    (change array size, shallow-copy content from other array most likely causes
    application crash). To allow arbitrary numpy operations on a volume array:

      1. Make a deep-copy of the returned VTK-managed array using :func:`numpy.copy`.
      2. Perform any computations using the copied array.
      3. Write results back to the image data using :py:meth:`updateVolumeFromArray`.
  """
    scalarTypes = ['vtkMRMLScalarVolumeNode', 'vtkMRMLLabelMapVolumeNode']
    vectorTypes = ['vtkMRMLVectorVolumeNode', 'vtkMRMLMultiVolumeNode', 'vtkMRMLDiffusionWeightedVolumeNode']
    tensorTypes = ['vtkMRMLDiffusionTensorVolumeNode']
    vimage = volumeNode.GetImageData()
    nshape = tuple(reversed(volumeNode.GetImageData().GetDimensions()))
    import vtk.util.numpy_support
    narray = None
    if volumeNode.GetClassName() in scalarTypes:
        narray = vtk.util.numpy_support.vtk_to_numpy(vimage.GetPointData().GetScalars()).reshape(nshape)
    elif volumeNode.GetClassName() in vectorTypes:
        components = vimage.GetNumberOfScalarComponents()
        if components > 1:
            nshape = nshape + (components,)
        narray = vtk.util.numpy_support.vtk_to_numpy(vimage.GetPointData().GetScalars()).reshape(nshape)
    elif volumeNode.GetClassName() in tensorTypes:
        narray = vtk.util.numpy_support.vtk_to_numpy(vimage.GetPointData().GetTensors()).reshape(nshape + (3, 3))
    else:
        raise RuntimeError("Unsupported volume type: " + volumeNode.GetClassName())
    return narray


def arrayFromVolumeModified(volumeNode):
    """Indicate that modification of a numpy array returned by :py:meth:`arrayFromVolume` has been completed."""
    imageData = volumeNode.GetImageData()
    pointData = imageData.GetPointData() if imageData else None
    if pointData:
        if pointData.GetScalars():
            pointData.GetScalars().Modified()
        if pointData.GetTensors():
            pointData.GetTensors().Modified()
    volumeNode.Modified()


def arrayFromModelPoints(modelNode):
    """Return point positions of a model node as numpy array.

  Point coordinates can be modified by modifying the numpy array.
  After all modifications has been completed, call :py:meth:`arrayFromModelPointsModified`.

  .. warning:: Important: memory area of the returned array is managed by VTK,
    therefore values in the array may be changed, but the array must not be reallocated.
    See :py:meth:`arrayFromVolume` for details.
  """
    import vtk.util.numpy_support
    pointData = modelNode.GetPolyData().GetPoints().GetData()
    narray = vtk.util.numpy_support.vtk_to_numpy(pointData)
    return narray


def arrayFromModelPointsModified(modelNode):
    """Indicate that modification of a numpy array returned by :py:meth:`arrayFromModelPoints` has been completed."""
    if modelNode.GetPolyData():
        modelNode.GetPolyData().GetPoints().GetData().Modified()
    # Trigger re-render
    modelNode.GetDisplayNode().Modified()


def _vtkArrayFromModelData(modelNode, arrayName, location):
    """Helper function for getting VTK point data array that throws exception

  with informative error message if the data array is not found.
  Point or cell data can be selected by setting 'location' argument to 'point' or 'cell'.

  :raises ValueError: in case of failure
  """
    if location == 'point':
        modelData = modelNode.GetMesh().GetPointData()
    elif location == 'cell':
        modelData = modelNode.GetMesh().GetCellData()
    else:
        raise ValueError("Location attribute must be set to 'point' or 'cell'")
    if not modelData or modelData.GetNumberOfArrays() == 0:
        raise ValueError(f"Input modelNode does not contain {location} data")
    arrayVtk = modelData.GetArray(arrayName)
    if not arrayVtk:
        availableArrayNames = [modelData.GetArrayName(i) for i in range(modelData.GetNumberOfArrays())]
        raise ValueError("Input modelNode does not contain {0} data array '{1}'. Available array names: '{2}'".format(
            location, arrayName, "', '".join(availableArrayNames)))
    return arrayVtk


def arrayFromModelPointData(modelNode, arrayName):
    """Return point data array of a model node as numpy array.

  .. warning:: Important: memory area of the returned array is managed by VTK,
    therefore values in the array may be changed, but the array must not be reallocated.
    See :py:meth:`arrayFromVolume` for details.
  """
    import vtk.util.numpy_support
    arrayVtk = _vtkArrayFromModelData(modelNode, arrayName, 'point')
    narray = vtk.util.numpy_support.vtk_to_numpy(arrayVtk)
    return narray


def arrayFromModelPointDataModified(modelNode, arrayName):
    """Indicate that modification of a numpy array returned by :py:meth:`arrayFromModelPointData` has been completed."""
    arrayVtk = _vtkArrayFromModelData(modelNode, arrayName, 'point')
    arrayVtk.Modified()


def arrayFromModelCellData(modelNode, arrayName):
    """Return cell data array of a model node as numpy array.

  .. warning:: Important: memory area of the returned array is managed by VTK,
    therefore values in the array may be changed, but the array must not be reallocated.
    See :py:meth:`arrayFromVolume` for details.
  """
    import vtk.util.numpy_support
    arrayVtk = _vtkArrayFromModelData(modelNode, arrayName, 'cell')
    narray = vtk.util.numpy_support.vtk_to_numpy(arrayVtk)
    return narray


def arrayFromModelCellDataModified(modelNode, arrayName):
    """Indicate that modification of a numpy array returned by :py:meth:`arrayFromModelCellData` has been completed."""
    arrayVtk = _vtkArrayFromModelData(modelNode, arrayName, 'cell')
    arrayVtk.Modified()


def arrayFromMarkupsControlPointData(markupsNode, arrayName):
    """Return control point data array of a markups node as numpy array.

  .. warning:: Important: memory area of the returned array is managed by VTK,
    therefore values in the array may be changed, but the array must not be reallocated.
    See :py:meth:`arrayFromVolume` for details.
  """
    import vtk.util.numpy_support
    for measurementIndex in range(markupsNode.GetNumberOfMeasurements()):
        measurement = markupsNode.GetNthMeasurement(measurementIndex)
        doubleArrayVtk = measurement.GetControlPointValues()
        if doubleArrayVtk and doubleArrayVtk.GetName() == arrayName:
            narray = vtk.util.numpy_support.vtk_to_numpy(doubleArrayVtk)
            return narray


def arrayFromMarkupsControlPointDataModified(markupsNode, arrayName):
    """Indicate that modification of a numpy array returned by :py:meth:`arrayFromMarkupsControlPointData` has been completed."""
    for measurementIndex in range(markupsNode.GetNumberOfMeasurements()):
        measurement = markupsNode.GetNthMeasurement(measurementIndex)
        doubleArrayVtk = measurement.GetControlPointValues()
        if doubleArrayVtk and doubleArrayVtk.GetName() == arrayName:
            doubleArrayVtk.Modified()


def arrayFromModelPolyIds(modelNode):
    """Return poly id array of a model node as numpy array.

  These ids are the following format:
  [ n(0), i(0,0), i(0,1), ... i(0,n(00),..., n(j), i(j,0), ... i(j,n(j))...]
  where n(j) is the number of vertices in polygon j
  and i(j,k) is the index into the vertex array for vertex k of poly j.

  As described here:
  https://vtk.org/wp-content/uploads/2015/04/file-formats.pdf

  Typically in Slicer n(j) will always be 3 because a model node's
  polygons will be triangles.

  .. warning:: Important: memory area of the returned array is managed by VTK,
    therefore values in the array may be changed, but the array must not be reallocated.
    See :py:meth:`arrayFromVolume` for details.
  """
    import vtk.util.numpy_support
    arrayVtk = modelNode.GetPolyData().GetPolys().GetData()
    narray = vtk.util.numpy_support.vtk_to_numpy(arrayVtk)
    return narray


def arrayFromGridTransform(gridTransformNode):
    """Return voxel array from transform node as numpy array.

  Vector values are not copied. Values in the transform node can be modified
  by changing values in the numpy array.
  After all modifications has been completed, call :py:meth:`arrayFromGridTransformModified`.

  .. warning:: Important: memory area of the returned array is managed by VTK,
    therefore values in the array may be changed, but the array must not be reallocated.
    See :py:meth:`arrayFromVolume` for details.
  """
    transformGrid = gridTransformNode.GetTransformFromParent()
    displacementGrid = transformGrid.GetDisplacementGrid()
    nshape = tuple(reversed(displacementGrid.GetDimensions()))
    import vtk.util.numpy_support
    nshape = nshape + (3,)
    narray = vtk.util.numpy_support.vtk_to_numpy(displacementGrid.GetPointData().GetScalars()).reshape(nshape)
    return narray


def arrayFromVTKMatrix(vmatrix):
    """Return vtkMatrix4x4 or vtkMatrix3x3 elements as numpy array.

  :raises RuntimeError: in case of failure

  The returned array is just a copy and so any modification in the array will not affect the input matrix.
  To set VTK matrix from a numpy array, use :py:meth:`vtkMatrixFromArray` or
  :py:meth:`updateVTKMatrixFromArray`.
  """
    from vtk import vtkMatrix4x4
    from vtk import vtkMatrix3x3
    import numpy as np
    if isinstance(vmatrix, vtkMatrix4x4):
        matrixSize = 4
    elif isinstance(vmatrix, vtkMatrix3x3):
        matrixSize = 3
    else:
        raise RuntimeError("Input must be vtk.vtkMatrix3x3 or vtk.vtkMatrix4x4")
    narray = np.eye(matrixSize)
    vmatrix.DeepCopy(narray.ravel(), vmatrix)
    return narray


def vtkMatrixFromArray(narray):
    """Create VTK matrix from a 3x3 or 4x4 numpy array.

  :param narray: input numpy array
  :raises RuntimeError: in case of failure

  The returned matrix is just a copy and so any modification in the array will not affect the output matrix.
  To set numpy array from VTK matrix, use :py:meth:`arrayFromVTKMatrix`.
  """
    from vtk import vtkMatrix4x4
    from vtk import vtkMatrix3x3
    narrayshape = narray.shape
    if narrayshape == (4, 4):
        vmatrix = vtkMatrix4x4()
        updateVTKMatrixFromArray(vmatrix, narray)
        return vmatrix
    elif narrayshape == (3, 3):
        vmatrix = vtkMatrix3x3()
        updateVTKMatrixFromArray(vmatrix, narray)
        return vmatrix
    else:
        raise RuntimeError("Unsupported numpy array shape: " + str(narrayshape) + " expected (4,4)")


def updateVTKMatrixFromArray(vmatrix, narray):
    """Update VTK matrix values from a numpy array.

  :param vmatrix: VTK matrix (vtkMatrix4x4 or vtkMatrix3x3) that will be update
  :param narray: input numpy array
  :raises RuntimeError: in case of failure

  To set numpy array from VTK matrix, use :py:meth:`arrayFromVTKMatrix`.
  """
    from vtk import vtkMatrix4x4
    from vtk import vtkMatrix3x3
    if isinstance(vmatrix, vtkMatrix4x4):
        matrixSize = 4
    elif isinstance(vmatrix, vtkMatrix3x3):
        matrixSize = 3
    else:
        raise RuntimeError("Output vmatrix must be vtk.vtkMatrix3x3 or vtk.vtkMatrix4x4")
    if narray.shape != (matrixSize, matrixSize):
        raise RuntimeError("Input narray size must match output vmatrix size ({0}x{0})".format(matrixSize))
    vmatrix.DeepCopy(narray.ravel())


def arrayFromTransformMatrix(transformNode, toWorld=False):
    """Return 4x4 transformation matrix as numpy array.

  :param toWorld: if set to True then the transform to world coordinate system is returned
    (effect of parent transform to the node is applied), otherwise transform to parent transform is returned.
  :return: numpy array
  :raises RuntimeError: in case of failure

  The returned array is just a copy and so any modification in the array will not affect the transform node.

  To set transformation matrix from a numpy array, use :py:meth:`updateTransformMatrixFromArray`.
  """
    import numpy as np
    from vtk import vtkMatrix4x4
    vmatrix = vtkMatrix4x4()
    if toWorld:
        success = transformNode.GetMatrixTransformToWorld(vmatrix)
    else:
        success = transformNode.GetMatrixTransformToParent(vmatrix)
    if not success:
        raise RuntimeError("Failed to get transformation matrix from node " + transformNode.GetID())
    return arrayFromVTKMatrix(vmatrix)


def updateTransformMatrixFromArray(transformNode, narray, toWorld=False):
    """Set transformation matrix from a numpy array of size 4x4 (toParent).

  :param world: if set to True then the transform will be set so that transform
    to world matrix will be equal to narray; otherwise transform to parent will be
    set as narray.
  :raises RuntimeError: in case of failure
  """
    import numpy as np
    from vtk import vtkMatrix4x4
    narrayshape = narray.shape
    if narrayshape != (4, 4):
        raise RuntimeError("Unsupported numpy array shape: " + str(narrayshape) + " expected (4,4)")
    if toWorld and transformNode.GetParentTransformNode():
        # thisToParent = worldToParent * thisToWorld = inv(parentToWorld) * toWorld
        narrayParentToWorld = arrayFromTransformMatrix(transformNode.GetParentTransformNode())
        thisToParent = np.dot(np.linalg.inv(narrayParentToWorld), narray)
        updateTransformMatrixFromArray(transformNode, thisToParent, toWorld=False)
    else:
        vmatrix = vtkMatrix4x4()
        updateVTKMatrixFromArray(vmatrix, narray)
        transformNode.SetMatrixTransformToParent(vmatrix)


def arrayFromGridTransformModified(gridTransformNode):
    """Indicate that modification of a numpy array returned by :py:meth:`arrayFromGridTransform` has been completed."""
    transformGrid = gridTransformNode.GetTransformFromParent()
    displacementGrid = transformGrid.GetDisplacementGrid()
    displacementGrid.GetPointData().GetScalars().Modified()
    displacementGrid.Modified()


def GetNthSegment(segmentationNode, i):
    if not segmentationNode:
        return None
    segmentation = segmentationNode.GetSegmentation()
    if not segmentation:
        return None
    return segmentation.GetNthSegment(i)


def GetControlPointNumber(node):
    if node is None:
        return 0
    return node.GetNumberOfControlPoints()


def RemoveNthControlPoint(node, n):
    node.RemoveNthControlPoint(n)


def GetSliceNormal(slicename):
    import slicer
    sliceWidget = slicer.app.layoutManager().sliceWidget(slicename)
    node = sliceWidget.mrmlSliceNode()
    sliceToRAS = node.GetSliceToRAS()
    sliceNormalVector = []
    sliceNormalVector.append([sliceToRAS.GetElement(0, 2), sliceToRAS.GetElement(1, 2), sliceToRAS.GetElement(2, 2)])
    return sliceNormalVector[0]


def GetNthControlPointPosition(node, n):
    world = [0.0, 0.0, 0.0]
    node.GetNthControlPointPosition(n, world)
    return world


def GetSegmentNumber(segmentationNode):
    if not segmentationNode:
        return None
    return segmentationNode.GetSegmentation().GetNumberOfSegments()


def GetNthSegmentID(segmentationNode, i):
    if not segmentationNode:
        return None
    segmentation = segmentationNode.GetSegmentation()
    if not segmentation:
        return None
    return segmentation.GetNthSegmentID(i)


def RemoveNthSegmentIDSoft(node, i):
    sid = GetNthSegmentID(node, i)
    vimage = node.GetBinaryLabelmapInternalRepresentation(sid)
    if vimage is None or vimage.GetPointData() is None or vimage.GetPointData().GetScalars() is None:
        return
    array = arrayFromSegmentInternalBinaryLabelmap(node, sid)
    array.fill(0)
    GetNthSegment(node, i).Modified()
    node.RemoveClosedSurfaceRepresentation()
    node.CreateClosedSurfaceRepresentation()


def RemoveNthSegmentID(segmentationNode, i):
    if not segmentationNode:
        return None
    segmentation = segmentationNode.GetSegmentation()
    if not segmentation:
        return None
    sid = segmentation.GetNthSegmentID(i)
    segmentation.RemoveSegment(sid)


def AddSegment(segment_node):
    import slicer
    segment = slicer.vtkSegment()
    segment.SetName("layer_1")
    segment.SetColor(0.9, 0.9, 0)
    segment_node.GetSegmentation().AddSegment(segment)
    segment_node.CreateDefaultDisplayNodes()


def arrayFromSegment(segmentationNode, segmentId):
    """Get segment as numpy array.

  .. warning:: Important: binary labelmap representation may be shared between multiple segments.

  .. deprecated:: 4.13.0
    Use arrayFromSegmentBinaryLabelmap to access a copy of the binary labelmap that will not modify the original labelmap."
    Use arrayFromSegmentInternalBinaryLabelmap to access a modifiable internal lablemap representation that may be shared"
    between multiple segments.
  """
    import logging
    logging.warning(
        "arrayFromSegment is deprecated. Binary labelmap representation may be shared between multiple segments.")
    return arrayFromSegmentBinaryLabelmap(segmentationNode, segmentId)


def arrayFromSegmentInternalBinaryLabelmap(segmentationNode, segmentId):
    """Return voxel array of a segment's binary labelmap representation as numpy array.

  Voxels values are not copied.
  The labelmap containing the specified segment may be a shared labelmap containing multiple segments.

  To get and modify the array for a single segment, calling::

    segmentationNode->GetSegmentation()->SeparateSegment(segmentId)

  will transfer the segment from a shared labelmap into a new layer.

  Layers can be merged by calling::

    segmentationNode->GetSegmentation()->CollapseBinaryLabelmaps()

  If binary labelmap is the master representation then voxel values in the volume node can be modified
  by changing values in the numpy array. After all modifications has been completed, call::

    segmentationNode.GetSegmentation().GetSegment(segmentID).Modified()

  .. warning:: Important: memory area of the returned array is managed by VTK,
    therefore values in the array may be changed, but the array must not be reallocated.
    See :py:meth:`arrayFromVolume` for details.
  """
    vimage = segmentationNode.GetBinaryLabelmapInternalRepresentation(segmentId)
    nshape = tuple(reversed(vimage.GetDimensions()))
    import vtk.util.numpy_support
    narray = vtk.util.numpy_support.vtk_to_numpy(vimage.GetPointData().GetScalars()).reshape(nshape)
    return narray


def arrayFromSegmentBinaryLabelmap(segmentationNode, segmentId):
    """Return voxel array of a segment's binary labelmap representation as numpy array.

  Voxels values are copied.

  If binary labelmap is the master representation then voxel values in the volume node can be modified
  by changing values in the numpy array.

  After all modifications have been completed, call::

    segmentationNode.GetSegmentation().GetSegment(segmentID).Modified()

  .. warning:: Important: memory area of the returned array is managed by VTK,
    therefore values in the array may be changed, but the array must not be reallocated.
    See :py:meth:`arrayFromVolume` for details.
  """
    import slicer
    vimage = slicer.vtkOrientedImageData()
    segmentationNode.GetBinaryLabelmapRepresentation(segmentId, vimage)
    nshape = tuple(reversed(vimage.GetDimensions()))
    import vtk.util.numpy_support
    narray = vtk.util.numpy_support.vtk_to_numpy(vimage.GetPointData().GetScalars()).reshape(nshape)
    return narray


def arrayFromMarkupsControlPoints(markupsNode, world=False):
    """Return control point positions of a markups node as rows in a numpy array (of size Nx3).

  :param world: if set to True then the control points coordinates are returned in world coordinate system
    (effect of parent transform to the node is applied).

  The returned array is just a copy and so any modification in the array will not affect the markup node.

  To modify markup control points based on a numpy array, use :py:meth:`updateMarkupsControlPointsFromArray`.
  """
    numberOfControlPoints = markupsNode.GetNumberOfControlPoints()
    import numpy as np
    narray = np.zeros([numberOfControlPoints, 3])
    for controlPointIndex in range(numberOfControlPoints):
        if world:
            markupsNode.GetNthControlPointPositionWorld(controlPointIndex, narray[controlPointIndex, :])
        else:
            markupsNode.GetNthControlPointPosition(controlPointIndex, narray[controlPointIndex, :])
    return narray


def updateMarkupsControlPointsFromArray(markupsNode, narray, world=False):
    """Sets control point positions in a markups node from a numpy array of size Nx3.

  :param world: if set to True then the control point coordinates are expected in world coordinate system.
  :raises RuntimeError: in case of failure

  All previous content of the node is deleted.
  """
    narrayshape = narray.shape
    if narrayshape == (0,):
        markupsNode.RemoveAllControlPoints()
        return
    if len(narrayshape) != 2 or narrayshape[1] != 3:
        raise RuntimeError("Unsupported numpy array shape: " + str(narrayshape) + " expected (N,3)")
    numberOfControlPoints = narrayshape[0]
    oldNumberOfControlPoints = markupsNode.GetNumberOfControlPoints()
    # Update existing control points
    for controlPointIndex in range(min(numberOfControlPoints, oldNumberOfControlPoints)):
        if world:
            markupsNode.SetNthControlPointPositionWorldFromArray(controlPointIndex, narray[controlPointIndex, :])
        else:
            markupsNode.SetNthControlPointPositionFromArray(controlPointIndex, narray[controlPointIndex, :])
    if numberOfControlPoints >= oldNumberOfControlPoints:
        # Add new points to the markup node
        from vtk import vtkVector3d
        for controlPointIndex in range(oldNumberOfControlPoints, numberOfControlPoints):
            if world:
                markupsNode.AddControlPointWorld(vtkVector3d(narray[controlPointIndex, :]))
            else:
                markupsNode.AddControlPoint(vtkVector3d(narray[controlPointIndex, :]))
    else:
        # Remove extra point from the markup node
        for controlPointIndex in range(oldNumberOfControlPoints, numberOfControlPoints, -1):
            markupsNode.RemoveNthControlPoint(controlPointIndex - 1)


def arrayFromMarkupsCurvePoints(markupsNode, world=False):
    """Return interpolated curve point positions of a markups node as rows in a numpy array (of size Nx3).

  :param world: if set to True then the point coordinates are returned in world coordinate system
    (effect of parent transform to the node is applied).

  The returned array is just a copy and so any modification in the array will not affect the markup node.
  """
    import numpy as np
    import vtk.util.numpy_support
    if world:
        pointData = markupsNode.GetCurvePointsWorld().GetData()
    else:
        pointData = markupsNode.GetCurvePoints().GetData()
    narray = vtk.util.numpy_support.vtk_to_numpy(pointData)
    return narray


def updateSegmentBinaryLabelmapFromArray(narray, segmentationNode, segmentId, referenceVolumeNode=None):
    """Sets binary labelmap representation of a segment from a numpy array.
  :param segmentationNode: segmentation node that will be updated.
  :param segmentId: ID of the segment that will be updated.
    Can be determined from segment name by calling ``segmentationNode.GetSegmentation().GetSegmentIdBySegmentName(segmentName)``.
  :param referenceVolumeNode: a volume node that determines geometry (origin, spacing, axis directions, extents) of the array.
    If not specified then the volume that was used for setting the segmentation's geometry is used as reference volume.
  :raises RuntimeError: in case of failure
  Voxels values are deep-copied, therefore if the numpy array is modified after calling this method, segmentation node will not change.
  """

    # Export segment as vtkImageData (via temporary labelmap volume node)
    import slicer
    import vtk

    # Get reference volume
    if not referenceVolumeNode:
        referenceVolumeNode = segmentationNode.GetNodeReference(
            slicer.vtkMRMLSegmentationNode.GetReferenceImageGeometryReferenceRole())
        if not referenceVolumeNode:
            raise RuntimeError(
                "No reference volume is found in the input segmentationNode, therefore a valid referenceVolumeNode input is required.")

    # Update segment in segmentation
    labelmapVolumeNode = slicer.modules.volumes.logic().CreateAndAddLabelVolume(referenceVolumeNode, "__temp__")
    try:
        updateVolumeFromArray(labelmapVolumeNode, narray)
        segmentIds = vtk.vtkStringArray()
        segmentIds.InsertNextValue(segmentId)
        if not slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(labelmapVolumeNode,
                                                                                     segmentationNode, segmentIds):
            raise RuntimeError("Importing of segment failed.")
    finally:
        slicer.mrmlScene.RemoveNode(labelmapVolumeNode)


def updateVolumeFromArray(volumeNode, narray):
    """Sets voxels of a volume node from a numpy array.

  :raises RuntimeError: in case of failure

  Voxels values are deep-copied, therefore if the numpy array
  is modified after calling this method, voxel values in the volume node will not change.
  Dimensions and data size of the source numpy array does not have to match the current
  content of the volume node.
  """

    vshape = tuple(reversed(narray.shape))
    if len(vshape) == 3:
        # Scalar volume
        vcomponents = 1
    elif len(vshape) == 4:
        # Vector volume
        vcomponents = vshape[0]
        vshape = vshape[1:4]
    else:
        # TODO: add support for tensor volumes
        raise RuntimeError("Unsupported numpy array shape: " + str(narray.shape))

    vimage = volumeNode.GetImageData()
    if not vimage:
        import vtk
        vimage = vtk.vtkImageData()
        volumeNode.SetAndObserveImageData(vimage)
    import vtk.util.numpy_support
    vtype = vtk.util.numpy_support.get_vtk_array_type(narray.dtype)

    # Volumes with "long long" scalar type are not rendered corectly.
    # Probably this could be fixed in VTK or Slicer but for now just reject it.
    if vtype == vtk.VTK_LONG_LONG:
        raise RuntimeError("Unsupported numpy array type: long long")

    vimage.SetDimensions(vshape)
    vimage.AllocateScalars(vtype, vcomponents)

    narrayTarget = arrayFromVolume(volumeNode)
    narrayTarget[:] = narray

    # Notify the application that image data is changed
    # (same notifications as in vtkMRMLVolumeNode.SetImageDataConnection)
    import slicer
    volumeNode.StorableModified()
    volumeNode.Modified()
    volumeNode.InvokeEvent(slicer.vtkMRMLVolumeNode.ImageDataModifiedEvent, volumeNode)


def addVolumeFromArray(narray, ijkToRAS=None, name=None, nodeClassName=None, hide=False):
    """Create a new volume node from content of a numpy array and add it to the scene.

  Voxels values are deep-copied, therefore if the numpy array
  is modified after calling this method, voxel values in the volume node will not change.

  :param narray: numpy array containing volume voxels.
  :param ijkToRAS: 4x4 numpy array or vtk.vtkMatrix4x4 that defines mapping from IJK to RAS coordinate system (specifying origin, spacing, directions)
  :param name: volume node name
  :param nodeClassName: type of created volume, default: ``vtkMRMLScalarVolumeNode``.
    Use ``vtkMRMLLabelMapVolumeNode`` for labelmap volume, ``vtkMRMLVectorVolumeNode`` for vector volume.
  :return: created new volume node

  Example::

    # create zero-filled volume
    import numpy as np
    volumeNode = slicer.util.addVolumeFromArray(np.zeros((30, 40, 50)))

  Example::

    # create labelmap volume filled with voxel value of 120
    import numpy as np
    volumeNode = slicer.util.addVolumeFromArray(np.ones((30, 40, 50), 'int8') * 120,
      np.diag([0.2, 0.2, 0.5, 1.0]), nodeClassName="vtkMRMLLabelMapVolumeNode")
  """
    import slicer
    from vtk import vtkMatrix4x4
    import numpy as np

    if name is None:
        name = ""
    if nodeClassName is None:
        nodeClassName = "vtkMRMLScalarVolumeNode"
    if hide == True:
        volumeNode = CreateNodeByClass(nodeClassName)
        volumeNode.SetAttribute("hide_in_manager", "1")
        AddNode(volumeNode)
    else:
        volumeNode = slicer.mrmlScene.AddNewNodeByClass(nodeClassName, name)
    if ijkToRAS is not None:
        if not isinstance(ijkToRAS, vtkMatrix4x4):
            ijkToRAS = vtkMatrixFromArray(ijkToRAS)
        volumeNode.SetIJKToRASMatrix(ijkToRAS)
    updateVolumeFromArray(volumeNode, narray)
    volumeNode.CreateDefaultDisplayNodes()

    return volumeNode


def get_windowlevel():
    import slicer
    applicationLogic = slicer.app.applicationLogic()
    selectionNode = applicationLogic.GetSelectionNode()
    if not selectionNode:
        print("selection node is None2")
        return
    volumeid = selectionNode.GetActiveVolumeID()
    if not volumeid:
        return
    node = GetNodeByID(volumeid)
    if not node:
        print("selection node is None")
        return
    displaynode = GetDisplayNode(node)
    if not displaynode:
        print("selection node display node is None")
        return
    displaynode.AutoWindowLevelOff()
    return displaynode.GetWindow(), displaynode.GetLevel()


def set_windowlevel(window, level):
    import slicer
    applicationLogic = slicer.app.applicationLogic()
    layoutManager = slicer.app.layoutManager()
    view = layoutManager.sliceWidget("Red").sliceView()
    sliceNode = view.mrmlSliceNode()
    sliceLogic = slicer.app.applicationLogic().GetSliceLogic(sliceNode)
    compositeNode = sliceLogic.GetSliceCompositeNode()
    vid = compositeNode.GetBackgroundVolumeID()
    vid2 = compositeNode.GetForegroundVolumeID()
    if vid is None:
        vid = vid2
    print(f"set_windowlevel background node is is {vid}")
    if vid:
        node = GetNodeByID(vid)
        if not node:
            print("empty node of set window/level")
            return
        else:
            pass
    else:
        return

    displaynode = GetDisplayNode(node)
    if not displaynode:
        print("selection node display node is None")
        return
    print(f"Set {node.GetID()} window level of {window}-{level}")
    displaynode.AutoWindowLevelOff()
    displaynode.SetWindow(window)
    displaynode.SetLevel(level)


def get_all_nodes():
    import slicer
    return slicer.mrmlScene.GetNodes()


def arrayFromTableColumn(tableNode, columnName):
    """Return values of a table node's column as numpy array.

  Values can be modified by modifying the numpy array.
  After all modifications has been completed, call :py:meth:`arrayFromTableColumnModified`.

  .. warning:: Important: memory area of the returned array is managed by VTK,
    therefore values in the array may be changed, but the array must not be reallocated.
    See :py:meth:`arrayFromVolume` for details.
  """
    import vtk.util.numpy_support
    columnData = tableNode.GetTable().GetColumnByName(columnName)
    narray = vtk.util.numpy_support.vtk_to_numpy(columnData)
    return narray


def arrayFromTableColumnModified(tableNode, columnName):
    """Indicate that modification of a numpy array returned by :py:meth:`arrayFromTableColumn` has been completed."""
    import vtk.util.numpy_support
    columnData = tableNode.GetTable().GetColumnByName(columnName)
    columnData.Modified()
    tableNode.GetTable().Modified()


def easyPlot(array, columnNames, title, tag, chatnode=None, table_node=None, series_Nodes=None, is_unique_color=True,
             color=[1, 1, 1]):
    import slicer
    import numpy as np
    # s
    nodes = getNodesByClass("vtkMRMLPlotSeriesNode")
    for node in nodes:
        if node.GetAttribute(tag) == "1":
            RemoveNode(node)

    if table_node is None:
        tableNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLTableNode")
    else:
        tableNode = table_node

    nparrays = []
    seriesNodes = []
    if series_Nodes is not None:
        seriesNodes = series_Nodes

    for i in range(len(array)):
        if series_Nodes is None:
            sn = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLPlotSeriesNode")
            sn.SetAttribute(tag, "1")
            if is_unique_color:
                sn.SetUniqueColor()
            else:
                sn.SetColor(color[0], color[1], color[2])
            seriesNodes.append(sn)
        ay = np.array(array[i])
        nparrays.append(ay)
    nodes = {}
    nodes['table'] = tableNode
    nodes['series'] = seriesNodes
    nodes['chart'] = chatnode
    chartNode = plot(np.vstack(nparrays).T, xColumnIndex=0, columnNames=columnNames, title=title, nodes=nodes)
    chartNode.SetLegendVisibility(False)
    return tableNode


def easyUpdatePlot(array, tableNode, columnNames, title, tag):
    import slicer
    import numpy as np
    nodes = getNodesByClass("vtkMRMLPlotSeriesNode")
    for node in nodes:
        if node.GetAttribute(tag) == "1":
            RemoveNode(node)

    nparrays = []
    for i in range(len(array)):
        ay = np.array(array[i])
        nparrays.append(ay)
    nodes = {}
    nodes['table'] = tableNode
    chartNode = slicer.util.plot(np.vstack(nparrays).T, xColumnIndex=0, nodes=nodes, columnNames=columnNames,
                                 title=title)
    chartNode.SetLegendVisibility(False)
    return tableNode


def updateTableFromArray(tableNode, narrays, columnNames=None):
    """Set values in a table node from a numpy array.

  :param columnNames: may contain a string or list of strings that will be used as column name(s).
  :raises ValueError: in case of failure

  Values are copied, therefore if the numpy array  is modified after calling this method,
  values in the table node will not change.
  All previous content of the table is deleted.

  Example::

    import numpy as np
    histogram = np.histogram(arrayFromVolume(getNode('MRHead')))
    tableNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLTableNode")
    updateTableFromArray(tableNode, histogram, ["Count", "Intensity"])
  """
    import numpy as np
    import vtk.util.numpy_support
    import slicer

    if tableNode is None:
        tableNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLTableNode")
    if isinstance(narrays, np.ndarray) and len(narrays.shape) == 1:
        ncolumns = [narrays]
    elif isinstance(narrays, np.ndarray) and len(narrays.shape) == 2:
        ncolumns = narrays.T
    elif isinstance(narrays, tuple) or isinstance(narrays, list):
        ncolumns = narrays
    else:
        raise ValueError('Expected narrays is a numpy ndarray, or tuple or list of numpy ndarrays, got %s instead.' % (
            str(type(narrays))))
    tableNode.RemoveAllColumns()
    # Convert single string to a single-element string list
    if columnNames is None:
        columnNames = []
    if isinstance(columnNames, str):
        columnNames = [columnNames]
    for columnIndex, ncolumn in enumerate(ncolumns):
        vcolumn = vtk.util.numpy_support.numpy_to_vtk(num_array=ncolumn.ravel(), deep=True, array_type=vtk.VTK_FLOAT)
        if (columnNames is not None) and (columnIndex < len(columnNames)):
            vcolumn.SetName(columnNames[columnIndex])
        tableNode.AddColumn(vcolumn)
    return tableNode


def dataframeFromTable(tableNode):
    """Convert table node content to pandas dataframe.

  Table content is copied. Therefore, changes in table node do not affect the dataframe,
  and dataframe changes do not affect the original table node.
  """
    try:
        # Suppress "lzma compression not available" UserWarning when loading pandas
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter(action='ignore', category=UserWarning)
            import pandas as pd
    except ImportError:
        raise ImportError(
            "Failed to convert to pandas dataframe. Please install pandas by running `slicer.util.pip_install('pandas')`")
    dataframe = pd.DataFrame()
    vtable = tableNode.GetTable()
    for columnIndex in range(vtable.GetNumberOfColumns()):
        vcolumn = vtable.GetColumn(columnIndex)
        column = []
        numberOfComponents = vcolumn.GetNumberOfComponents()
        if numberOfComponents == 1:
            # most common, simple case
            for rowIndex in range(vcolumn.GetNumberOfValues()):
                column.append(vcolumn.GetValue(rowIndex))
        else:
            # rare case: column contains multiple components
            valueIndex = 0
            for rowIndex in range(vcolumn.GetNumberOfTuples()):
                item = []
                for componentIndex in range(numberOfComponents):
                    item.append(vcolumn.GetValue(valueIndex))
                    valueIndex += 1
                column.append(item)
        dataframe[vcolumn.GetName()] = column
    return dataframe


def dataframeFromMarkups(markupsNode):
    """Convert table node content to pandas dataframe.

  Table content is copied. Therefore, changes in table node do not affect the dataframe,
  and dataframe changes do not affect the original table node.
  """
    try:
        # Suppress "lzma compression not available" UserWarning when loading pandas
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter(action='ignore', category=UserWarning)
            import pandas as pd
    except ImportError:
        raise ImportError(
            "Failed to convert to pandas dataframe. Please install pandas by running `slicer.util.pip_install('pandas')`")

    label = []
    description = []
    positionWorldR = []
    positionWorldA = []
    positionWorldS = []
    selected = []
    visible = []

    numberOfControlPoints = markupsNode.GetNumberOfControlPoints()
    for controlPointIndex in range(numberOfControlPoints):
        label.append(markupsNode.GetNthControlPointLabel(controlPointIndex))
        description.append(markupsNode.GetNthControlPointDescription(controlPointIndex))
        p = [0, 0, 0]
        markupsNode.GetNthControlPointPositionWorld(controlPointIndex, p)
        positionWorldR.append(p[0])
        positionWorldA.append(p[1])
        positionWorldS.append(p[2])
        selected.append(markupsNode.GetNthControlPointSelected(controlPointIndex) != 0)
        visible.append(markupsNode.GetNthControlPointVisibility(controlPointIndex) != 0)

    dataframe = pd.DataFrame({
        'label': label,
        'position.R': positionWorldR,
        'position.A': positionWorldA,
        'position.S': positionWorldS,
        'selected': selected,
        'visible': visible,
        'description': description})
    return dataframe


#
# VTK
#

class VTKObservationMixin(object):
    def __init__(self):
        super(VTKObservationMixin, self).__init__()
        self.Observations = []

    def removeObservers(self, method=None):
        for o, e, m, g, t, p in list(self.Observations):
            if method == m or method is None:
                o.RemoveObserver(t)
                self.Observations.remove([o, e, m, g, t, p])

    def addObserver(self, object, event, method, group='none', priority=0.0):
        if self.hasObserver(object, event, method):
            print('already has observer')
            return
        tag = object.AddObserver(event, method, priority)
        self.Observations.append([object, event, method, group, tag, priority])

    def removeObserver(self, object, event, method):
        for o, e, m, g, t, p in self.Observations:
            if o == object and e == event and m == method:
                o.RemoveObserver(t)
                self.Observations.remove([o, e, m, g, t, p])

    def hasObserver(self, object, event, method):
        for o, e, m, g, t, p in self.Observations:
            if o == object and e == event and m == method:
                return True
        return False

    def observer(self, event, method):
        for o, e, m, g, t, p in self.Observations:
            if e == event and m == method:
                return o
        return None


def toVTKString(text):
    """Convert unicode string into VTK string.

  .. deprecated:: 4.11.0
    Since now VTK assumes that all strings are in UTF-8 and all strings in Slicer are UTF-8, too,
    conversion is no longer necessary.
    The method is only kept for backward compatibility and will be removed in the future.
  """
    import logging
    logging.warning("toVTKString is deprecated! Conversion is no longer necessary.")
    return text


def toLatin1String(text):
    """Convert string to latin1 encoding."""
    vtkStr = ""
    for c in text:
        try:
            cc = c.encode("latin1", "ignore").decode()
        except (UnicodeDecodeError):
            cc = "?"
        vtkStr = vtkStr + cc
    return vtkStr


#
# File Utilities
#

def tempDirectory(key='__SlicerTemp__', tempDir=None, includeDateTime=True):
    """Come up with a unique directory name in the temp dir and make it and return it

  Note: this directory is not automatically cleaned up
  """
    # TODO: switch to QTemporaryDir in Qt5.
    import qt, slicer
    if not tempDir:
        tempDir = qt.QDir(slicer.app.temporaryPath)
    if includeDateTime:
        tempDirName = key + qt.QDateTime().currentDateTime().toString("yyyy-MM-dd_hh+mm+ss.zzz")
    else:
        tempDirName = key
    fileInfo = qt.QFileInfo(qt.QDir(tempDir), tempDirName)
    dirPath = fileInfo.absoluteFilePath()
    qt.QDir().mkpath(dirPath)
    return dirPath


def delayDisplay(message, autoCloseMsec=1000):
    """Display an information message in a popup window for a short time.

  If ``autoCloseMsec < 0`` then the window is not closed until the user clicks on it

  If ``0 <= autoCloseMsec < 400`` then only ``slicer.app.processEvents()`` is called.

  If ``autoCloseMsec >= 400`` then the window is closed after waiting for autoCloseMsec milliseconds
  """
    import qt, slicer
    import logging
    logging.info(message)
    if 0 <= autoCloseMsec < 400:
        slicer.app.processEvents()
        return
    messagePopup = qt.QDialog()
    layout = qt.QVBoxLayout()
    messagePopup.setLayout(layout)
    label = qt.QLabel(message, messagePopup)
    layout.addWidget(label)
    if autoCloseMsec >= 0:
        qt.QTimer.singleShot(autoCloseMsec, messagePopup.close)
    else:
        okButton = qt.QPushButton("OK")
        layout.addWidget(okButton)
        okButton.connect('clicked()', messagePopup.close)
    messagePopup.exec_()


def infoDisplay(text, windowTitle=None, parent=None, standardButtons=None, **kwargs):
    """Display popup with a info message.

  If there is no main window then the text is only logged (at info level).
  """
    import qt, slicer
    import logging
    if not windowTitle:
        windowTitle = slicer.app.applicationName + " information"
    logging.info(text)
    mw = mainWindow()
    if mw:
        standardButtons = standardButtons if standardButtons else qt.QMessageBox.Ok
        messageBox(text, parent, windowTitle=windowTitle, icon=qt.QMessageBox.Information,
                   standardButtons=standardButtons, **kwargs)


def warningDisplay(text, windowTitle=None, parent=None, standardButtons=None, **kwargs):
    """Display popup with a warning message.

  If there is no main window then the text is only logged (at warning level).
  """
    import qt, slicer
    import logging
    if not windowTitle:
        windowTitle = slicer.app.applicationName + " warning"
    logging.warning(text)
    mw = mainWindow()
    if mw:
        standardButtons = standardButtons if standardButtons else qt.QMessageBox.Ok
        messageBox(text, parent, windowTitle=windowTitle, icon=qt.QMessageBox.Warning, standardButtons=standardButtons,
                   **kwargs)


def errorDisplay(text, windowTitle=None, parent=None, standardButtons=None, **kwargs):
    """Display an error popup.

  If there is no main window then the text is only logged (at error level).
  """
    import qt, slicer
    import logging
    if not windowTitle:
        windowTitle = slicer.app.applicationName + " error"
    logging.error(text)
    mw = mainWindow()
    if mw:
        standardButtons = standardButtons if standardButtons else qt.QMessageBox.Ok
        messageBox(text, parent, windowTitle=windowTitle, icon=qt.QMessageBox.Critical, standardButtons=standardButtons,
                   **kwargs)


def confirmOkCancelDisplay(text, windowTitle=None, parent=None, **kwargs):
    """Display an confirmation popup. Return if confirmed with OK."""
    import qt, slicer
    if not windowTitle:
        windowTitle = slicer.app.applicationName + " confirmation"
    result = messageBox(text, parent=parent, windowTitle=windowTitle, icon=qt.QMessageBox.Question,
                        standardButtons=qt.QMessageBox.Ok | qt.QMessageBox.Cancel, **kwargs)
    return result == 1


def confirmYesNoDisplay(text, windowTitle=None, parent=None, **kwargs):
    """Display an confirmation popup. Return if confirmed with Yes."""
    import qt, slicer
    if not windowTitle:
        windowTitle = slicer.app.applicationName + " confirmation"
    result = messageBox(text, parent=parent, windowTitle=windowTitle, icon=qt.QMessageBox.Question,
                        standardButtons=qt.QMessageBox.Yes | qt.QMessageBox.No, **kwargs)
    return result == 1


def confirmRetryCloseDisplay(text, windowTitle=None, parent=None, **kwargs):
    """Display an confirmation popup. Return if confirmed with Retry."""
    import qt, slicer
    if not windowTitle:
        windowTitle = slicer.app.applicationName + " error"
    result = messageBox(text, parent=parent, windowTitle=windowTitle, icon=qt.QMessageBox.Critical,
                        standardButtons=qt.QMessageBox.Retry | qt.QMessageBox.Close, **kwargs)
    return result == qt.QMessageBox.Retry


def HideProgressBar():
    import qt, slicer
    mainWindow().HideProgressBar()


def clear_all_tableinfo():
    nodes = getNodesByClass("vtkMRMLPlotChartNode")
    for node in nodes:
        RemoveNode(node)
    nodes = getNodesByClass("vtkMRMLTableNode")
    for node in nodes:
        RemoveNode(node)
    nodes = getNodesByClass("vtkMRMLPlotSeriesNode")
    for node in nodes:
        RemoveNode(node)


def reset_pixel_label_style(btn, index):
    label = findChild(btn, "labelText")
    label.setStyleSheet("font: 16px 'Source Han Sans CN-Regular, Source Han Sans CN';")
    labelPic = findChild(btn, "labelPic")
    labelPic.setStyleSheet("")

    # 正常状态
    if index == 0:
        btnScissors_stylesheet = ""
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: transparent;: 1px solid #ffffff}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:hover{background-color: transparent;: 1px solid #999999}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:pressed{background-color: transparent;: 1px solid #ffffff}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:checked{background-color: #3377FF;: 1px solid #ffffff};"
        btn.setStyleSheet(btnScissors_stylesheet)
    # 黄色边框
    if index == 1:
        btnScissors_stylesheet = ""
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: transparent;: 1px solid #ffff00}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:hover{background-color: transparent;: 1px solid #aaaa00}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:pressed{background-color: transparent;: 1px solid #ffffff}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:checked{background-color: transparent;: 1px solid #555500};"
        btn.setStyleSheet(btnScissors_stylesheet)
    # 黄色边框+黄色文字
    if index == 2:
        btnScissors_stylesheet = ""
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: transparent;: 1px solid #ffff00};"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:hover{background-color: transparent;: 1px solid #ffff00}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:pressed{background-color: transparent;: 1px solid #aaaa00}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:checked{background-color: transparent;: 1px solid #555500};"
        btn.setStyleSheet(btnScissors_stylesheet)
        label = findChild(btn, "labelText")
        label.setStyleSheet("color: rgb(166, 166, 0);")
    # 绿色边框
    if index == 3:
        btnScissors_stylesheet = ""
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: transparent;: 1px solid #00ff00}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:hover{background-color: transparent;: 1px solid #999999}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:pressed{background-color: transparent;: 1px solid #ffffff}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:checked{background-color: transparent;: 1px solid #555555};"
        btn.setStyleSheet(btnScissors_stylesheet)
    # 绿色背景
    if index == 4:
        btnScissors_stylesheet = ""
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: rgb(0, 166, 0);}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:hover{background-color: transparent;: 1px solid #00ff00}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:pressed{background-color: transparent;: 1px solid #00aa00}"
        btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:checked{background-color: transparent;: 1px solid #005500};"
        btn.setStyleSheet(btnScissors_stylesheet)

        label = findChild(btn, "labelText")
        label.setStyleSheet("background-color: rgb(0, 166, 0);")

        labelPic = findChild(btn, "labelPic")
        labelPic.setStyleSheet("background-color: rgb(0, 166, 0);")


def set_dark_normal_pushbutton_style(btn):
    btnScissors_stylesheet = ""
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: rgb(21, 21, 21);}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:hover{background-color: rgb(70, 70, 70);}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: rgb(51, 51, 51);}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: rgb(151, 151, 151);}"
    btn.setStyleSheet(btnScissors_stylesheet)


def set_dark_highlight_pushbutton_style(btn):
    btnScissors_stylesheet = ""
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: rgb(151, 151, 151);}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:hover{background-color: rgb(70, 70, 70);}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: rgb(151, 151, 151);}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: rgb(251, 251, 251);}"
    btn.setStyleSheet(btnScissors_stylesheet)


def add_pixel_label(pic_path, label, btn):
    import qt
    labelPic = qt.QLabel()
    labelText = qt.QLabel()
    labelPic.setPixmap(qt.QPixmap(pic_path))
    labelPic.setObjectName("labelPic")
    labelText.setText(label)
    labelText.setObjectName("labelText")
    labelText.setStyleSheet("font: 16px 'Source Han Sans CN-Regular, Source Han Sans CN';")
    layout = qt.QHBoxLayout()
    layout.addSpacing(2)
    layout.addWidget(labelPic)
    layout.addSpacing(10)
    layout.addWidget(labelText)
    layout.addStretch()
    btn.setLayout(layout)

    btnScissors_stylesheet = ""
    btnScissors_stylesheet = btnScissors_stylesheet + "QToolTip { color: #000000; background-color: #ffffff; border: 0px; }"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: transparent;border: 1px solid #ffffff}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:hover{background-color: #527fd9;border: 1px solid #527fd9}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:pressed{background-color: transparent;border: 1px solid #ffffff}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:checked{background-color: #204aa6;border: 1px solid #ffffff};"
    btn.setStyleSheet(btnScissors_stylesheet)


def add_pixel_middle(pic_path, btn, tooltip, size=24):
    import qt
    labelPic = qt.QLabel()
    pixelmap_scaled = qt.QPixmap(pic_path)
    pixelmap_scaled = pixelmap_scaled.scaled(size, size, 0, 1)
    labelPic.setPixmap(pixelmap_scaled)
    labelPic.setObjectName("labelPic")
    layout = qt.QHBoxLayout()
    layout.addStretch()
    layout.addWidget(labelPic)
    layout.addStretch()
    btn.setLayout(layout)
    labelPic.setToolTip(tooltip)


def add_pixel_label_middle(pic_path, label, btn):
    import qt
    labelPic = qt.QLabel()
    labelText = qt.QLabel()
    pixelmap_scaled = qt.QPixmap(pic_path)
    pixelmap_scaled = pixelmap_scaled.scaled(24, 24, 0, 1)
    labelPic.setPixmap(pixelmap_scaled)
    labelPic.setObjectName("labelPic")
    labelText.setText(label)
    labelText.setObjectName("labelText")
    labelText.setStyleSheet("font: 16px 'Source Han Sans CN-Regular, Source Han Sans CN';")
    layout = qt.QHBoxLayout()
    layout.addSpacing(2)
    layout.addWidget(labelPic)
    layout.addSpacing(10)
    layout.addWidget(labelText)
    layout.addStretch()
    btn.setLayout(layout)

    btnScissors_stylesheet = ""
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{border: 1px solid #ffffff}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:hover{border: 1px solid #999999}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:pressed{border: 1px solid #ffffff}"
    btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:checked{border: 1px solid #555555};"
    btn.setStyleSheet(btnScissors_stylesheet)


def zoom_2d(factor, sliceNodes=None):
    import slicer
    """Zoom slice nodes by factor.
    factor: "Fit" or +/- amount to zoom
    sliceNodes: list of slice nodes to change, None means all.
    """
    layoutManager = slicer.app.layoutManager()
    for sliceNode in sliceNodes:
        if factor == "Fit":
            sliceWidget = layoutManager.sliceWidget(sliceNode.GetLayoutName())
            if sliceWidget:
                sliceWidget.sliceLogic().FitSliceToAll()
        else:
            newFOVx = sliceNode.GetFieldOfView()[0] * factor
            newFOVy = sliceNode.GetFieldOfView()[1] * factor
            newFOVz = sliceNode.GetFieldOfView()[2]
            sliceNode.SetFieldOfView(newFOVx, newFOVy, newFOVz)
            sliceNode.UpdateMatrices()


def showWarningText(text):
    return mainWindow().ShowWarningText(text)


def write_test_log(txt):
    print(txt)
    # 创建并打开日志文件
    with open(unit_test["log_file_path"], "a") as log_file:
        log_file.write(f"{txt}\n")


def create_log_file(directory_path="D:/"):
    import os
    import datetime
    import slicer
    # 获取当前日期和时间
    now = datetime.datetime.now()
    date_time = now.strftime("%Y-%m-%d_%H-%M-%S")  # 格式化日期和时间为字符串

    # 创建一个以日期和时间为单位的日志文件名
    log_file_name = f"log_{date_time}.txt"

    # 指定日志文件的路径（可根据需要修改）
    log_directory = os.path.join(directory_path, "log")

    # 确保日志目录存在，如果不存在则创建
    if not os.path.exists(log_directory):
        os.makedirs(log_directory)

    # 构建完整的日志文件路径
    unit_test["log_file_path"] = os.path.join(log_directory, log_file_name)

    # 创建并打开日志文件
    with open(unit_test["log_file_path"], "w") as log_file:
        log_file.write(f"Log file created at {now}\n")


def kill_process_by_name(process_name):
    import psutil

    for process in psutil.process_iter(attrs=['name']):
        if process.info['name'] == process_name:
            try:
                process.terminate()
                print(f"Killed process with name: {process_name}")
            except psutil.NoSuchProcess:
                pass


def messageBoxReject(text, parent=None, **kwargs):
    import slicer
    windowTitle = None
    for key, value in kwargs.items():
        if key == 'windowTitle':
            windowTitle = value
    return mainWindow().ShowInformationReject(windowTitle, text)


def trace(tag=1):
    import traceback
    str1 = "".join(traceback.format_stack()[:-1])
    print(str1)
    if tag == 1:
        import pdb
        pdb.set_trace()


def messageBox(text, parent=None, **kwargs):
    """Displays a messagebox.

  ctkMessageBox is used instead of a default qMessageBox to provide "Don't show again" checkbox.

  For example::

    slicer.util.messageBox("Some message", dontShowAgainSettingsKey = "MainWindow/DontShowSomeMessage")
  """
    # import qt
    # import ctk
    # mbox = ctk.ctkMessageBox(parent if parent else mainWindow())
    # mbox.text = text
    # for key, value in kwargs.items():
    #   if hasattr(mbox, key):
    #     setattr(mbox, key, value)
    # # Windows 10 peek feature in taskbar shows all hidden but not destroyed windows
    # # (after creating and closing a messagebox, hovering over the mouse on Slicer icon, moving up the
    # # mouse to the peek thumbnail would show it again).
    # # Popup windows in other Qt applications often show closed popups (such as
    # # Paraview's Edit / Find data dialog, MeshMixer's File/Preferences dialog).
    # # By calling deleteLater, the messagebox is permanently deleted when the current call is completed.
    # mbox.deleteLater()
    # return mbox.exec_()
    import slicer
    windowTitle = None
    for key, value in kwargs.items():
        if key == 'windowTitle':
            windowTitle = value
    return mainWindow().ShowInformation(windowTitle, text)


def messageBox2(text, parent=None, title="消息", **kwargs):
    """Displays a messagebox.

  ctkMessageBox is used instead of a default qMessageBox to provide "Don't show again" checkbox.

  For example::

    slicer.util.messageBox("Some message", dontShowAgainSettingsKey = "MainWindow/DontShowSomeMessage")
  """
    import qt
    import ctk
    mbox = ctk.ctkMessageBox(parent if parent else mainWindow())
    mbox.text = text
    mbox.setWindowTitle(title)
    for key, value in kwargs.items():
        if hasattr(mbox, key):
            setattr(mbox, key, value)
    # Windows 10 peek feature in taskbar shows all hidden but not destroyed windows
    # (after creating and closing a messagebox, hovering over the mouse on Slicer icon, moving up the
    # mouse to the peek thumbnail would show it again).
    # Popup windows in other Qt applications often show closed popups (such as
    # Paraview's Edit / Find data dialog, MeshMixer's File/Preferences dialog).
    # By calling deleteLater, the messagebox is permanently deleted when the current call is completed.
    mbox.deleteLater()
    return mbox.exec_()


def createProgressDialog2(parent=None, value=0, maximum=100, labelText="", windowTitle="请稍候.....", **kwargs):
    if parent is None:
        parent = mainWindow()
    getModuleWidget("JProgressBar").InitProgressBar(windowTitle, value, maximum)
    return getModuleWidget("JProgressBar").bar_item


def createProgressDialog2Old(parent=None, value=0, maximum=100, labelText="", windowTitle="请稍候.....", **kwargs):
    if parent is None:
        parent = mainWindow()
    progressIndicator = mainWindow().ShowProgressBar(windowTitle)
    return progressIndicator


def createProgressDialog(parent=None, value=0, maximum=100, labelText="", windowTitle="请稍候.....", **kwargs):
    """Display a modal QProgressDialog.

  Go to `QProgressDialog documentation <https://doc.qt.io/qt-5/qprogressdialog.html>`_ to
  learn about the available keyword arguments.

  Examples::

    # Prevent progress dialog from automatically closing
    progressbar = createProgressDialog(autoClose=False)

    # Update progress value
    progressbar.value = 50

    # Update label text
    progressbar.labelText = "processing XYZ"
  """
    import qt, slicer
    progressIndicator = qt.QProgressDialog(parent if parent else mainWindow())
    progressIndicator.minimumDuration = 0
    progressIndicator.maximum = maximum
    progressIndicator.value = value
    progressIndicator.windowTitle = windowTitle
    progressIndicator.labelText = labelText
    progressIndicator.setWindowModality(1)
    progressIndicator.resize(500, 300)
    progressIndicator.setCancelButtonText("取消")
    return progressIndicator


def setProgress(value):
    import qt, slicer
    print("setProgress", value, int(value))
    getModuleWidget("JProgressBar").SetProgress(value)


def updateProgress():
    getModuleWidget("JProgressBar").UpdateProgressStep(1)


def toBool(value):
    """Convert any type of value to a boolean.

  The function uses the following heuristic:

  1. If the value can be converted to an integer, the integer is then
     converted to a boolean.
  2. If the value is a string, return True if it is equal to 'true'. False otherwise.
     Note that the comparison is case insensitive.
  3. If the value is neither an integer or a string, the bool() function is applied.

  >>> [toBool(x) for x in range(-2, 2)]
  [True, True, False, True]
  >>> [toBool(x) for x in ['-2', '-1', '0', '1', '2', 'Hello']]
  [True, True, False, True, True, False]
  >>> toBool(object())
  True
  >>> toBool(None)
  False
  """
    try:
        return bool(int(value))
    except (ValueError, TypeError):
        return value.lower() in ['true'] if isinstance(value, str) else bool(value)


def settingsValue(key, default, converter=lambda v: v, settings=None):
    """Return settings value associated with key if it exists or the provided default otherwise.

  ``settings`` parameter is expected to be a valid ``qt.Settings`` object.
  """
    import qt
    settings = qt.QSettings() if settings is None else settings
    return converter(settings.value(key)) if settings.contains(key) else default


def clickAndDrag(widget, button='Left', start=(10, 10), end=(10, 40), steps=20, modifiers=[]):
    """Send synthetic mouse events to the specified widget (qMRMLSliceWidget or qMRMLThreeDView)

  :param button: "Left", "Middle", "Right", or "None"
   start, end : window coordinates for action
  :param steps: number of steps to move in, if <2 then mouse jumps to the end position
  :param modifiers: list containing zero or more of "Shift" or "Control"
  :raises RuntimeError: in case of failure

  .. hint::

    For generating test data you can use this snippet of code::

      layoutManager = slicer.app.layoutManager()
      threeDView = layoutManager.threeDWidget(0).threeDView()
      style = threeDView.interactorStyle()
      interactor = style.GetInteractor()

      def onClick(caller,event):
          print(interactor.GetEventPosition())

      interactor.AddObserver(vtk.vtkCommand.LeftButtonPressEvent, onClick)
  """
    style = widget.interactorStyle()
    interactor = style.GetInteractor()
    if button == 'Left':
        down = interactor.LeftButtonPressEvent
        up = interactor.LeftButtonReleaseEvent
    elif button == 'Right':
        down = interactor.RightButtonPressEvent
        up = interactor.RightButtonReleaseEvent
    elif button == 'Middle':
        down = interactor.MiddleButtonPressEvent
        up = interactor.MiddleButtonReleaseEvent
    elif button == 'None' or not button:
        down = lambda: None
        up = lambda: None
    else:
        raise RuntimeError("Bad button - should be Left or Right, not %s" % button)
    if 'Shift' in modifiers:
        interactor.SetShiftKey(1)
    if 'Control' in modifiers:
        interactor.SetControlKey(1)
    interactor.SetEventPosition(*start)
    down()
    if (steps < 2):
        interactor.SetEventPosition(end[0], end[1])
        interactor.MouseMoveEvent()
    else:
        for step in range(steps):
            frac = float(step) / (steps - 1)
            x = int(start[0] + frac * (end[0] - start[0]))
            y = int(start[1] + frac * (end[1] - start[1]))
            interactor.SetEventPosition(x, y)
            interactor.MouseMoveEvent()
    up()
    interactor.SetShiftKey(0)
    interactor.SetControlKey(0)


def downloadFile(url, targetFilePath, checksum=None, reDownloadIfChecksumInvalid=True):
    """Download ``url`` to local storage as ``targetFilePath``

  Target file path needs to indicate the file name and extension as well

  If specified, the ``checksum`` is used to verify that the downloaded file is the expected one.
  It must be specified as ``<algo>:<digest>``. For example, ``SHA256:cc211f0dfd9a05ca3841ce1141b292898b2dd2d3f08286affadf823a7e58df93``.
  """
    import os
    import logging
    try:
        (algo, digest) = extractAlgoAndDigest(checksum)
    except ValueError as excinfo:
        logging.error('Failed to parse checksum: ' + excinfo.message)
        return False
    if not os.path.exists(targetFilePath) or os.stat(targetFilePath).st_size == 0:
        logging.info('Downloading from\n  %s\nas file\n  %s\nIt may take a few minutes...' % (url, targetFilePath))
        try:
            import urllib.request, urllib.parse, urllib.error
            urllib.request.urlretrieve(url, targetFilePath)
        except Exception as e:
            import traceback
            traceback.print_exc()
            logging.error('Failed to download file from ' + url)
            return False
        if algo is not None:
            logging.info('Verifying checksum\n  %s' % targetFilePath)
            current_digest = computeChecksum(algo, targetFilePath)
            if current_digest != digest:
                logging.error('Downloaded file does not have expected checksum.'
                              '\n   current checksum: %s'
                              '\n  expected checksum: %s' % (current_digest, digest))
                return False
            else:
                logging.info('Checksum OK')
    else:
        if algo is not None:
            current_digest = computeChecksum(algo, targetFilePath)
            if current_digest != digest:
                if reDownloadIfChecksumInvalid:
                    logging.info(
                        'Requested file has been found but its checksum is different: deleting and re-downloading')
                    os.remove(targetFilePath)
                    return downloadFile(url, targetFilePath, checksum, reDownloadIfChecksumInvalid=False)
                else:
                    logging.error('Requested file has been found but its checksum is different:'
                                  '\n   current checksum: %s'
                                  '\n  expected checksum: %s' % (current_digest, digest))
                    return False
            else:
                logging.info('Requested file has been found and checksum is OK: ' + targetFilePath)
        else:
            logging.info('Requested file has been found: ' + targetFilePath)
    return True


def extractArchive(archiveFilePath, outputDir, expectedNumberOfExtractedFiles=None):
    """ Extract file ``archiveFilePath`` into folder ``outputDir``.

  Number of expected files unzipped may be specified in ``expectedNumberOfExtractedFiles``.
  If folder contains the same number of files as expected (if specified), then it will be
  assumed that unzipping has been successfully done earlier.
  """
    import os
    import logging
    from slicer import app
    if not os.path.exists(archiveFilePath):
        logging.error('Specified file %s does not exist' % (archiveFilePath))
        return False
    fileName, fileExtension = os.path.splitext(archiveFilePath)
    if fileExtension.lower() != '.zip':
        # TODO: Support other archive types
        logging.error('Only zip archives are supported now, got ' + fileExtension)
        return False

    numOfFilesInOutputDir = len(getFilesInDirectory(outputDir, False))
    if expectedNumberOfExtractedFiles is not None \
            and numOfFilesInOutputDir == expectedNumberOfExtractedFiles:
        logging.info('File %s already unzipped into %s' % (archiveFilePath, outputDir))
        return True

    extractSuccessful = app.applicationLogic().Unzip(archiveFilePath, outputDir)
    numOfFilesInOutputDirTest = len(getFilesInDirectory(outputDir, False))
    if extractSuccessful is False or (expectedNumberOfExtractedFiles is not None \
                                      and numOfFilesInOutputDirTest != expectedNumberOfExtractedFiles):
        logging.error('Unzipping %s into %s failed' % (archiveFilePath, outputDir))
        return False
    logging.info('Unzipping %s into %s successful' % (archiveFilePath, outputDir))
    return True


def computeChecksum(algo, filePath):
    """Compute digest of ``filePath`` using ``algo``.

  Supported hashing algorithms are SHA256, SHA512, and MD5.

  It internally reads the file by chunk of 8192 bytes.

  :raises ValueError: if algo is unknown.
  :raises IOError: if filePath does not exist.
  """
    import hashlib

    if algo not in ['SHA256', 'SHA512', 'MD5']:
        raise ValueError("unsupported hashing algorithm %s" % algo)

    with open(filePath, 'rb') as content:
        hash = hashlib.new(algo)
        while True:
            chunk = content.read(8192)
            if not chunk:
                break
            hash.update(chunk)
        return hash.hexdigest()


def extractAlgoAndDigest(checksum):
    """Given a checksum string formatted as ``<algo>:<digest>`` returns the tuple ``(algo, digest)``.

  ``<algo>`` is expected to be `SHA256`, `SHA512`, or `MD5`.
  ``<digest>`` is expected to be the full length hexdecimal digest.

  :raises ValueError: if checksum is incorrectly formatted.
  """
    if checksum is None:
        return None, None
    if len(checksum.split(':')) != 2:
        raise ValueError("invalid checksum '%s'. Expected format is '<algo>:<digest>'." % checksum)
    (algo, digest) = checksum.split(':')
    expected_algos = ['SHA256', 'SHA512', 'MD5']
    if algo not in expected_algos:
        raise ValueError("invalid algo '%s'. Algo must be one of %s" % (algo, ", ".join(expected_algos)))
    expected_digest_length = {'SHA256': 64, 'SHA512': 128, 'MD5': 32}
    if len(digest) != expected_digest_length[algo]:
        raise ValueError("invalid digest length %d. Expected digest length for %s is %d" % (
            len(digest), algo, expected_digest_length[algo]))
    return algo, digest


def downloadAndExtractArchive(url, archiveFilePath, outputDir, \
                              expectedNumberOfExtractedFiles=None, numberOfTrials=3, checksum=None):
    """ Downloads an archive from ``url`` as ``archiveFilePath``, and extracts it to ``outputDir``.

  This combined function tests the success of the download by the extraction step,
  and re-downloads if extraction failed.

  If specified, the ``checksum`` is used to verify that the downloaded file is the expected one.
  It must be specified as ``<algo>:<digest>``. For example, ``SHA256:cc211f0dfd9a05ca3841ce1141b292898b2dd2d3f08286affadf823a7e58df93``.
  """
    import os
    import shutil
    import logging

    maxNumberOfTrials = numberOfTrials

    def _cleanup():
        # If there was a failure, delete downloaded file and empty output folder
        logging.warning(
            'Download and extract failed, removing archive and destination folder and retrying. Attempt #%d...' % (
                    maxNumberOfTrials - numberOfTrials))
        os.remove(archiveFilePath)
        shutil.rmtree(outputDir)
        os.mkdir(outputDir)

    while numberOfTrials:
        if not downloadFile(url, archiveFilePath, checksum):
            numberOfTrials -= 1
            _cleanup()
            continue
        if not extractArchive(archiveFilePath, outputDir, expectedNumberOfExtractedFiles):
            numberOfTrials -= 1
            _cleanup()
            continue
        return True

    _cleanup()
    return False


def getFilesInDirectory(directory, absolutePath=True):
    """Collect all files in a directory and its subdirectories in a list."""
    import os
    allFiles = []
    for root, subdirs, files in os.walk(directory):
        for fileName in files:
            if absolutePath:
                fileAbsolutePath = os.path.abspath(os.path.join(root, fileName)).replace('\\', '/')
                allFiles.append(fileAbsolutePath)
            else:
                allFiles.append(fileName)
    return allFiles


def plot(narray, xColumnIndex=-1, columnNames=None, title=None, show=True, nodes=None):
    """Create a plot from a numpy array that contains two or more columns.

  :param narray: input numpy array containing data series in columns.
  :param xColumnIndex: index of column that will be used as x axis.
    If it is set to negative number (by default) then row index will be used as x coordinate.
  :param columnNames: names of each column of the input array.
  :param title: title of the chart. Plot node names are set based on this value.
  :param nodes: plot chart, table, and list of plot series nodes.
    Specified in a dictionary, with keys: 'chart', 'table', 'series'.
    Series contains a list of plot series nodes (one for each table column).
    The parameter is used both as an input and output.
  :return: plot chart node. Plot chart node provides access to chart properties and plot series nodes.

  Example 1: simple plot

  .. code-block:: python

    # Get sample data
    import numpy as np
    import SampleData
    volumeNode = SampleData.downloadSample("MRHead")

    # Create new plot
    histogram = np.histogram(arrayFromVolume(volumeNode), bins=50)
    chartNode = plot(histogram, xColumnIndex = 1)

    # Change some plot properties
    chartNode.SetTitle("My histogram")
    chartNode.GetNthPlotSeriesNode(0).SetPlotType(slicer.vtkMRMLPlotSeriesNode.PlotTypeScatterBar)

  Example 2: plot with multiple updates

  .. code-block:: python

    # Get sample data
    import numpy as np
    import SampleData
    volumeNode = SampleData.downloadSample("MRHead")

    # Create variable that will store plot nodes (chart, table, series)
    plotNodes = {}

    # Create new plot
    histogram = np.histogram(arrayFromVolume(volumeNode), bins=80)
    plot(histogram, xColumnIndex = 1, nodes = plotNodes)

    # Update plot
    histogram = np.histogram(arrayFromVolume(volumeNode), bins=40)
    plot(histogram, xColumnIndex = 1, nodes = plotNodes)
  """
    import slicer

    chartNode = None
    tableNode = None
    seriesNodes = []

    # Retrieve nodes that must be reused
    if nodes is not None:
        if 'chart' in nodes:
            chartNode = nodes['chart']
        if 'table' in nodes:
            tableNode = nodes['table']
        if 'series' in nodes:
            seriesNodes = nodes['series']

    # Create table node
    if tableNode is None:
        tableNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLTableNode")

    if title is not None:
        tableNode.SetName(title + ' table')

    updateTableFromArray(tableNode, narray)

    # Update column names
    numberOfColumns = tableNode.GetTable().GetNumberOfColumns()
    yColumnIndex = 0
    for columnIndex in range(numberOfColumns):
        if (columnNames is not None) and (len(columnNames) > columnIndex):
            columnName = columnNames[columnIndex]
        else:
            if columnIndex == xColumnIndex:
                columnName = "X"
            elif yColumnIndex == 0:
                columnName = "Y"
                yColumnIndex += 1
            else:
                columnName = "Y" + str(yColumnIndex)
                yColumnIndex += 1
        tableNode.GetTable().GetColumn(columnIndex).SetName(columnName)

    # Create chart and add plot
    if chartNode is None:
        chartNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLPlotChartNode")
    if title is not None:
        chartNode.SetName(title + ' chart')
        chartNode.SetTitle(title)

    # Create plot series node(s)
    xColumnName = columnNames[xColumnIndex] if (columnNames is not None) and (len(columnNames) > 0) else "X"
    seriesIndex = -1
    for columnIndex in range(numberOfColumns):
        if columnIndex == xColumnIndex:
            continue
        seriesIndex += 1
        if len(seriesNodes) > seriesIndex:
            seriesNode = seriesNodes[seriesIndex]
        else:
            seriesNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLPlotSeriesNode")
            seriesNodes.append(seriesNode)
            seriesNode.SetUniqueColor()
        seriesNode.SetAndObserveTableNodeID(tableNode.GetID())
        if xColumnIndex < 0:
            seriesNode.SetXColumnName("")
            seriesNode.SetPlotType(seriesNode.PlotTypeLine)
        else:
            seriesNode.SetXColumnName(xColumnName)
            seriesNode.SetPlotType(seriesNode.PlotTypeScatter)
        yColumnName = tableNode.GetTable().GetColumn(columnIndex).GetName()
        seriesNode.SetYColumnName(yColumnName)
        if title:
            seriesNode.SetName(title + " " + yColumnName)

        if not chartNode.HasPlotSeriesNodeID(seriesNode.GetID()):
            chartNode.AddAndObservePlotSeriesNodeID(seriesNode.GetID())

    # Show plot in layout
    if nodes['chart'] is None:
        if show:
            slicer.modules.plots.logic().ShowChartInLayout(chartNode)

    # Without this, chart view may show up completely empty when the same nodes are updated
    # (this is probably due to a bug in plotting nodes or widgets).
    chartNode.Modified()

    if nodes is not None:
        nodes['table'] = tableNode
        nodes['chart'] = chartNode
        nodes['series'] = seriesNodes
    return chartNode


def launchConsoleProcess(args, useStartupEnvironment=True, updateEnvironment=None, cwd=None):
    """Launch a process. Hiding the console and captures the process output.

  The console window is hidden when running on Windows.

  :param args: executable name, followed by command-line arguments
  :param useStartupEnvironment: launch the process in the original environment as the original Slicer process
  :param updateEnvironment: map containing optional additional environment variables (existing variables are overwritten)
  :param cwd: current working directory
  :return: process object.
  """
    import subprocess
    import os
    if useStartupEnvironment:
        startupEnv = startupEnvironment()
        if updateEnvironment:
            startupEnv.update(updateEnvironment)
    else:
        if updateEnvironment:
            startupEnv = os.environ.copy()
            startupEnv.update(updateEnvironment)
        else:
            startupEnv = None
    if os.name == 'nt':
        # Hide console window (only needed on Windows)
        info = subprocess.STARTUPINFO()
        info.dwFlags = 1
        info.wShowWindow = 0
        proc = subprocess.Popen(args, env=startupEnv, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                universal_newlines=True, startupinfo=info, cwd=cwd)
    else:
        proc = subprocess.Popen(args, env=startupEnv, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                universal_newlines=True, cwd=cwd)
    return proc


def removeSegmentByName(segmentnode, name):
    segmentid = segmentnode.GetSegmentation().GetSegmentIdBySegmentName(name)
    segmentnode.GetSegmentation().RemoveSegment(segmentid)


def clone(node):
    import slicer
    shNode = slicer.vtkMRMLSubjectHierarchyNode.GetSubjectHierarchyNode(slicer.mrmlScene)
    itemIDToClone = shNode.GetItemByDataNode(node)
    clonedItemID = slicer.modules.subjecthierarchy.logic().CloneSubjectHierarchyItem(shNode, itemIDToClone)
    clonedNode = shNode.GetItemDataNode(clonedItemID)
    return clonedNode


def convert_segment_to_labelmap(segmentation_node):
    import slicer
    labelmapVolumeNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLLabelMapVolumeNode")
    slicer.modules.segmentations.logic().ExportAllSegmentsToLabelmapNode(segmentation_node, labelmapVolumeNode,
                                                                         slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY)
    return labelmapVolumeNode


def convert_segment_to_model(segmentation_node):
    import slicer, vtk
    opacity = GetDisplayNode(segmentation_node).GetOpacity()
    shNode = slicer.vtkMRMLSubjectHierarchyNode.GetSubjectHierarchyNode(slicer.mrmlScene)
    exportFolderItemId = shNode.CreateFolderItem(shNode.GetSceneItemID(), "Segments")
    slicer.modules.segmentations.logic().ExportAllSegmentsToModels(segmentation_node, exportFolderItemId)
    segmentModels = vtk.vtkCollection()
    shNode.GetDataNodesInBranch(exportFolderItemId, segmentModels)
    ModelNode = segmentModels.GetItemAsObject(0)
    ModelNode.GetDisplayNode().SetVisibility2D(True)
    ModelNode.GetDisplayNode().SetOpacity(opacity)
    ModelNode.GetDisplayNode().SetSliceIntersectionOpacity(opacity)
    return ModelNode


def convert_volume_to_segment(volume_node, threhold=0):
    segment_node = CreateDefaultSegmentationNode("tmp")
    nparray = arrayFromVolume(volume_node)
    nparray[arrayFromVolume(volume_node) > threhold] = 1  # create segment by simple thresholding of an image
    updateSegmentBinaryLabelmapFromArray(nparray, segment_node, GetNthSegmentID(segment_node, 0), volume_node)
    return segment_node


def convert_model_to_volume(model_node, master_node):
    import slicer
    complete_segment_node = convert_model_to_segment(model_node, master_node)
    complete_segment_node.SetName("complete_segment_node")
    complete_segment_volume = clone(master_node)
    complete_segment_volume.SetName(model_node.GetName() + "_volume")
    segmentEditorWidget = slicer.modules.segmenteditor.widgetRepresentation().self().editor
    segmentEditorWidget.setSegmentationNode(complete_segment_node)
    segmentEditorWidget.setSourceVolumeNode(master_node)
    segmentEditorWidget.setActiveEffectByName("Mask volume")
    sid = GetNthSegmentID(complete_segment_node, 0)
    segmentEditorWidget.setCurrentSegmentID(sid)
    effect = segmentEditorWidget.activeEffect()
    effect.setParameter('BinaryMaskFillValueInside', -1000)
    effect.setParameter('BinaryMaskFillValueOutside', 800)
    effect.setParameter('Operation', 'FILL_INSIDE_AND_OUTSIDE')
    effect.self().inputVolumeSelector.setCurrentNode(master_node)
    effect.self().outputVolumeSelector.setCurrentNode(complete_segment_volume)
    effect.self().onApply()
    RemoveNode(complete_segment_node)
    return complete_segment_volume


def convert_segment_to_volume(segment_node, master_node):
    import slicer
    complete_segment_volume = clone(master_node)
    complete_segment_volume.SetName(segment_node.GetName() + "_volume")

    segmentEditorWidget = slicer.modules.segmenteditor.widgetRepresentation().self().editor
    segmentEditorWidget.setSegmentationNode(segment_node)
    segmentEditorWidget.setSourceVolumeNode(master_node)
    segmentEditorWidget.setActiveEffectByName("Mask volume")
    sid = GetNthSegmentID(segment_node, 0)
    segmentEditorWidget.setCurrentSegmentID(sid)
    effect = segmentEditorWidget.activeEffect()
    effect.setParameter('BinaryMaskFillValueInside', -1000)
    effect.setParameter('BinaryMaskFillValueOutside', 800)
    effect.setParameter('Operation', 'FILL_INSIDE_AND_OUTSIDE')
    effect.self().inputVolumeSelector.setCurrentNode(master_node)
    effect.self().outputVolumeSelector.setCurrentNode(complete_segment_volume)
    effect.self().onApply()
    return complete_segment_volume


def convert_model_to_segment(model_node, master_node):
    import slicer, vtk
    seg = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentationNode")
    seg.SetReferenceImageGeometryParameterFromVolumeNode(master_node)
    slicer.modules.segmentations.logic().ImportModelToSegmentationNode(model_node, seg)
    seg.CreateClosedSurfaceRepresentation()
    if not seg.GetDisplayNode():
        seg.CreateDefaultDisplayNodes()
    return seg


def convert_scalar_to_labelmap(scalar_node):
    import slicer
    targetVolumeNode = AddNewNodeByClass("vtkMRMLLabelMapVolumeNode")
    getModuleLogic("Volumes").CreateLabelVolumeFromVolume(slicer.mrmlScene, targetVolumeNode, scalar_node)
    return targetVolumeNode


def convert_model_to_segment_accuracy(model_node):
    import slicer, vtk
    segmentationNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentationNode")
    segmentationNode.CreateDefaultDisplayNodes()

    # Set segmentation resolution (for closed surface to binary labelmap conversion)
    geometryImageData = slicer.vtkOrientedImageData()
    geometryImageData.SetSpacing(0.2, 0.2, 0.2)
    geometryString = slicer.vtkSegmentationConverter.SerializeImageGeometry(geometryImageData)
    segmentationNode.GetSegmentation().SetConversionParameter(
        slicer.vtkSegmentationConverter.GetReferenceImageGeometryParameterName(), geometryString)

    segmentationNode.AddSegmentFromClosedSurfaceRepresentation(model_node.GetPolyData(), "Conv", [1.0, 0.0, 0.0])
    return segmentationNode


def logProcessOutput(proc):
    """Continuously write process output to the application log and the Python console.

  :param proc: process object.
  """
    from subprocess import Popen, PIPE, CalledProcessError
    import logging
    try:
        from slicer import app
        guiApp = app
    except ImportError:
        # Running from console
        guiApp = None
    for line in proc.stdout:
        if guiApp:
            logging.info(line.rstrip())
            guiApp.processEvents()  # give a chance the application to refresh GUI
        else:
            print(line.rstrip())
    proc.wait()
    retcode = proc.returncode
    if retcode != 0:
        raise CalledProcessError(retcode, proc.args, output=proc.stdout, stderr=proc.stderr)


def _executePythonModule(module, args):
    """Execute a Python module as a script in Slicer's Python environment.

  Internally python -m is called with the module name and additional arguments.

  :raises RuntimeError: in case of failure
  """
    # Determine pythonSlicerExecutablePath
    try:
        from slicer import app
        # If we get to this line then import from "app" is succeeded,
        # which means that we run this function from Slicer Python interpreter.
        # PythonSlicer is added to PATH environment variable in Slicer
        # therefore shutil.which will be able to find it.
        import shutil
        import subprocess
        pythonSlicerExecutablePath = shutil.which('PythonSlicer')
        if not pythonSlicerExecutablePath:
            raise RuntimeError("PythonSlicer executable not found")
    except ImportError:
        # Running from console
        import os
        import sys
        pythonSlicerExecutablePath = os.path.dirname(sys.executable) + "/PythonSlicer"
        if os.name == 'nt':
            pythonSlicerExecutablePath += ".exe"

    commandLine = [pythonSlicerExecutablePath, "-m", module, *args]
    proc = launchConsoleProcess(commandLine, useStartupEnvironment=False)
    logProcessOutput(proc)


def pip_install(requirements):
    """Install python packages.

  Currently, the method simply calls ``python -m pip install`` but in the future further checks, optimizations,
  user confirmation may be implemented, therefore it is recommended to use this method call instead of a plain
  pip install.
  :param requirements: requirement specifier, same format as used by pip (https://docs.python.org/3/installing/index.html)

  Example: calling from Slicer GUI

  .. code-block:: python

    pip_install("tensorflow keras scikit-learn ipywidgets")

  Example: calling from PythonSlicer console

  .. code-block:: python

    from slicer.util import pip_install
    pip_install("tensorflow")

  """
    args = 'install', *requirements.split()
    _executePythonModule('pip', args)


def pip_uninstall(requirements):
    """Uninstall python packages.

  Currently, the method simply calls ``python -m pip uninstall`` but in the future further checks, optimizations,
  user confirmation may be implemented, therefore it is recommended to use this method call instead of a plain
  pip uninstall.

  :param requirements: requirement specifier, same format as used by pip (https://docs.python.org/3/installing/index.html)

  Example: calling from Slicer GUI

  .. code-block:: python

    pip_uninstall("tensorflow keras scikit-learn ipywidgets")

  Example: calling from PythonSlicer console

  .. code-block:: python

    from slicer.util import pip_uninstall
    pip_uninstall("tensorflow")

  """
    args = 'uninstall', *requirements.split(), '--yes'
    _executePythonModule('pip', args)


def _get_logger(filename, level='info'):
    import logging
    level_relations = {
        'debug': logging.DEBUG,
        'info': logging.INFO,
        'warning': logging.WARNING,
        'error': logging.ERROR,
        'crit': logging.CRITICAL
    }
    from logging import handlers
    import sys
    # 创建日志对象
    log = logging.getLogger(filename)
    # 设置日志级别
    log.setLevel(level_relations.get(level))
    # 日志输出格式
    fmt = logging.Formatter('%(asctime)s %(thread)d %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s')
    # 输出到控制台
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(fmt)
    # 输出到文件
    # 日志文件按天进行保存，每天一个日志文件
    file_handler = handlers.TimedRotatingFileHandler(filename=filename, when='D', backupCount=1, encoding='utf-8')
    # 按照大小自动分割日志文件，一旦达到指定的大小重新生成文件
    # file_handler = handlers.RotatingFileHandler(filename=filename, maxBytes=1*1024*1024*1024, backupCount=1, encoding='utf-8')
    file_handler.setFormatter(fmt)

    log.addHandler(console_handler)
    log.addHandler(file_handler)
    return log


def info_logger():
    import logging as log
    import slicer
    from slicer import app
    if hasattr(slicer, 'info'):
        logger = getattr(slicer, 'info')
    else:
        # logger = _get_logger(app.currentLogFilePath())
        logger = _get_logger("E:/123.log")
        setattr(slicer, 'info', logger)
    return logger


def error_logger():
    import logging as log
    import slicer
    from slicer import app
    if hasattr(slicer, 'error'):
        logger = getattr(slicer, 'error')
    else:
        # logger = _get_logger(app.currentLogFilePath())
        logger = _get_logger("E:/123.log")
        setattr(slicer, 'error', logger)
    return logger


# def logger():
#     import logging as log
#     import slicer
#     from slicer import app
#     if hasattr(slicer, 'logger'):
#         logger = getattr(slicer, 'logger')
#     else:
#         logger = _get_logger(app.currentLogFilePath())
#         # logger = _get_logger("c:/med3d.log")
#         setattr(slicer, 'logger', logger)
#     return logger


def longPath(path):
    """Make long paths work on Windows, where the maximum path length is 260 characters.

  For example, the files in the DICOM database may have paths longer than this limit.
  Accessing these can be made safe by prefixing it with the UNC prefix ('\\?\').

  :param string path: Path to be made safe if too long

  :return string: Safe path
  """
    # Return path as is if conversion is disabled
    longPathConversionEnabled = settingsValue('General/LongPathConversionEnabled', True, converter=toBool)
    if not longPathConversionEnabled:
        return path
    # Return path as is on operating systems other than Windows
    import qt
    sysInfo = qt.QSysInfo()
    if sysInfo.productType() != 'windows':
        return path
    # Skip prefixing relative paths as UNC prefix wors only on absolute paths
    if not qt.QDir.isAbsolutePath(path):
        return path
    # Return path as is if UNC prefix is already applied
    if path[:4] == '\\\\?\\':
        return path
    return u"\\\\?\\" + path.replace('/', '\\')


def load_extra_module_by_config(filepath):
    settings = mainWindow().GetProjectSettings()
    p_module_need_to_load = settings.value("JModules/p_module_need_to_load")
    c_module_need_to_load = settings.value("JModules/c_module_need_to_load")
    p_module_release = settings.value("JModules/p_module_release")
    print("--------------------------------------------------------------------", c_module_need_to_load)
    add_search_paths(p_module_need_to_load, c_module_need_to_load, p_module_release, filepath)

    p_module_need_to_load = settings.value("DModules/p_module_need_to_load")
    c_module_need_to_load = settings.value("DModules/c_module_need_to_load")
    add_d_search_paths(p_module_need_to_load, c_module_need_to_load, filepath)


def savejson2(project_name, key1, key2, value):
    import json, os
    base_project_path = mainWindow().GetProjectBasePath()
    project_json = os.path.join(base_project_path, "ProjectCache", project_name, "sub_project.json")
    # util.send_event_str(util.ProgressValue,"30")
    # 如果已经存在的json文件
    import time
    if os.path.exists(project_json):
        with open(project_json, encoding='utf-8') as json_file:
            JsonData = json.load(json_file)
    JsonData[key1][key2] = value
    with open(project_json, "w", encoding="utf-8") as json_file:
        json.dump(JsonData, json_file, ensure_ascii=False, indent=4)


def project_enter():
    # print("project enter")
    show_gif_progress()
    # 判断是否需要下载,如果需要下载,先下载项目
    import time, os, json
    project_name = get_from_PAAA("current_project_selector_project_name")
    base_project_path = mainWindow().GetProjectBasePath()
    project_json = os.path.join(base_project_path, "ProjectCache", project_name, "sub_project.json")
    # util.send_event_str(util.ProgressValue,"30")
    # 如果已经存在的json文件
    import time
    if os.path.exists(project_json):
        with open(project_json, encoding='utf-8') as json_file:
            # JsonData = json.load(json_file)
            # util.send_event_str(util.ProgressValue,"40")
            if not get_from_PAAA("init_all_at_begin") == "2":
                send_event_str(InitPageList, "1")

    send_event_str(GotoNextPage)
    singleShot(0, lambda: hide_gif_progress())

    update_controller_settings()


def update_controller_settings():
    import slicer, os, qt
    use_left_mouse_mode_1 = (getjson2("controller", "use_left_mouse_mode_1") == "2")
    show_slice_pinButton = (getjson2("controller", "show_slice_pinButton") == "2")
    show_slice_viewLabel = (getjson2("controller", "show_slice_viewLabel") == "2")
    show_slice_MaximizeViewButton = (getjson2("controller", "show_slice_MaximizeViewButton") == "2")
    show_slice_fitToWindowToolButton = (getjson2("controller", "show_slice_fitToWindowToolButton") == "2")
    show_sliceOffsetSlider = (getjson2("controller", "show_sliceOffsetSlider") == "2")
    show_sliceOffsetSlider_SpinBox = (getjson2("controller", "show_sliceOffsetSlider_SpinBox") == "2")
    show_3d_pinButton = (getjson2("controller", "show_3d_pinButton") == "2")
    show_3d_viewLabel = (getjson2("controller", "show_3d_viewLabel") == "2")
    show_3d_MaximizeViewButton = (getjson2("controller", "show_3d_MaximizeViewButton") == "2")
    show_3d_fitToWindowToolButton = (getjson2("controller", "show_3d_fitToWindowToolButton") == "2")
    show_sliceVisibility = (getjson2("controller", "show_sliceVisibility") == "2")
    for viewName in slicer.app.layoutManager().sliceViewNames():
        widget = slicer.app.layoutManager().sliceWidget(viewName).sliceView().displayableManagerByClassName(
            "vtkMRMLCrosshairDisplayableManager").GetSliceIntersectionWidget()
        if use_left_mouse_mode_1:
            widget.SetEventTranslationClickAndDrag(widget.WidgetStateIdle, vtk.vtkCommand.LeftButtonPressEvent,
                                                   vtk.vtkEvent.NoModifier, widget.WidgetStateTranslate,
                                                   widget.WidgetEventTranslateStart, widget.WidgetEventTranslateEnd)
        controller = slicer.app.layoutManager().sliceWidget(viewName).sliceController()

        if show_sliceOffsetSlider_SpinBox:
            controller.sliceOffsetSlider().setSpinBoxVisible(True)
        else:
            controller.sliceOffsetSlider().setSpinBoxVisible(False)

        layout = controller.barLayout()
        barWidget = controller.barWidget()

        if not show_slice_pinButton:
            controller.pinButton().setVisible(False)
            layout.removeWidget(controller.pinButton())
        if not show_slice_viewLabel:
            layout.removeWidget(controller.viewLabel())
        if not show_slice_MaximizeViewButton:
            MaximizeViewButton = findChild(barWidget, name="MaximizeViewButton")
            layout.removeWidget(MaximizeViewButton)
        if not show_slice_fitToWindowToolButton:
            layout.removeWidget(controller.fitToWindowToolButton())
        if not show_sliceOffsetSlider:
            layout.removeWidget(controller.sliceOffsetSlider())

        # 如果只有一个slider，那么必须要加这一句，不然会出现显示问题
        if layout.count() == 2 and show_sliceOffsetSlider:
            item_to_remove = layout.itemAt(0)
            layout.removeItem(item_to_remove)

        if show_sliceVisibility:
            def inner_slice_visible(c, v):
                islinked = c.isLinked()
                c.setSliceLink(False)
                c.setSliceVisible(v)
                c.setSliceLink(islinked)

            layout.setSpacing(0)
            btnScissors_stylesheet = ""
            btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton{background-color: #ffff00}"
            btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:hover{background-color: #3377FF;}"
            btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:pressed{background-color: #ffff00;}"
            btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton:checked{background-color: #3377FF;};"
            btnScissors_stylesheet = btnScissors_stylesheet + "QPushButton { border-radius: 0px; }"
            btn = qt.QPushButton()
            btn.setStyleSheet(btnScissors_stylesheet)
            btn.setFixedSize(qt.QSize(15, 15))
            btn.setCheckable(True)
            layout.insertWidget(0, btn)
            btn.connect('toggled(bool)', lambda v, c=controller: inner_slice_visible(c, v))

    for threeDViewIndex in range(slicer.app.layoutManager().threeDViewCount):
        controller = slicer.app.layoutManager().threeDWidget(threeDViewIndex).threeDController()
        layout = controller.barLayout()
        barWidget = controller.barWidget()

        if not show_3d_pinButton:
            controller.pinButton().setVisible(False)
            layout.removeWidget(controller.pinButton())
        if not show_3d_viewLabel:
            layout.removeWidget(controller.viewLabel())
        if not show_3d_MaximizeViewButton:
            MaximizeViewButton = findChild(barWidget, name="MaximizeViewButton")
            layout.removeWidget(MaximizeViewButton)
        if not show_3d_fitToWindowToolButton:
            fitToWindowToolButton = findChild(barWidget, name="CenterButton_Header")
            layout.removeWidget(fitToWindowToolButton)


def save_to_PAAA(key, value):
    import qt, os
    key = "PAAA/" + key
    path = os.path.join(get_project_base_path(), "ProjectCache", "settings.ini").replace("\\", "/")
    config_file = qt.QSettings(path, qt.QSettings.IniFormat)
    config_file.setIniCodec("UTF-8")
    config_file.setValue(key, value)


def get_from_PAAA(key, default=None):
    import qt, os
    key = "PAAA/" + key
    path = os.path.join(get_project_base_path(), "ProjectCache", "settings.ini").replace("\\", "/")
    config_file = qt.QSettings(path, qt.QSettings.IniFormat)
    if config_file.contains(key):
        return config_file.value(key)
    return default


def get_project_base_path():
    return mainWindow().GetProjectBasePath()


def getcustomjson(json, key, default_value=""):
    if not json:
        return default_value
    if key not in json:
        return default_value
    return json[key]


def getcustomjson2(json, key, key2, default_value=""):
    if not json:
        return default_value
    if key not in json:
        return default_value
    if key2 not in json[key][key2]:
        return default_value
    return json[key][key2]


def getjson(key, default_value=""):
    if not JsonData:
        return default_value
    if key not in JsonData:
        return default_value
    return JsonData[key]


def getjson2(key1, key2, default_value=""):
    if not JsonData:
        return default_value
    if key1 not in JsonData:
        return default_value
    pad = JsonData[key1]
    if key2 not in pad:
        return default_value
    return pad[key2]


def processEvents():
    import slicer
    slicer.app.processEvents()


def get_world_position(node):
    target_point_world = [0, 0, 0]
    node.GetNthControlPointPositionWorld(0, target_point_world)
    return target_point_world


def update_setting_from_config(sett):
    import slicer, qt, os
    # from DICOMLib import DICOMUtils
    slicer.util.QSS_QWidget_MV_color = settingsValue("General/QSS_QWidget_MV_color", "#000000", settings=sett)
    slicer.util.QSS_QWidget_background_color = settingsValue("General/QSS_QWidget_background_color", "#000000",
                                                             settings=sett)
    is_axis_visible = settingsValue("General/show_axis", 0, settings=sett) == "2"
    is_box_visible = settingsValue("General/show_box", 0, settings=sett) == "2"
    is_link_control = settingsValue("General/link_control", 0, settings=sett) == "2"
    show_controller = settingsValue("General/show_controller", 0, settings=sett) == "2"
    slice_view_background_color = settingsValue("General/slice_view_background_color", "0|0|0", settings=sett)
    sl = slice_view_background_color.split("|")
    threeD_view_background_color = settingsValue("General/threeD_view_background_color", "0|0|0", settings=sett)
    tl = threeD_view_background_color.split("|")
    threeD_view_background_color2 = settingsValue("General/threeD_view_background_color2", "0|0|0", settings=sett)
    tl2 = threeD_view_background_color2.split("|")
    sliceCompositeNodes = slicer.util.getNodesByClass("vtkMRMLSliceCompositeNode")
    for sliceCompositeNode in sliceCompositeNodes:
        sliceCompositeNode.SetLinkedControl(is_link_control)
    viewNode = slicer.app.layoutManager().threeDWidget(0).mrmlViewNode()
    viewNode.SetAxisLabelsVisible(is_axis_visible)
    viewNode.SetBoxVisible(is_box_visible)

    layoutManager = slicer.app.layoutManager()
    for sliceViewName in layoutManager.sliceViewNames():
        view = layoutManager.sliceWidget(sliceViewName).sliceView()
        view.setBackgroundColor(qt.QColor.fromRgbF(int(sl[0]) / 255.0, int(sl[1]) / 255.0, int(sl[2]) / 255.0))
        view.forceRender()

    settings = qt.QSettings()
    color1_str = settings.value("General/bg_3d_color1")
    color2_str = settings.value("General/bg_3d_color2")
    if color1_str == "" or color1_str is None:
        color1_str = "0,0,0"
    if color2_str == "" or color2_str is None:
        color2_str = "0,0,0"
    color1list = color1_str.split(",")
    color2list = color2_str.split(",")
    viewNode = slicer.app.layoutManager().threeDWidget(0).mrmlViewNode()
    viewNode.SetBackgroundColor(float(tl[0]), float(tl[1]), float(tl[2]))
    viewNode.SetBackgroundColor2(float(tl2[0]), float(tl2[1]), float(tl2[2]))
    setViewControllersVisible(show_controller)

    slicer.util.show_file_dialog_flag = settingsValue("General/qfiledialog_fixed", 0, settings=sett) != "2"
    unique_color_controller = settingsValue("General/unique_color_controller", 0, settings=sett) == "2"
    controller_direction = settingsValue("General/controller_direction", 0, settings=sett) == "2"
    if controller_direction:
        set_controller_vertical_layout()
    if not unique_color_controller:
        nodes = getNodesByClass("vtkMRMLAbstractViewNode")
        for node in nodes:
            node.SetLayoutColor([0, 0, 0])
    is_auto_test = settingsValue("Test/auto_test", 0, settings=sett) == "2"
    if is_auto_test:
        # modulenames = sett.value("Test/to_be_test")
        # if modulenames == "" or modulenames is None:
        #   qt.QTimer.singleShot(100, lambda:showWarningText("自动化测试结束"))
        #   return
        # if isinstance(modulenames, tuple):
        #   modulenames = list(modulenames)
        # else:
        #   modulenames = [modulenames]

        # for modulename_pack in modulenames:
        #   modulename_li = modulename_pack.split('-')
        #   modulename = modulename_li[0]
        #   moduleversion = modulename_li[1]
        #   widget = getModuleWidget(modulename)
        #   scripts_folders = widget.testPath('scripts').replace('\\','/')
        #   script_file = scripts_folders+"/ModuleSelf.py"
        #   test_row_info['test_module'] = modulename
        #   test_row_info['test_version'] = moduleversion
        #   import sys
        #   sys.path.append(scripts_folders)
        #   import ModuleSelf
        #   slicer.mrmlScene.AddObserver(TestEvent, ModuleSelf.OnTestEvent)
        #   sys.path.remove(scripts_folders)
        #   break
        # qt.QTimer.singleShot(100, lambda:send_test_event(0,0))
        record_mode = settingsValue("Test/record_mode", 0, settings=sett) == "2"
        record_module = settingsValue("Test/record_module", "", settings=sett)
        if record_mode:
            if record_module != "":
                modulename_li = record_module.split('-')
                modulename = modulename_li[0]
                moduleversion = modulename_li[1]
                widget = getModuleWidget(modulename)
                if widget is None:
                    showWarningText("模块没有加载:", record_module)
                    return
                if widget.test is None:
                    showWarningText("模块没有测试模块:", record_module)
                    return
                qt.QTimer.singleShot(100, lambda: widget.test.record())
            return

        playback_mode = settingsValue("Test/playback_mode", 0, settings=sett) == "2"
        playback_module = settingsValue("Test/playback_module", "", settings=sett)
        if playback_mode:
            if playback_module != "":
                modulename_li = playback_module.split('-')
                modulename = modulename_li[0]
                moduleversion = modulename_li[1]
                widget = getModuleWidget(modulename)
                if widget is None:
                    showWarningText("模块没有加载:", playback_module)
                    return
                if widget.test is None:
                    showWarningText("模块没有测试模块:", playback_module)
                    return
                qt.QTimer.singleShot(100, lambda: widget.test.playback())
            return


def show_screen_label(str1):
    import qt, slicer
    mainWindow().ShowGlobalPopup(str1, 2000)


def save_list_to_ini(sett, key, val):
    if val == "" or val is None:
        sett.setValue(key, None)
    elif len(val) == 1:
        sett.setValue(key, val[0])
    elif isinstance(val, list):
        sett.setValue(key, tuple(val))
    else:
        sett.setValue(key, val)


def get_list_from_init(sett, key):
    modulenames = settingsValue(key, "", settings=sett)
    if modulenames == "":
        return []
    if modulenames == None:
        return []
    elif isinstance(modulenames, tuple):
        modulenames = list(modulenames)
    else:
        modulenames = [modulenames]
    return modulenames


def test_case_finished(file_path):
    module = test_row_info['test_module']
    version = test_row_info['test_version']
    modulewithversion = module + "-" + version
    sett = mainWindow().GetProjectSettings()
    is_persist_test = settingsValue("Test/persist_test", 0, settings=sett) == "2"
    if not is_persist_test:
        modulenames = get_list_from_init(sett, 'Test/to_be_test')
        modulenames.remove(modulewithversion)
        save_list_to_ini(sett, 'Test/to_be_test', modulenames)

        modulenames = get_list_from_init(sett, 'Test/has_tested')
        modulenames.append(modulewithversion)
        save_list_to_ini(sett, 'Test/has_tested', modulenames)
        return True

    # 将报告复制到项目的文件报告中
    folder_path = get_project_base_path() + "/Test/reports"
    import os
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    file_folder_path, file_name = os.path.split(file_path)
    import shutil
    dst_path = folder_path + "/" + file_name
    print("copy test case from " + file_path + " to " + dst_path)
    shutil.move(file_path, dst_path)
    return False


'''
  如果用激活码激活的话,激活WindowTitle的激活状态
'''


def update_activate_state():
    verify = deshacoc2()
    if verify:
        mainWindow().setWindowTitle(mainWindow().GetProjectSettings().value("General/window_title") + ("(已激活)"))
    else:
        mainWindow().setWindowTitle(mainWindow().GetProjectSettings().value("General/window_title") + ("(未激活)"))


def deshacoc(str1, str2):
    str1 = str1 + "olkjhgfakjhwefghjeewr"
    import hashlib
    loc = hashlib.sha256(str1.encode()).hexdigest()
    loc = loc[-6:]
    if loc == str2:
        print("activate ok")
        return True
    print("activate wrong")
    return False


def deshacoc3():
    if mainWindow().check_verification():
        print("already activate")
        return True


def deshacoc2():
    import slicer, os
    if mainWindow().check_verification():
        print("already activate")
        return True
    key_path = slicer.app.slicerHome + "/key.txt"
    lock_path = slicer.app.slicerHome + "/lock.txt"
    key_content = ""
    lock_content = ""
    if os.path.exists(key_path):
        with open(key_path, 'r') as f:
            key_content = f.read()
    else:
        print("miss key file")
        return False

    if os.path.exists(lock_path):
        with open(lock_path, 'r') as f:
            lock_content = f.read()
    else:
        print("miss lock file")
        return False
    return deshacoc(key_content, lock_content)


'''
  更新所有传入的模块的Reload按钮的可见性
  如果General/module_reload==2,那么可见
'''


def update_module_reload_visibility(all_jmodule_name, settings1):
    btnr = settingsValue("General/module_reload", 0, settings=settings1) == "2"
    for name in all_jmodule_name:
        module_logic = getModuleLogic(name)
        if module_logic:
            module_logic.hide_reload(btnr)


def pyautogui_clickwidget(widget, interval=TestInterval):
    import qt, pyautogui, time, slicer
    buttonPos = widget.mapToGlobal(qt.QPoint(widget.width / 2, widget.height / 2))
    pyautogui.moveTo(buttonPos.x(), buttonPos.y())
    slicer.app.processEvents()
    time.sleep(interval)
    pyautogui.click()
    slicer.app.processEvents()


def test_file_save(workbook, filepath):
    try:
        import openpyxl
    except Exception as e:
        pip_install('openpyxl')
        import openpyxl
    # 遍历所有单元格并设置它们的文本居中
    sheet = workbook.active
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical='center')
    # 保存工作簿
    # 遍历所有单元格并设置它们的文本居中
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
    # 创建样式
    wrap_text_style = openpyxl.styles.alignment.Alignment(wrap_text=True, horizontal="center", vertical='center')

    # 遍历每个单元格，并设置样式
    for row in sheet.rows:
        for cell in row:
            cell.alignment = wrap_text_style
    workbook.save(filepath)


def test_take_screenshot(folder_path, width=1920, height=1040, rwidth=0.3, rheight=0.3, filename="1.png"):
    try:
        import pyautogui, time
    except Exception as e:
        pip_install('pyautogui')
        import pyautogui

    if folder_path is not None:
        import os
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

    from PIL import Image
    # 获取屏幕截图
    screenshot = pyautogui.screenshot(region=(0, 0, width, height))
    # 调整图像大小
    screenshot = screenshot.resize((int(width * rwidth), int(height * rheight)), Image.ANTIALIAS)
    import time
    timestamp_str = int(time.time()).__str__()
    screenshot.save(folder_path + "/" + timestamp_str + filename)
    return folder_path + "/" + timestamp_str + filename


def test_add_row(sheet, test_row_info):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import column_index_from_string
    # 获取当前时间的datetime对象
    import datetime
    now = datetime.datetime.now()
    # 将datetime对象格式化为指定的字符串
    timestamp_str = now.strftime('%Y-%m-%d')

    img_path = test_row_info['image_path']
    img = openpyxl.drawing.image.Image(img_path)
    row = []
    row.append(test_row_info['test_serial_number'])  # 序号为当前最后一行序号+1
    row.append(f"JSS-000{test_row_info['test_serial_number']}")  # 用例编号为"TC00"+序号
    row.append(test_row_info['test_type'])  # 随机选择测试类型
    row.append(test_row_info['test_method'])  # 随机选择测试方式
    row.append(test_row_info['test_description'])  # 测试用例描述
    row.append(test_row_info['test_input'])  # 测试输入数据
    row.append(test_row_info['test_steps'])  # 测试执行步骤
    row.append(test_row_info['test_expects'])  # 测试预期结果
    row.append(' ')  # 测试记录和说明
    row.append(' ')  # 随机选择判定
    row.append(" ")  # 测试人
    row.append(timestamp_str)  # 测试日期
    row.append("无")  # 测试日期

    for j in range(len(row)):
        column = get_column_letter(j + 1)
        sheet[f"{column}{test_row_info['test_serial_number'] + 1}"] = row[j]
    sheet.add_image(img, 'I' + (test_row_info['test_serial_number'] + 1).__str__())
    sheet.row_dimensions[(test_row_info['test_serial_number'] + 1)].height = img.height * 0.75
    sheet.column_dimensions["I"].width = img.width / 7.0 * 0.87


def test_file_create(folder_path, filename):
    # 保存工作簿
    import os
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    import datetime
    # 获取当前时间的datetime对象
    now = datetime.datetime.now()
    # 将datetime对象格式化为指定的字符串
    timestamp_str = now.strftime('%Y-%m-%d-%H-%M-%S')
    print(timestamp_str)
    filepath = folder_path + "/" + timestamp_str + "--" + filename

    try:
        import openpyxl
    except Exception as e:
        pip_install('openpyxl')
        import openpyxl
    from openpyxl.styles import Font, PatternFill
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()

    # 选择默认的工作表
    ws = wb.active

    # 添加标题行
    t0 = test_row_info['test_module'] + '(' + test_row_info['test_version'] + ')' + '\n' + '序号'
    ws.append([t0, '用例编号', '测试类型', '测试方式', '测试用例描述', '测试输入数据', '测试执行步骤', '测试预期结果',
               '测试记录', '判定', '测试人', '测试日期', '备注'])

    # 设置标题行的样式
    font = Font(bold=False, name='宋体', size=12)
    fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    for cell in ws[1]:
        cell.font = font
        cell.fill = fill
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    sheet = wb.active

    # 遍历所有单元格并设置它们的文本居中
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    # 将每一列的自动换行设置为True
    for col in ws.columns:
        col[0].alignment = openpyxl.styles.Alignment(wrap_text=True)

    # 设置每个单元格自适应列宽
    for col in ws.columns:
        column_letter = col[0].column_letter
        ws.column_dimensions[column_letter].auto_size = True

    # 格式化每一列
    for col in ws.columns:
        for cell in col:
            # 设置文字居中
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    wb.save(filepath)
    return filepath


def getModuleVersion(module_name):
    import os, qt
    widget = getModuleWidget(module_name)
    module_conifg_path = widget.resourcePath("config.ini").replace('\\', '/')
    setting_obj = qt.QSettings(module_conifg_path, qt.QSettings.IniFormat)
    version_str = settingsValue("General/version", "1.0.0", settings=setting_obj)
    return version_str


# 生成模块里所有的详细注释
def generate_comment_docs(comment_classes, comment_path):
    for comment_class in comment_classes:
        class_obj = eval(comment_class)
        _generate_comment_doc(comment_class, class_obj, comment_path)


def _generate_comment_doc(comment_class_name, class_obj, comment_path):
    import inspect, time, os
    members = inspect.getmembers(class_obj, predicate=inspect.isfunction)

    # 只保留不是以'__'和'_'开头的函数名称
    custom_functions = []
    for function_name, function_obj in members:
        if not function_name.startswith('__') and not function_name.startswith(
                '_') and function_obj.__qualname__.startswith(class_obj.__name__):
            custom_functions.append((function_name, function_obj))

    # 通过函数名获取到函数体
    function_name_map = {}
    # 打印自定义函数名
    for function_name, _ in custom_functions:
        # 通过类名和函数名获取函数对象
        function_obj = getattr(class_obj, function_name)
        # 获取函数的完整代码
        function_code = inspect.getsource(function_obj)
        function_name_map[function_name] = function_code

    keys = (function_name_map.keys())
    gpt = getModuleWidget("JChatGPT")
    response = {}
    for key in function_name_map:
        source_code = function_name_map[key]
        question_list = get_comment_prompt(keys, source_code)
        try:
            res = gpt.ask_question_detail(question_list, temperature=0)
            response[key] = res
        except Exception as e:
            response[key] = e.__str__()

        import json, pickle
        cache_folder = comment_path
        if not os.path.exists(cache_folder):
            os.makedirs(cache_folder)
        cache_file = comment_path + "/cmt_%s.txt" % (comment_class_name).replace('\\', '/')
        with open(cache_file, "a", encoding="utf-8") as file:
            for key in response:
                file.write("\n函数名:\n" + key)
                file.write("\n\n注释:\n" + response[key].replace('\n\n', '\n'))
                file.write("\n-----------------------------------------------------------------------------\n\n\n\n")
        response = {}
        time.sleep(18)


def remove_directory(folder_path):
    import shutil, os
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)


def make_directory(path):
    import os
    if not os.path.exists(path):
        os.makedirs(path)


def make_parent_directory(path):
    import os
    folder_path = os.path.dirname(path)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)


def get_comment_prompt(keys, source_code):
    question_list = []
    question_list.append({"role": "system", "content": "You are a senior software engineer"})
    question_list.append({"role": "system", "content": "You are preparing to write a detailed design document"})
    question_list.append({"role": "user", "content": "All functions in this module are as follows '%s'" % (keys)})
    question_list.append({"role": "user",
                          "content": "Please use chinese to provide a detailed comment above the function, explaining its purpose(using the format '@detail  :purpose description'), the parameters(using the format '@para : description'), and the return value(using the format '@return : description'). '%s'" % (
                              source_code)})
    return question_list


def do_load_jmodule(module_name_pack, scriptedDirPath, moduleFactoryManager, is_c=False, is_d=False):
    import os, slicer, qt
    # 1.get module name and module version
    module_names = module_name_pack.split("-")
    if len(module_names) == 1:
        module_name = module_names[0]
        module_version = "1.0.0"
    else:
        module_name = module_names[0]
        module_version = module_names[1]

    # 2.check module version
    # TODO

    # 3.add module to search path
    if is_c:
        if is_d:
            module_path = (
                os.path.join(scriptedDirPath, 'DExtension', 'CExtension', module_name, 'qt-scripted-modules')).replace(
                '\\', '/')
            moduleFactoryManager.pyAddSearchPath(module_path)
            module_path = (
                os.path.join(scriptedDirPath, 'DExtension', 'CExtension', module_name, 'qt-loadable-modules')).replace(
                '\\', '/')
            moduleFactoryManager.pyAddSearchPath(module_path)
            module_path = (
                os.path.join(scriptedDirPath, 'DExtension', 'CExtension', module_name, 'cli-modules')).replace('\\',
                                                                                                               '/')
            moduleFactoryManager.pyAddSearchPath(module_path)
            module_conifg_path = (os.path.join(scriptedDirPath, 'DExtension', 'CExtension', module_name, "Resources",
                                               "config.ini")).replace('\\', '/')
        else:
            module_path = (os.path.join(scriptedDirPath, 'CExtension', module_name, 'qt-scripted-modules')).replace(
                '\\', '/')
            moduleFactoryManager.pyAddSearchPath(module_path)
            module_path = (os.path.join(scriptedDirPath, 'CExtension', module_name, 'qt-loadable-modules')).replace(
                '\\', '/')
            moduleFactoryManager.pyAddSearchPath(module_path)
            module_path = (os.path.join(scriptedDirPath, 'CExtension', module_name, 'cli-modules')).replace('\\', '/')
            moduleFactoryManager.pyAddSearchPath(module_path)
            module_conifg_path = (
                os.path.join(scriptedDirPath, 'CExtension', module_name, "Resources", "config.ini")).replace('\\', '/')
    else:
        if is_d:
            module_path = (os.path.join(scriptedDirPath, 'DExtension', 'JExtension', module_name)).replace('\\', '/')
        else:
            module_path = (os.path.join(scriptedDirPath, 'JExtension', module_name)).replace('\\', '/')
        moduleFactoryManager.pyAddSearchPath(module_path)
        module_conifg_path = (os.path.join(module_path, "Resources", "config.ini")).replace('\\', '/')

    loaded_modules[module_name] = 1

    # 4.module config
    module_conifg_path = (os.path.join(module_path, "Resources", "config.ini")).replace('\\', '/')
    if os.path.exists(module_conifg_path):
        setting_obj = qt.QSettings(module_conifg_path, qt.QSettings.IniFormat)
        version_str = settingsValue("General/version", "1.0.0", settings=setting_obj)

        # load jmodule
        module_depends_str = settingsValue("JModules/p_module_need_to_load", "", settings=setting_obj)
        if isinstance(module_depends_str, str):
            module_depends_str = [module_depends_str]
        print("module_depends_str:", module_depends_str, type(module_depends_str))
        if module_depends_str != "":
            for module_depend in module_depends_str:
                if module_depend in loaded_modules:
                    continue
                else:
                    do_load_jmodule(module_depend, scriptedDirPath, moduleFactoryManager)

        # load cmodule
        module_depends_str = settingsValue("JModules/c_module_need_to_load", "", settings=setting_obj)
        if isinstance(module_depends_str, str):
            module_depends_str = [module_depends_str]
        # print("module_depends_str:",module_depends_str,type(module_depends_str))
        if module_depends_str != "":
            for module_depend in module_depends_str:
                if module_depend in loaded_modules:
                    continue
                else:
                    do_load_jmodule(module_depend, scriptedDirPath, moduleFactoryManager, is_c=True)

        # load djmodule
        module_depends_str = settingsValue("DModules/p_module_need_to_load", "", settings=setting_obj)
        if isinstance(module_depends_str, str):
            module_depends_str = [module_depends_str]
        # print("module_depends_str:",module_depends_str,type(module_depends_str))
        if module_depends_str != "":
            for module_depend in module_depends_str:
                if module_depend in loaded_modules:
                    continue
                else:
                    do_load_jmodule(module_depend, scriptedDirPath, moduleFactoryManager, is_d=True)

        # load dcmodule
        module_depends_str = settingsValue("DModules/c_module_need_to_load", "", settings=setting_obj)
        if isinstance(module_depends_str, str):
            module_depends_str = [module_depends_str]
        # print("module_depends_str:",module_depends_str,type(module_depends_str))
        if module_depends_str != "":
            for module_depend in module_depends_str:
                if module_depend in loaded_modules:
                    continue
                else:
                    do_load_jmodule(module_depend, scriptedDirPath, moduleFactoryManager, is_d=True, is_c=True)


def add_d_search_paths(module_names, c_module_names, filepath):
    import slicer, os
    moduleManager = slicer.app.moduleManager()
    moduleFactoryManager = moduleManager.factoryManager()
    scriptedDirPath = os.path.abspath(os.path.dirname(filepath))
    scriptedDirPath = os.path.abspath(os.path.join(scriptedDirPath, os.pardir))

    if module_names is not None:
        for module_name_pack in module_names:
            do_load_jmodule(module_name_pack, scriptedDirPath, moduleFactoryManager, is_d=True)

    if c_module_names is not None:
        for module_name_pack in c_module_names:
            module_names = module_name_pack.split("-")
            if len(module_names) == 1:
                module_name = module_names[0]
                module_version = "1.0.0"
            else:
                module_name = module_names[0]
                module_version = module_names[1]
            if len(module_name) > 0:
                module_path = (os.path.join(scriptedDirPath, 'DExtension', 'CExtension', module_name,
                                            'qt-scripted-modules')).replace('\\', '/')
                moduleFactoryManager.pyAddSearchPath(module_path)
                module_path = (os.path.join(scriptedDirPath, 'DExtension', 'CExtension', module_name,
                                            'qt-loadable-modules')).replace('\\', '/')
                moduleFactoryManager.pyAddSearchPath(module_path)
                module_path = (
                    os.path.join(scriptedDirPath, 'DExtension', 'CExtension', module_name, 'cli-modules')).replace('\\',
                                                                                                                   '/')
                moduleFactoryManager.pyAddSearchPath(module_path)

    moduleFactoryManager.registerModules()
    # moduleFactoryManager.instantiateModules()
    moduleFactoryManager.instantiateModules()
    moduleFactoryManager.loadModules()


KEY = b'\x1b\x5e\x30\x27\x77\x34\x6a\x50\x3f\x63\x3f\x2b\x44\x3a\x6c\x52'


def decrypt_file(input_file_path, output_file_path):
    import os, shutil
    with open(input_file_path, 'rb') as input_file:
        ciphertext = input_file.read()

    # 对每个字节进行异或操作
    plaintext = bytes([byte ^ KEY[i % len(KEY)] for i, byte in enumerate(ciphertext)])

    with open(output_file_path, 'wb') as output_file:
        output_file.write(plaintext)


def decrypt_folder_without_remove(folder_path):
    import os, shutil
    for dirpath, dirnames, filenames in os.walk(folder_path):
        for filename in filenames:
            if filename.endswith('.yp'):
                input_file_path = os.path.join(dirpath, filename)
                output_file_path = os.path.join(dirpath, filename[:-3] + '.py')

                # 解密文件
                decrypt_file(input_file_path, output_file_path)


def delete_files_with_extensions(folder_path, extensions):
    print("VVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVV:", folder_path, extensions)
    import os, shutil
    for root, _, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            if any(file_path.endswith(ext) for ext in extensions):
                os.remove(file_path)


def decrypt_PAAA_sub(full_path, GLPyDirPath):
    import os, qt
    module_conifg_path = (os.path.join(full_path, "Resources", "config.ini")).replace('\\', '/')
    if os.path.exists(module_conifg_path):
        setting_obj = qt.QSettings(module_conifg_path, qt.QSettings.IniFormat)
        # load jmodule
        p_module_need_to_load = settingsValue("JModules/p_module_need_to_load", "", settings=setting_obj)
        c_module_need_to_load = settingsValue("JModules/c_module_need_to_load", "", settings=setting_obj)
        if isinstance(p_module_need_to_load, str):
            p_module_need_to_load = [p_module_need_to_load]
        if isinstance(c_module_need_to_load, str):
            c_module_need_to_load = [c_module_need_to_load]
        decrypt_PAAA(GLPyDirPath, p_module_need_to_load, c_module_need_to_load)


def decrypt_PAAA(GLPyDirPath, p_module_need_to_load, c_module_need_to_load):
    print("decrypt_PAAA:", GLPyDirPath)
    if len(p_module_need_to_load) + len(c_module_need_to_load) == 0:
        return
    if not isPackage():
        return
    import os, qt
    do_hash = get_from_PAAA("do_hash") == "2"
    if not do_hash:
        return
    for module_s in p_module_need_to_load:
        if len(module_s) > 0:
            full_path = os.path.join(GLPyDirPath, "JExtension", module_s)
            decrypt_folder_without_remove(full_path)
            decrypt_PAAA_sub(full_path, GLPyDirPath)
    for module_s in c_module_need_to_load:
        if len(module_s) > 0:
            full_path = os.path.join(GLPyDirPath, "CExtension", module_s)
            decrypt_folder_without_remove(full_path)
            decrypt_PAAA_sub(full_path, GLPyDirPath)


def decrypt_remove_PAAA_Sub(full_path, GLPyDirPath):
    import os, qt
    module_conifg_path = (os.path.join(full_path, "Resources", "config.ini")).replace('\\', '/')
    if os.path.exists(module_conifg_path):
        setting_obj = qt.QSettings(module_conifg_path, qt.QSettings.IniFormat)
        # load jmodule
        p_module_need_to_load = settingsValue("JModules/p_module_need_to_load", "", settings=setting_obj)
        c_module_need_to_load = settingsValue("JModules/c_module_need_to_load", "", settings=setting_obj)
        if isinstance(p_module_need_to_load, str):
            p_module_need_to_load = [p_module_need_to_load]
        if isinstance(c_module_need_to_load, str):
            c_module_need_to_load = [c_module_need_to_load]
        decrypt_remove_PAAA(GLPyDirPath, p_module_need_to_load, c_module_need_to_load)


def decrypt_remove_PAAA(GLPyDirPath, p_module_need_to_load, c_module_need_to_load):
    print("decrypt_remove_PAAA:", decrypt_remove_PAAA)
    if len(p_module_need_to_load) + len(c_module_need_to_load) == 0:
        return
    if not isPackage():
        return
    import os
    extensions_to_delete = ['.py', '.pyc']
    whitelist = get_from_PAAA("whitelist", "")
    do_hash = get_from_PAAA("do_hash") == "2"
    if not do_hash:
        return
    for module_s in p_module_need_to_load:
        in_whitelist = False
        if module_s in whitelist:
            in_whitelist = True
        print("UIA:", module_s, whitelist, in_whitelist)
        if len(module_s) > 0:
            full_path = os.path.join(GLPyDirPath, "JExtension", module_s)
        if not in_whitelist:
            delete_files_with_extensions(full_path, extensions_to_delete)
            decrypt_remove_PAAA_Sub(full_path, GLPyDirPath)
    for module_s in c_module_need_to_load:
        in_whitelist = False
        if module_s in whitelist:
            in_whitelist = True
        if len(module_s) > 0:

            full_path = os.path.join(GLPyDirPath, "CExtension", module_s)
            if not in_whitelist:
                delete_files_with_extensions(full_path, extensions_to_delete)
                decrypt_remove_PAAA_Sub(full_path, GLPyDirPath)


def add_search_paths(module_names, c_module_names, r_module_names, filepath):
    import slicer, os, qt
    moduleManager = slicer.app.moduleManager()
    moduleFactoryManager = moduleManager.factoryManager()
    scriptedDirPath = os.path.abspath(os.path.dirname(filepath))
    scriptedDirPath = os.path.abspath(os.path.join(scriptedDirPath, os.pardir))
    scriptedDirPath = os.path.abspath(os.path.join(scriptedDirPath, os.pardir))
    if module_names is not None:
        for module_name_pack in module_names:
            do_load_jmodule(module_name_pack, scriptedDirPath, moduleFactoryManager)

            # do_load_jmodule(module_path,module_name,module_version,scriptedDirPath,moduleFactoryManager)
    if r_module_names is not None:
        if type(r_module_names) is str:
            r_module_names = [r_module_names]
        for module_name in r_module_names:
            if isRelease():
                module_path = (os.path.join(scriptedDirPath, 'RExtension', module_name, 'Release')).replace('\\', '/')
                moduleFactoryManager.pyAddSearchPath(module_path)
                print("Add Search Path R:", module_path)
            else:
                module_path = (os.path.join(scriptedDirPath, 'RExtension', module_name, 'Debug')).replace('\\', '/')
                print("Add Search Path D:", module_path)
                moduleFactoryManager.pyAddSearchPath(module_path)

    if c_module_names is not None:
        if type(c_module_names) is str:
            c_module_names = [c_module_names]
        for module_name_pack in c_module_names:
            module_names = module_name_pack.split("-")
            if len(module_names) == 1:
                module_name = module_names[0]
                module_version = "1.0.0"
            else:
                module_name = module_names[0]
                module_version = module_names[1]

            module_path = (os.path.join(scriptedDirPath, 'CExtension', module_name, 'qt-scripted-modules')).replace(
                '\\', '/')
            moduleFactoryManager.pyAddSearchPath(module_path)
            module_path = (os.path.join(scriptedDirPath, 'CExtension', module_name, 'qt-loadable-modules')).replace(
                '\\', '/')
            moduleFactoryManager.pyAddSearchPath(module_path)
            module_path = (os.path.join(scriptedDirPath, 'CExtension', module_name, 'cli-modules')).replace('\\', '/')
            moduleFactoryManager.pyAddSearchPath(module_path)

    moduleFactoryManager.registerModules()
    # moduleFactoryManager.instantiateModules()
    moduleFactoryManager.instantiateModules()
    moduleFactoryManager.loadModules()


def save_global_value(save_key, save_value):
    global_data_map[save_key] = save_value


def get_global_value(save_key, default_value=-1):
    print(global_data_map)
    value_of_key = global_data_map.get(save_key, default_value)
    return value_of_key


def del_global_value(save_key):
    value_of_key = global_data_map.get(save_key, None)
    if value_of_key3 is not None:
        del global_data_map[save_key]


def shut_down():
    result = getModuleWidget("JMessageBox").show_three_popup('正在准备关机中')
    if result == qt.QDialog.Rejected:
        return
    os.system("shutdown /s /t 0")


def reboot():
    result = getModuleWidget("JMessageBox").show_three_popup('正在准备重启中')
    if result == qt.QDialog.Rejected:
        return
    os.system("shutdown /r /t 0")


def start_move_animation(widget, parent_widget, move_time, pos_start, pos_end, callback_func):
    import qt
    print('start_move_animation')
    animation = qt.QPropertyAnimation(widget, "pos", parent_widget)
    animation.setDuration(move_time)
    animation.setStartValue(pos_start)
    animation.setEndValue(pos_end)
    animation.start()
    animation.connect('finished()', callback_func)


def clear_save_info():
    global_data_map.clear()
    save_screeen_capture_list.clear()


def set_check_state(list):
    check_state_list = list
    num = 0
    global_data_map["check_error"] = ""
    error_item_des = ""
    print("get_error_info", check_state_list)
    for item in check_state_list:
        if item == 2:
            error_item_des = check_des_list[num]
            break
        num = num + 1
    if error_item_des == "":
        return
    global_data_map["check_error"] = f'{error_item_des}单元初始化失败--XXX\n请联系售后维修'


def get_error_info():
    print(global_data_map["check_error"], "get_error_info")
    return global_data_map["check_error"]


def add_to_clipboard(info):
    try:
        import pyperclip
    except Exception as e:
        pip_install('pyperclip')
        import pyperclip
    pyperclip.copy(info)


def get_md5_value(value_list):
    import qt
    tmp_str = ",".join(value_list)
    print(tmp_str)
    input_byte = bytes(tmp_str, 'utf-8')
    hash_object = qt.QCryptographicHash(qt.QCryptographicHash.Md5)
    hash_object.addData(input_byte)
    result = hash_object.result().toHex().data().decode('utf-8')
    return result


def get_unique_id():
    time = datetime.now().strftime('%Y%m%d%H%M%S')
    characters = string.ascii_letters + string.digits
    result = ''.join(random.choice(characters) for _ in range(5))
    # print(result, time)
    return time + result


def convert_float_to_bytes(val):
    import struct
    # Convert float to 4 bytes using IEEE 754 single-precision
    bytes_value = struct.pack('>f', val)
    # Alternatively, convert each byte to hexadecimal for readability
    hex_bytes = [f"0x{b:02x}" for b in bytes_value]
    int_bytes = [b for b in bytes_value]
    return hex_bytes, int_bytes


def convert_int_to_2byte_bytes(val):
    # 将整数转换为2字节大端序字节数据
    bytes_value = struct.pack('>H', val)
    # 将每个字节转换为十六进制字符串，方便阅读
    hex_bytes = [f"0x{b:02x}" for b in bytes_value]
    # 将每个字节转换为整数表示
    int_bytes = [b for b in bytes_value]

    return hex_bytes, int_bytes


def convert_2byte_to_int(hex_bytes):
    # 将每个字符串转换为实际的字节值
    byte_values = bytes(int(byte, 16) for byte in hex_bytes)
    return int.from_bytes(byte_values, byteorder='little')


def convert_bytes_to_float(bytes_value_big_endian):
    import struct
    # Convert list of hex strings to a bytes object
    bytes_value_big_endian = bytes(bytes_value_big_endian)
    # Convert the bytes object back to a float using big-endian format
    float_value = struct.unpack('>f', bytes_value_big_endian)[0]
    # Print the float value
    return float_value


def convert_hex_to_float(hex_string):
    # 将清理后的十六进制字符串转换为字节数据
    byte_data = bytes.fromhex(hex_string)

    # 将字节数据解释为浮点数（大端序）
    float_num = struct.unpack('>f', byte_data)[0]
    return float_num


def convert_hex_to_half_float(hex_string):
    # 将清理后的十六进制字符串转换为字节数据
    byte_data = bytes.fromhex(hex_string)

    # 将字节数据从大端序转换为小端序
    byte_data_little_endian = byte_data[::-1]

    # 使用numpy将字节数据转换为半精度浮点数
    half_float = np.frombuffer(byte_data_little_endian, dtype=np.float16)[0]
    return half_float


def convert_hex_to_half_int(hex_string):
    # 将十六进制字符串转换为字节数据
    byte_data = bytes.fromhex(hex_string)

    # 使用struct模块将字节数据解包为大端序的无符号短整数
    integer_value = struct.unpack('>H', byte_data)[0]
    return integer_value


def extract_states_from_hex_string(hex_string):
    # 将十六进制字符串转换为整数
    hex_value = int(hex_string, 16)

    # 提取每个位的值
    status_1 = (hex_value >> 0) & 0x01  # 取出位0
    status_2 = (hex_value >> 1) & 0x01  # 取出位1
    status_3 = (hex_value >> 2) & 0x01  # 取出位2
    status_4 = (hex_value >> 3) & 0x01  # 取出位3
    status_5 = (hex_value >> 4) & 0x01  # 取出位4
    status_6 = (hex_value >> 5) & 0x01  # 取出位5
    status_7 = (hex_value >> 6) & 0x01  # 取出位6
    status_8 = (hex_value >> 7) & 0x01  # 取出位7

    return status_1, status_2, status_3, status_4, status_5, status_6, status_7, status_8
